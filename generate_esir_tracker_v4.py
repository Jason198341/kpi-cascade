"""
ESIR Master Tracker V4 -- Excel Generator
Automotive Interior Parts (12 items, 18 tests, 6 suppliers)
Hyundai Motor India Engineering Sample / Initial Run Test Preparation

V4 Architecture (28 cols A-AB, 4 reference tables, DB Schema v2):
  Sheet 1: Config (hidden helper)
  Sheet 2: Reference Data (SSOT: Vehicle + TestCatalog + ItemMaster + ItemTests)
  Sheet 3: Master Tracker (28 cols, formula-driven)
  Sheet 4: Dashboard (COUNTIFS + vehicle filter)
  Sheet 5: Supplier Checklist

Reference Data (4 tables):
  - TBL_Vehicle (10 cols, 3 rows): Vehicle schedule + LeadTime offsets
  - TBL_TestCatalog (5 cols, 18 rows): Tests with type/spec/duration/scope
  - TBL_ItemMaster (4 cols, 12 rows): Item→Category+Supplier+DrawingNo
  - TBL_ItemTests (3 cols, ~17 rows): Item-specific test mappings + exclusions
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# == Constants ==
OUTPUT = r"C:\Users\USER\kpi-cascade\ESIR_Master_Tracker_v4.xlsx"
TODAY = datetime.date(2026, 2, 27)
TRACKER_END = 200  # max data row for formulas/validations

# == Explicit cross-sheet range references for VLOOKUPs ==
REF = "'Reference Data 기준데이터'"
RNG_VEHICLE     = f"{REF}!$A$3:$J$5"       # 10 cols, data only (skip header row 2)
RNG_TESTCATALOG = f"{REF}!$A$9:$E$26"      # 18 tests, data only (skip header row 8)
RNG_ITEMMASTER  = f"{REF}!$A$30:$D$41"     # 12 items, data only (skip header row 29)

# == Mapping Data (SSOT for table generation + sim data) ==
VEHICLE_ORDER = ["AE_PE", "SU2i", "EN_SUV"]

# 12 interior items with (PartCategory, Supplier, DrawingNo)
ITEM_DATA = {
    "센터콘솔":     ("Console", "HS R&A",           "DWG-CONSOLE-001"),
    "콘솔리드":     ("Console", "HS R&A",           "DWG-CONSOLE-002"),
    "칵핏모듈":     ("Cockpit", "Hyundai Mobis",    "DWG-COCKPIT-001"),
    "도어트림(앞)": ("Trim",    "Faurecia",         "DWG-DOORTRIM-FR"),
    "도어트림(뒤)": ("Trim",    "Faurecia",         "DWG-DOORTRIM-RR"),
    "필러트림 A":   ("Trim",    "Seoyon",           "DWG-PILLAR-A"),
    "필러트림 B":   ("Trim",    "Seoyon",           "DWG-PILLAR-B"),
    "필러트림 C":   ("Trim",    "Seoyon",           "DWG-PILLAR-C"),
    "패키지트레이": ("Trim",    "HS R&A",           "DWG-PKGTRAY-001"),
    "헤드라이너":   ("Trim",    "Seoyon E-Hwa",     "DWG-HEADLINER-001"),
    "시트(앞)":     ("Seat",    "Hyundai Transys",  "DWG-SEAT-FR"),
    "시트(뒤)":     ("Seat",    "Hyundai Transys",  "DWG-SEAT-RR"),
}
ITEM_NAMES = list(ITEM_DATA.keys())

T0_DATES = {
    "AE_PE":  datetime.date(2026, 1, 15),
    "SU2i":   datetime.date(2026, 3, 1),
    "EN_SUV": datetime.date(2026, 2, 1),
}

# 18 tests: TestNameKR -> (TestType, Spec, DurationDays, Scope)
TEST_CATALOG = {
    # 9 COMMON
    "EMC 시험":       ("단품", "ES 95400-05", 14, "COMMON"),
    "VOC 방출 시험":  ("단품", "MS-300-57",   14, "COMMON"),
    "연소성 시험":    ("단품", "FMVSS 302",    3, "COMMON"),
    "포깅 시험":      ("단품", "ES 80062",     7, "COMMON"),
    "냄새 시험":      ("단품", "MS-300-55",    5, "COMMON"),
    "내후성 시험":    ("단품", "MS-300-34",   21, "COMMON"),
    "내열성 시험":    ("단품", "MS-300-31",   14, "COMMON"),
    "내마모성 시험":  ("단품", "MS-300-35",    7, "COMMON"),
    "내변색 시험":    ("단품", "MS-300-36",   10, "COMMON"),
    # 9 ITEM_SPECIFIC
    "힌지 내구 시험":     ("단품",     "ES-92030",  10, "ITEM_SPECIFIC"),
    "시트 내구 시험":     ("단품",     "ES-93100",  21, "ITEM_SPECIFIC"),
    "H-Point 검증":       ("차량장착", "ES-93000",  14, "ITEM_SPECIFIC"),
    "안전벨트 앵커리지":  ("차량장착", "FMVSS 210", 14, "ITEM_SPECIFIC"),
    "ISOFIX 강도 시험":   ("차량장착", "ECE R14",   10, "ITEM_SPECIFIC"),
    "BSR 시험":           ("차량장착", "ES-90500",  14, "ITEM_SPECIFIC"),
    "치수 검사":          ("단품",     "ES-90101",   3, "ITEM_SPECIFIC"),
    "수하물 충격 시험":   ("차량장착", "ECE R17",    7, "ITEM_SPECIFIC"),
    "조작력 시험":        ("단품",     "ES-91300",   5, "ITEM_SPECIFIC"),
}

# Item-specific test mappings (only ITEM_SPECIFIC tests need explicit mapping)
ITEM_TESTS_MAP = [
    ("센터콘솔",     "BSR 시험",         False),
    ("센터콘솔",     "치수 검사",        False),
    ("콘솔리드",     "힌지 내구 시험",   False),
    ("콘솔리드",     "조작력 시험",      False),
    ("칵핏모듈",     "BSR 시험",         False),
    ("도어트림(앞)", "치수 검사",        False),
    ("도어트림(앞)", "조작력 시험",      False),
    ("도어트림(뒤)", "치수 검사",        False),
    ("필러트림 A",   "치수 검사",        False),
    ("필러트림 B",   "치수 검사",        False),
    ("필러트림 C",   "치수 검사",        False),
    ("시트(앞)",     "시트 내구 시험",   False),
    ("시트(앞)",     "H-Point 검증",     False),
    ("시트(앞)",     "안전벨트 앵커리지", False),
    ("시트(뒤)",     "시트 내구 시험",   False),
    ("시트(뒤)",     "안전벨트 앵커리지", False),
    ("시트(뒤)",     "ISOFIX 강도 시험", False),
    ("시트(뒤)",     "수하물 충격 시험", False),
]

# Demo items and their tests for sim data (4 items x varied tests = 23 combos per vehicle)
DEMO_ITEMS = ["센터콘솔", "콘솔리드", "시트(앞)", "시트(뒤)"]
DEMO_COMMON = ["EMC 시험", "VOC 방출 시험", "연소성 시험"]
DEMO_SPECIFIC = {
    "센터콘솔": ["BSR 시험", "치수 검사"],
    "콘솔리드": ["힌지 내구 시험", "조작력 시험"],
    "시트(앞)": ["시트 내구 시험", "H-Point 검증", "안전벨트 앵커리지"],
    "시트(뒤)": ["시트 내구 시험", "안전벨트 앵커리지", "ISOFIX 강도 시험", "수하물 충격 시험"],
}

# 23 target states per vehicle (all 12 states appear):
STATE_PATTERN = [
    1, 5, 9, 11, 3,         # 센터콘솔 (5 tests: 3 common + 2 specific)
    2, 8, 10, 6, 11,        # 콘솔리드 (5 tests: 3 common + 2 specific)
    4, 9, 11, 5, 7, 3,      # 시트(앞) (6 tests: 3 common + 3 specific)
    8, 12, 1, 11, 10, 9, 11 # 시트(뒤) (7 tests: 3 common + 4 specific)
]

# == Color Palette ==
C_DARK_NAVY    = "1B2A4A"
C_NAVY         = "2C3E6B"
C_MID_BLUE     = "3B5998"
C_LIGHT_BLUE   = "D6E4F0"
C_VERY_LIGHT   = "EBF1F8"
C_WHITE        = "FFFFFF"
C_RED          = "E74C3C"
C_RED_LIGHT    = "FADBD8"
C_ORANGE       = "F39C12"
C_ORANGE_LIGHT = "FDEBD0"
C_YELLOW       = "F1C40F"
C_YELLOW_LIGHT = "FEF9E7"
C_GREEN        = "27AE60"
C_GREEN_LIGHT  = "D5F5E3"
C_GRAY         = "BDC3C7"
C_GRAY_LIGHT   = "F2F3F4"
C_CHARCOAL     = "2C3E50"
C_BLACK        = "000000"
C_LIGHT_GREEN  = "ABEBC6"

C_MANUAL_YELLOW = "FFF9C4"

# == Reusable Styles ==
THIN_BORDER = Border(
    left=Side(style="thin", color=C_GRAY),
    right=Side(style="thin", color=C_GRAY),
    top=Side(style="thin", color=C_GRAY),
    bottom=Side(style="thin", color=C_GRAY),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color=C_NAVY),
    right=Side(style="medium", color=C_NAVY),
    top=Side(style="medium", color=C_NAVY),
    bottom=Side(style="medium", color=C_NAVY),
)

FONT_TITLE      = Font(name="Malgun Gothic", size=18, bold=True, color=C_WHITE)
FONT_SUBTITLE   = Font(name="Malgun Gothic", size=13, bold=True, color=C_WHITE)
FONT_HEADER     = Font(name="Malgun Gothic", size=10, bold=True, color=C_WHITE)
FONT_HEADER_KR  = Font(name="Malgun Gothic", size=9, bold=True, color=C_DARK_NAVY)
FONT_BODY       = Font(name="Malgun Gothic", size=10, color=C_CHARCOAL)
FONT_BODY_BOLD  = Font(name="Malgun Gothic", size=10, bold=True, color=C_CHARCOAL)
FONT_SMALL      = Font(name="Malgun Gothic", size=9, color=C_CHARCOAL)
FONT_KPI_BIG    = Font(name="Malgun Gothic", size=28, bold=True, color=C_DARK_NAVY)
FONT_KPI_LABEL  = Font(name="Malgun Gothic", size=10, color=C_MID_BLUE)
FONT_RED_BOLD   = Font(name="Malgun Gothic", size=11, bold=True, color=C_RED)
FONT_GREEN_BOLD = Font(name="Malgun Gothic", size=11, bold=True, color=C_GREEN)
FONT_LINK       = Font(name="Malgun Gothic", size=10, color="2980B9", underline="single")
FONT_REF_TITLE  = Font(name="Malgun Gothic", size=12, bold=True, color=C_DARK_NAVY)

FILL_NAVY        = PatternFill(start_color=C_DARK_NAVY, end_color=C_DARK_NAVY, fill_type="solid")
FILL_MID_BLUE    = PatternFill(start_color=C_MID_BLUE, end_color=C_MID_BLUE, fill_type="solid")
FILL_LIGHT_BLUE  = PatternFill(start_color=C_LIGHT_BLUE, end_color=C_LIGHT_BLUE, fill_type="solid")
FILL_VERY_LIGHT  = PatternFill(start_color=C_VERY_LIGHT, end_color=C_VERY_LIGHT, fill_type="solid")
FILL_WHITE       = PatternFill(start_color=C_WHITE, end_color=C_WHITE, fill_type="solid")
FILL_RED         = PatternFill(start_color=C_RED, end_color=C_RED, fill_type="solid")
FILL_RED_LIGHT   = PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid")
FILL_ORANGE_LIGHT= PatternFill(start_color=C_ORANGE_LIGHT, end_color=C_ORANGE_LIGHT, fill_type="solid")
FILL_YELLOW_LIGHT= PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid")
FILL_GREEN_LIGHT = PatternFill(start_color=C_GREEN_LIGHT, end_color=C_GREEN_LIGHT, fill_type="solid")
FILL_GREEN       = PatternFill(start_color=C_GREEN, end_color=C_GREEN, fill_type="solid")
FILL_GRAY_LIGHT  = PatternFill(start_color=C_GRAY_LIGHT, end_color=C_GRAY_LIGHT, fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color=C_LIGHT_GREEN, end_color=C_LIGHT_GREEN, fill_type="solid")
FILL_MANUAL      = PatternFill(start_color=C_MANUAL_YELLOW, end_color=C_MANUAL_YELLOW, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")


# ================================================================
# Helpers
# ================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def write_row(ws, row, col_start, values, font=FONT_BODY, fill=None, alignment=ALIGN_CENTER, border=THIN_BORDER):
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        apply_style(c, font=font, fill=fill, alignment=alignment, border=border)


def merge_and_style(ws, r1, c1, r2, c2, value, font, fill, alignment=ALIGN_CENTER, border=MEDIUM_BORDER):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    apply_style(cell, font=font, fill=fill, alignment=alignment, border=border)
    for r in range(r1, r2 + 1):
        for ci in range(c1, c2 + 1):
            ws.cell(row=r, column=ci).border = border


def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def styled_cell(ws, row, col, value, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    apply_style(c, font=font, fill=fill, alignment=alignment, border=border, number_format=number_format)
    return c


def write_table_header(ws, row, col_start, headers, font=FONT_HEADER, fill=FILL_NAVY, alignment=ALIGN_CENTER, border=THIN_BORDER):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        apply_style(c, font=font, fill=fill, alignment=alignment, border=border)


def add_openpyxl_table(ws, name, ref, style_name="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style_name, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


# ================================================================
# Simulation Data
# ================================================================

def generate_sim_data():
    """Generate 69 rows (3 vehicles x 23 item-test combos), all 12 states covered."""
    d = datetime.timedelta
    sim = []
    no = 0
    for v in VEHICLE_ORDER:
        t0 = T0_DATES[v]
        si = 0
        for item_name in DEMO_ITEMS:
            tests = DEMO_COMMON + DEMO_SPECIFIC[item_name]
            for t in tests:
                no += 1
                st = STATE_PATTERN[si]; si += 1
                pd = t0 + d(days=10 + no % 5)
                sd = t0 + d(days=45 + no % 8)
                ud = t0 + d(days=62 + no % 6)
                wd = t0 + d(days=85 + no % 7)
                yd = t0 + d(days=93 + no % 4)
                fields = {
                    1:  (None, None, None, None, None, None, None),
                    2:  (pd, None, None, None, None, None, None),
                    3:  (pd, "Under Review", None, None, None, None, None),
                    4:  (pd, "Revision Needed", None, None, None, None, None),
                    5:  (pd, "Approved", None, None, None, None, None),
                    6:  (pd, "Approved", None, None, None, None, "06-BIW중단"),
                    7:  (pd, "Approved", sd, None, None, None, "07-미인수"),
                    8:  (pd, "Approved", sd, None, None, None, None),
                    9:  (pd, "Approved", sd, ud, None, None, None),
                    10: (pd, "Approved", sd, ud, wd, None, None),
                    11: (pd, "Approved", sd, ud, wd, yd, None),
                    12: (pd, "Approved", sd, ud, wd, yd, "12-조건부승인"),
                }[st]
                # Tuple: (no, vehicle, item, testname, k_sub, l_rev, n_act, p_act, r_act, t_sub, u_ovr)
                sim.append((no, v, item_name, t, *fields))
    return sim


# ================================================================
# Sheet 1: Config (hidden helper)
# ================================================================

def build_config(wb):
    ws = wb.active
    ws.title = "Config"
    ws.sheet_properties.tabColor = C_GRAY

    # A1: Selected Vehicle (default first vehicle)
    ws.cell(row=1, column=1, value="AE_PE")
    apply_style(ws.cell(row=1, column=1), font=FONT_BODY_BOLD, fill=FILL_LIGHT_BLUE,
                alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Label
    ws.cell(row=1, column=2, value="<-- Select Vehicle Code")
    apply_style(ws.cell(row=1, column=2), font=FONT_SMALL, alignment=ALIGN_LEFT)

    # Vehicle dropdown from Reference Data
    dv = DataValidation(
        type="list",
        formula1="'Reference Data 기준데이터'!$A$3:$A$5",
        allow_blank=False
    )
    dv.prompt = "Select a vehicle code"
    dv.promptTitle = "Vehicle"
    ws.add_data_validation(dv)
    dv.add("A1")

    set_col_widths(ws, {"A": 16, "B": 30})
    ws.sheet_state = "hidden"
    return ws


# ================================================================
# Sheet 2: Reference Data (SSOT - 4 tables)
# ================================================================

def build_reference_data(wb):
    ws = wb.create_sheet("Reference Data 기준데이터")
    ws.sheet_properties.tabColor = C_MID_BLUE

    set_col_widths(ws, {
        "A": 30, "B": 28, "C": 18, "D": 18, "E": 18, "F": 18,
        "G": 16, "H": 16, "I": 16, "J": 16,
    })

    # ── TBL_Vehicle (A1:J5) — 10 cols, 3 vehicles ──
    r = 1
    styled_cell(ws, r, 1, "TBL_Vehicle (차종 + 일정 + LeadTime)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 2
    vh = ["VehicleCode", "VehicleName", "T0_DrawingRelease", "ProtoBuild",
          "MasterCar", "TestDeadline", "LT_PlanDue", "LT_PartsReady",
          "LT_TestStart", "LT_ReportDue"]
    write_table_header(ws, r, 1, vh)
    vehicles = [
        ("AE_PE", "AE PE (Creta EV)", datetime.date(2026, 1, 15),
         datetime.date(2026, 4, 15), datetime.date(2026, 6, 15), None,
         14, 49, 63, 98),
        ("SU2i", "SU2i (Venue Next)", datetime.date(2026, 3, 1),
         datetime.date(2026, 6, 1), datetime.date(2026, 8, 15), None,
         14, 49, 63, 98),
        ("EN_SUV", "EN SUV", datetime.date(2026, 2, 1),
         datetime.date(2026, 5, 1), datetime.date(2026, 7, 15), None,
         14, 49, 63, 98),
    ]
    for i, v in enumerate(vehicles):
        row = r + 1 + i
        for j, val in enumerate(v):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
            if isinstance(val, datetime.date):
                c.number_format = "YYYY-MM-DD"
        # TestDeadline (col F=6) = MasterCar (col E=5) - 30
        td_cell = ws.cell(row=row, column=6)
        td_cell.value = f"=E{row}-30"
        apply_style(td_cell, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
        td_cell.number_format = "YYYY-MM-DD"

    add_openpyxl_table(ws, "TBL_Vehicle", f"A2:J{r + len(vehicles)}", "TableStyleMedium2")

    # ── TBL_TestCatalog (A7:E26) — 5 cols, 18 tests ──
    r = 7
    styled_cell(ws, r, 1, "TBL_TestCatalog (시험 카탈로그 — 18 tests)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 8
    th = ["TestNameKR", "TestType", "Spec", "DurationDays", "Scope"]
    write_table_header(ws, r, 1, th)
    unique_tests = list(TEST_CATALOG.items())  # 18 unique tests
    for i, (tname, (ttype, spec, dur, scope)) in enumerate(unique_tests):
        row = r + 1 + i
        for j, val in enumerate([tname, ttype, spec, dur, scope]):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE,
                        alignment=ALIGN_LEFT if j == 0 else ALIGN_CENTER, border=THIN_BORDER)

    add_openpyxl_table(ws, "TBL_TestCatalog", f"A8:E{r + len(unique_tests)}", "TableStyleMedium6")

    # ── TBL_ItemMaster (A28:D41) — 4 cols, 12 items ──
    r = 28
    styled_cell(ws, r, 1, "TBL_ItemMaster (아이템 → 구분 + 협력사 + 도면번호)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 29
    ph = ["ItemNameKR", "PartCategory", "Supplier", "DrawingNo"]
    write_table_header(ws, r, 1, ph)
    for i, item_name in enumerate(ITEM_NAMES):
        row = r + 1 + i
        cat, supplier, drawing = ITEM_DATA[item_name]
        vals = [item_name, cat, supplier, drawing]
        for j, val in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE,
                        alignment=ALIGN_LEFT if j == 0 else ALIGN_CENTER, border=THIN_BORDER)

    add_openpyxl_table(ws, "TBL_ItemMaster", f"A29:D{r + len(ITEM_NAMES)}", "TableStyleMedium5")

    # ── TBL_ItemTests (A43:C61) — 3 cols, ~18 rows ──
    r = 43
    styled_cell(ws, r, 1, "TBL_ItemTests (아이템 × 시험 매핑 + 제외)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 44
    it_h = ["ItemNameKR", "TestNameKR", "Excluded"]
    write_table_header(ws, r, 1, it_h)
    for i, (item_name, test_name, excluded) in enumerate(ITEM_TESTS_MAP):
        row = r + 1 + i
        vals = [item_name, test_name, excluded]
        for j, val in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE,
                        alignment=ALIGN_LEFT if j < 2 else ALIGN_CENTER, border=THIN_BORDER)

    add_openpyxl_table(ws, "TBL_ItemTests", f"A44:C{r + len(ITEM_TESTS_MAP)}", "TableStyleMedium4")

    return ws


# ================================================================
# Sheet 3: Master Tracker (28 cols, A-AB)
# ================================================================

# V4 Column mapping (1-based): A=1 .. AB=28
# A(1):No  B(2):Vehicle[M]  C(3):Part[AUTO]  D(4):Item[M]  E(5):Supplier[AUTO]
# F(6):TestName[M]  G(7):TestType[AUTO]  H(8):Scope[AUTO]
# I(9):T0[AUTO]  J(10):PlanDue[AUTO]  K(11):PlanSubmitted[M]  L(12):ReviewResult[M]
# M(13):PartsReadyPlan[AUTO]  N(14):PartsReadyActual[M]
# O(15):TestStartPlan[AUTO]  P(16):TestStartActual[M]
# Q(17):TestCompletePlan[AUTO]  R(18):TestCompleteActual[M]
# S(19):ReportDue[AUTO]  T(20):ReportSubmitted[M]  U(21):Override[M]
# V(22):Status[AUTO]  W(23):Progress%[AUTO]  X(24):DaysToDeadline[AUTO]
# Y(25):RiskFlag[AUTO]  Z(26):DaysSinceActivity[AUTO]  AA(27):NextAction[AUTO]  AB(28):Note[M]

# Manual columns (1-based): B(2), D(4), F(6), K(11), L(12), N(14), P(16), R(18), T(20), U(21), AB(28)
MANUAL_COLS = {2, 4, 6, 11, 12, 14, 16, 18, 20, 21, 28}

def build_master_tracker(wb):
    ws = wb.create_sheet("Master Tracker 마스터추적표")
    ws.sheet_properties.tabColor = C_MID_BLUE

    # -- English headers (row 3) --
    headers_en = [
        "No.",                          # A  (1)
        "Vehicle Code",                 # B  (2)  MANUAL
        "Part Category",                # C  (3)  AUTO
        "Item Name",                    # D  (4)  MANUAL
        "Supplier",                     # E  (5)  AUTO
        "Test Name KR",                 # F  (6)  MANUAL
        "Test Type",                    # G  (7)  AUTO
        "Scope",                        # H  (8)  AUTO
        "T\u2080 Drawing\nRelease",    # I  (9)  AUTO
        "Plan Due\nDate",              # J  (10) AUTO
        "Plan Submitted\nDate",        # K  (11) MANUAL
        "Review\nResult",             # L  (12) MANUAL
        "Parts Ready\n(Plan)",         # M  (13) AUTO
        "Parts Ready\n(Actual)",       # N  (14) MANUAL
        "Test Start\n(Plan)",          # O  (15) AUTO
        "Test Start\n(Actual)",        # P  (16) MANUAL
        "Test Complete\n(Plan)",       # Q  (17) AUTO
        "Test Complete\n(Actual)",     # R  (18) MANUAL
        "Report Due\nDate",            # S  (19) AUTO
        "Report Submitted\nDate",      # T  (20) MANUAL
        "Override",                     # U  (21) MANUAL
        "Status",                       # V  (22) AUTO
        "Progress\n%",                 # W  (23) AUTO
        "Days to\nDeadline",          # X  (24) AUTO
        "Risk\nFlag",                  # Y  (25) AUTO
        "Days Since\nActivity",        # Z  (26) AUTO
        "Next Action",                  # AA (27) AUTO
        "Note",                         # AB (28) MANUAL
    ]

    # -- Korean headers (row 4) --
    headers_kr = [
        "번호",              # A
        "차종",              # B
        "파트구분",          # C
        "아이템명",          # D
        "협력사",            # E
        "시험명",            # F
        "시험유형",          # G
        "범위",              # H
        "T₀ 도면배포",      # I
        "계획서기한",        # J
        "계획서제출",        # K
        "검토결과",          # L
        "부품준비(계획)",    # M
        "부품준비(실적)",    # N
        "시험시작(계획)",    # O
        "시험시작(실적)",    # P
        "시험완료(계획)",    # Q
        "시험완료(실적)",    # R
        "성적서기한",        # S
        "성적서제출",        # T
        "오버라이드",        # U
        "상태",              # V
        "진행률%",          # W
        "잔여일",            # X
        "리스크",            # Y
        "정체일",            # Z
        "다음조치",          # AA
        "비고",              # AB
    ]

    num_cols = len(headers_en)  # 28

    # -- Title rows (1-2) --
    merge_and_style(ws, 1, 1, 2, num_cols,
                    "ESIR Master Tracker V4 — Engineering Sample / Initial Run Test (DB Schema v2)",
                    FONT_TITLE, FILL_NAVY)

    # -- Header rows (3-4) --
    for i, h in enumerate(headers_en):
        col = i + 1
        c = ws.cell(row=3, column=col, value=h)
        if col in MANUAL_COLS:
            apply_style(c, font=FONT_HEADER, fill=FILL_MID_BLUE, alignment=ALIGN_CENTER, border=THIN_BORDER)
        else:
            apply_style(c, font=FONT_HEADER, fill=FILL_NAVY, alignment=ALIGN_CENTER, border=THIN_BORDER)

    for i, h in enumerate(headers_kr):
        col = i + 1
        c = ws.cell(row=4, column=col, value=h)
        if col in MANUAL_COLS:
            apply_style(c, font=FONT_HEADER_KR, fill=FILL_MANUAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
        else:
            apply_style(c, font=FONT_HEADER_KR, fill=FILL_LIGHT_BLUE, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # -- Column widths --
    col_widths = {
        "A": 6, "B": 12, "C": 12, "D": 18, "E": 18,
        "F": 20, "G": 14, "H": 14, "I": 14, "J": 14,
        "K": 14, "L": 16, "M": 14, "N": 14, "O": 14,
        "P": 14, "Q": 14, "R": 14, "S": 14, "T": 14,
        "U": 18, "V": 22, "W": 10, "X": 10, "Y": 12,
        "Z": 10, "AA": 30, "AB": 20,
    }
    set_col_widths(ws, col_widths)

    # -- Row height for headers --
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 24

    # -- Generate sim data and write --
    sim = generate_sim_data()
    data_start = 5
    num_items = len(sim)  # 69

    for idx, item in enumerate(sim):
        r = data_start + idx
        no, vehicle, item_name, testname, k_sub, l_rev, n_act, p_act, r_act, t_sub, u_ovr = item

        # A: No. (static)
        styled_cell(ws, r, 1, no, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # B: Vehicle Code (MANUAL)
        styled_cell(ws, r, 2, vehicle, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # C: Part Category (AUTO) = VLOOKUP(D, RNG_ITEMMASTER, 2, FALSE)
        f_c = f'=IFERROR(VLOOKUP(D{r},{RNG_ITEMMASTER},2,FALSE),"")'
        styled_cell(ws, r, 3, f_c, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # D: Item Name (MANUAL)
        styled_cell(ws, r, 4, item_name, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # E: Supplier (AUTO) = VLOOKUP(D, RNG_ITEMMASTER, 3, FALSE)
        f_e = f'=IFERROR(VLOOKUP(D{r},{RNG_ITEMMASTER},3,FALSE),"")'
        styled_cell(ws, r, 5, f_e, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # F: Test Name KR (MANUAL)
        styled_cell(ws, r, 6, testname, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # G: Test Type (AUTO) = VLOOKUP(F, RNG_TESTCATALOG, 2, FALSE)
        f_g = f'=IFERROR(VLOOKUP(F{r},{RNG_TESTCATALOG},2,FALSE),"")'
        styled_cell(ws, r, 7, f_g, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # H: Scope (AUTO) = VLOOKUP(F, RNG_TESTCATALOG, 5, FALSE)
        f_h = f'=IFERROR(VLOOKUP(F{r},{RNG_TESTCATALOG},5,FALSE),"")'
        styled_cell(ws, r, 8, f_h, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # I: T0 Drawing Release (AUTO) = VLOOKUP(B, RNG_VEHICLE, 3, FALSE)
        f_i = f'=IFERROR(VLOOKUP(B{r},{RNG_VEHICLE},3,FALSE),"")'
        c_i = styled_cell(ws, r, 9, f_i, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_i.number_format = "YYYY-MM-DD"

        # J: Plan Due (AUTO) = I + VLOOKUP(B, VEHICLE, 7) [LT_PlanDue]
        f_j = f'=IFERROR(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},7,FALSE),"")'
        c_j = styled_cell(ws, r, 10, f_j, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_j.number_format = "YYYY-MM-DD"

        # K: Plan Submitted (MANUAL)
        c_k = styled_cell(ws, r, 11, k_sub, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(k_sub, datetime.date):
            c_k.number_format = "YYYY-MM-DD"

        # L: Review Result (MANUAL)
        styled_cell(ws, r, 12, l_rev, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # M: Parts Ready Plan (AUTO) = I + VLOOKUP(B, VEHICLE, 8) [LT_PartsReady]
        f_m = f'=IFERROR(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},8,FALSE),"")'
        c_m = styled_cell(ws, r, 13, f_m, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_m.number_format = "YYYY-MM-DD"

        # N: Parts Ready Actual (MANUAL)
        c_n = styled_cell(ws, r, 14, n_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(n_act, datetime.date):
            c_n.number_format = "YYYY-MM-DD"

        # O: Test Start Plan (AUTO) — vehicle-mounted tests use MAX with ProtoBuild
        f_o = (
            f'=IFERROR(IF(G{r}="차량장착",'
            f'MAX(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},4,FALSE)),'
            f'I{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE)),"")'
        )
        c_o = styled_cell(ws, r, 15, f_o, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_o.number_format = "YYYY-MM-DD"

        # P: Test Start Actual (MANUAL)
        c_p = styled_cell(ws, r, 16, p_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(p_act, datetime.date):
            c_p.number_format = "YYYY-MM-DD"

        # Q: Test Complete Plan (AUTO) = O + VLOOKUP(F, TestCatalog, 4=DurationDays)
        f_q = f'=IFERROR(O{r}+VLOOKUP(F{r},{RNG_TESTCATALOG},4,FALSE),"")'
        c_q = styled_cell(ws, r, 17, f_q, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_q.number_format = "YYYY-MM-DD"

        # R: Test Complete Actual (MANUAL)
        c_r = styled_cell(ws, r, 18, r_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(r_act, datetime.date):
            c_r.number_format = "YYYY-MM-DD"

        # S: Report Due (AUTO) = MIN(I+VLOOKUP(B,VEHICLE,10), TestDeadline)
        f_s = (
            f'=IFERROR(MIN(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},10,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},6,FALSE)),"")'
        )
        c_s = styled_cell(ws, r, 19, f_s, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_s.number_format = "YYYY-MM-DD"

        # T: Report Submitted (MANUAL)
        c_t = styled_cell(ws, r, 20, t_sub, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(t_sub, datetime.date):
            c_t.number_format = "YYYY-MM-DD"

        # U: Override (MANUAL)
        styled_cell(ws, r, 21, u_ovr, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # V: Status (AUTO) — 12-state formula
        f_v = (
            f'=IF(U{r}<>"",U{r},'
            f'IF(AND(T{r}<>"",R{r}<>""),"11-결과등록완료",'
            f'IF(AND(R{r}<>"",T{r}=""),"10-시험완료",'
            f'IF(AND(P{r}<>"",R{r}=""),"09-시험준비중",'
            f'IF(AND(N{r}<>"",P{r}=""),"08-단품준비완료",'
            f'IF(L{r}="Approved","05-계획서승인",'
            f'IF(L{r}="Revision Needed","04-검토중_보완",'
            f'IF(L{r}="Under Review","03-검토중",'
            f'IF(AND(K{r}<>"",L{r}=""),"02-제출_미검토",'
            f'"01-미제출")))))))))'
        )
        styled_cell(ws, r, 22, f_v, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # W: Progress % (AUTO) = IFS on Status
        f_w = (
            f'=IFS(V{r}="01-미제출",0,V{r}="02-제출_미검토",0.05,'
            f'V{r}="03-검토중",0.1,V{r}="04-검토중_보완",0.05,'
            f'V{r}="05-계획서승인",0.2,V{r}="06-BIW중단",0.15,'
            f'V{r}="07-미인수",0.3,V{r}="08-단품준비완료",0.4,'
            f'V{r}="09-시험준비중",0.55,V{r}="10-시험완료",0.85,'
            f'V{r}="11-결과등록완료",1,V{r}="12-조건부승인",0.9,TRUE,0)'
        )
        styled_cell(ws, r, 23, f_w, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # X: Days to Deadline (AUTO)
        f_x = f'=IF(T{r}<>"","-",IF(S{r}="","N/A",S{r}-TODAY()))'
        styled_cell(ws, r, 24, f_x, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Y: Risk Flag (AUTO) — 7 states including BLOCKED
        f_y = (
            f'=IF(V{r}="11-결과등록완료","DONE",'
            f'IF(U{r}="06-BIW중단","BLOCKED",'
            f'IF(AND(K{r}="",J{r}<TODAY()),"CRITICAL",'
            f'IF(AND(ISNUMBER(X{r}),X{r}<=0),"OVERDUE",'
            f'IF(AND(ISNUMBER(X{r}),X{r}<=14),"WARNING",'
            f'IF(U{r}<>"","ACTION","ON TRACK"))))))'
        )
        styled_cell(ws, r, 25, f_y, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Z: Days Since Activity (AUTO) — only when real activity exists (K~T)
        f_z = f'=IF(OR(T{r}<>"",R{r}<>"",P{r}<>"",N{r}<>"",K{r}<>""),TODAY()-MAX(T{r},R{r},P{r},N{r},K{r}),"—")'
        styled_cell(ws, r, 26, f_z, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # AA: Next Action (AUTO) — 12-branch IF
        f_aa = (
            f'=IF(V{r}="11-결과등록완료","— 완료",'
            f'IF(V{r}="01-미제출",IF(J{r}<TODAY(),"계획서 제출 독촉 (OVERDUE)","계획서 제출 대기"),'
            f'IF(V{r}="02-제출_미검토","설계자: 계획서 검토 착수",'
            f'IF(V{r}="03-검토중","설계자: 승인 처리",'
            f'IF(V{r}="04-검토중_보완","보완 요청 → 협력사 재제출 독촉",'
            f'IF(V{r}="05-계획서승인","시험편 준비 현황 확인",'
            f'IF(V{r}="06-BIW중단","BIW 생산팀 확인 → 일정 재수립",'
            f'IF(V{r}="07-미인수","협력사 인수 독촉",'
            f'IF(V{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(V{r}="09-시험준비중","시험 진행 모니터링",'
            f'IF(V{r}="10-시험완료","결과 보고서 전산 등록 독촉",'
            f'IF(V{r}="12-조건부승인","조건부 승인 여부 판단 → 관련부서 협의","—"))))))))))))'
        )
        styled_cell(ws, r, 27, f_aa, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # AB: Note (MANUAL)
        styled_cell(ws, r, 28, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

    # -- Empty rows (after sim data to row TRACKER_END) with wrapped formulas --
    data_end = data_start + num_items - 1  # last sim row
    for r in range(data_end + 1, TRACKER_END + 1):
        # A: No. — auto-increment if B has value
        f_a = f'=IF(B{r}="","",ROW()-{data_start - 1})'
        styled_cell(ws, r, 1, f_a, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # B: Vehicle Code (MANUAL - blank)
        styled_cell(ws, r, 2, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # C: Part Category (AUTO, wrapped)
        f_c = f'=IF(B{r}="","",IFERROR(VLOOKUP(D{r},{RNG_ITEMMASTER},2,FALSE),""))'
        styled_cell(ws, r, 3, f_c, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # D: Item Name (MANUAL - blank)
        styled_cell(ws, r, 4, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # E: Supplier (AUTO, wrapped)
        f_e = f'=IF(B{r}="","",IFERROR(VLOOKUP(D{r},{RNG_ITEMMASTER},3,FALSE),""))'
        styled_cell(ws, r, 5, f_e, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # F: Test Name KR (MANUAL - blank)
        styled_cell(ws, r, 6, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # G: Test Type (AUTO, wrapped)
        f_g = f'=IF(B{r}="","",IFERROR(VLOOKUP(F{r},{RNG_TESTCATALOG},2,FALSE),""))'
        styled_cell(ws, r, 7, f_g, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # H: Scope (AUTO, wrapped)
        f_h = f'=IF(B{r}="","",IFERROR(VLOOKUP(F{r},{RNG_TESTCATALOG},5,FALSE),""))'
        styled_cell(ws, r, 8, f_h, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # I: T0 (AUTO, wrapped)
        f_i = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},{RNG_VEHICLE},3,FALSE),""))'
        c_i = styled_cell(ws, r, 9, f_i, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_i.number_format = "YYYY-MM-DD"

        # J: Plan Due (AUTO, wrapped)
        f_j = f'=IF(B{r}="","",IFERROR(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},7,FALSE),""))'
        c_j = styled_cell(ws, r, 10, f_j, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_j.number_format = "YYYY-MM-DD"

        # K: Plan Submitted (MANUAL - blank)
        c_k = styled_cell(ws, r, 11, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_k.number_format = "YYYY-MM-DD"

        # L: Review Result (MANUAL - blank)
        styled_cell(ws, r, 12, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # M: Parts Ready Plan (AUTO, wrapped)
        f_m = f'=IF(B{r}="","",IFERROR(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},8,FALSE),""))'
        c_m = styled_cell(ws, r, 13, f_m, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_m.number_format = "YYYY-MM-DD"

        # N: Parts Ready Actual (MANUAL - blank)
        c_n = styled_cell(ws, r, 14, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_n.number_format = "YYYY-MM-DD"

        # O: Test Start Plan (AUTO, wrapped)
        f_o = (
            f'=IF(B{r}="","",IFERROR(IF(G{r}="차량장착",'
            f'MAX(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},4,FALSE)),'
            f'I{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE)),""))'
        )
        c_o = styled_cell(ws, r, 15, f_o, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_o.number_format = "YYYY-MM-DD"

        # P: Test Start Actual (MANUAL - blank)
        c_p = styled_cell(ws, r, 16, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_p.number_format = "YYYY-MM-DD"

        # Q: Test Complete Plan (AUTO, wrapped)
        f_q = f'=IF(B{r}="","",IFERROR(O{r}+VLOOKUP(F{r},{RNG_TESTCATALOG},4,FALSE),""))'
        c_q = styled_cell(ws, r, 17, f_q, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_q.number_format = "YYYY-MM-DD"

        # R: Test Complete Actual (MANUAL - blank)
        c_r = styled_cell(ws, r, 18, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_r.number_format = "YYYY-MM-DD"

        # S: Report Due (AUTO, wrapped)
        f_s = (
            f'=IF(B{r}="","",IFERROR(MIN(I{r}+VLOOKUP(B{r},{RNG_VEHICLE},10,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},6,FALSE)),""))'
        )
        c_s = styled_cell(ws, r, 19, f_s, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_s.number_format = "YYYY-MM-DD"

        # T: Report Submitted (MANUAL - blank)
        c_t = styled_cell(ws, r, 20, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_t.number_format = "YYYY-MM-DD"

        # U: Override (MANUAL - blank)
        styled_cell(ws, r, 21, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # V: Status (AUTO, wrapped)
        f_v = (
            f'=IF(B{r}="","",'
            f'IF(U{r}<>"",U{r},'
            f'IF(AND(T{r}<>"",R{r}<>""),"11-결과등록완료",'
            f'IF(AND(R{r}<>"",T{r}=""),"10-시험완료",'
            f'IF(AND(P{r}<>"",R{r}=""),"09-시험준비중",'
            f'IF(AND(N{r}<>"",P{r}=""),"08-단품준비완료",'
            f'IF(L{r}="Approved","05-계획서승인",'
            f'IF(L{r}="Revision Needed","04-검토중_보완",'
            f'IF(L{r}="Under Review","03-검토중",'
            f'IF(AND(K{r}<>"",L{r}=""),"02-제출_미검토",'
            f'"01-미제출"))))))))))'
        )
        styled_cell(ws, r, 22, f_v, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # W: Progress % (AUTO, wrapped with IFS)
        f_w = (
            f'=IF(B{r}="","",IFS(V{r}="01-미제출",0,V{r}="02-제출_미검토",0.05,'
            f'V{r}="03-검토중",0.1,V{r}="04-검토중_보완",0.05,'
            f'V{r}="05-계획서승인",0.2,V{r}="06-BIW중단",0.15,'
            f'V{r}="07-미인수",0.3,V{r}="08-단품준비완료",0.4,'
            f'V{r}="09-시험준비중",0.55,V{r}="10-시험완료",0.85,'
            f'V{r}="11-결과등록완료",1,V{r}="12-조건부승인",0.9,TRUE,0))'
        )
        styled_cell(ws, r, 23, f_w, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # X: Days to Deadline (AUTO, wrapped)
        f_x = f'=IF(B{r}="","",IF(T{r}<>"","-",IF(S{r}="","N/A",S{r}-TODAY())))'
        styled_cell(ws, r, 24, f_x, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Y: Risk Flag (AUTO, wrapped) — 7 states including BLOCKED
        f_y = (
            f'=IF(B{r}="","",'
            f'IF(V{r}="11-결과등록완료","DONE",'
            f'IF(U{r}="06-BIW중단","BLOCKED",'
            f'IF(AND(K{r}="",J{r}<TODAY()),"CRITICAL",'
            f'IF(AND(ISNUMBER(X{r}),X{r}<=0),"OVERDUE",'
            f'IF(AND(ISNUMBER(X{r}),X{r}<=14),"WARNING",'
            f'IF(U{r}<>"","ACTION","ON TRACK")))))))'
        )
        styled_cell(ws, r, 25, f_y, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Z: Days Since Activity (AUTO, wrapped) — only when real activity exists (K~T)
        f_z = f'=IF(B{r}="","",IF(OR(T{r}<>"",R{r}<>"",P{r}<>"",N{r}<>"",K{r}<>""),TODAY()-MAX(T{r},R{r},P{r},N{r},K{r}),"—"))'
        styled_cell(ws, r, 26, f_z, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # AA: Next Action (AUTO, wrapped)
        f_aa = (
            f'=IF(B{r}="","",'
            f'IF(V{r}="11-결과등록완료","— 완료",'
            f'IF(V{r}="01-미제출",IF(J{r}<TODAY(),"계획서 제출 독촉 (OVERDUE)","계획서 제출 대기"),'
            f'IF(V{r}="02-제출_미검토","설계자: 계획서 검토 착수",'
            f'IF(V{r}="03-검토중","설계자: 승인 처리",'
            f'IF(V{r}="04-검토중_보완","보완 요청 → 협력사 재제출 독촉",'
            f'IF(V{r}="05-계획서승인","시험편 준비 현황 확인",'
            f'IF(V{r}="06-BIW중단","BIW 생산팀 확인 → 일정 재수립",'
            f'IF(V{r}="07-미인수","협력사 인수 독촉",'
            f'IF(V{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(V{r}="09-시험준비중","시험 진행 모니터링",'
            f'IF(V{r}="10-시험완료","결과 보고서 전산 등록 독촉",'
            f'IF(V{r}="12-조건부승인","조건부 승인 여부 판단 → 관련부서 협의","—")))))))))))))'
        )
        styled_cell(ws, r, 27, f_aa, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # AB: Note (MANUAL - blank)
        styled_cell(ws, r, 28, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

    # -- Data Validations (rows 5:TRACKER_END) --
    # B: Vehicle Code dropdown
    dv_vehicle = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$3:$A$5", allow_blank=True)
    dv_vehicle.prompt = "Select vehicle"
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add(f"B5:B{TRACKER_END}")

    # D: Item Name dropdown (from TBL_ItemMaster)
    dv_item = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$30:$A$41", allow_blank=True)
    dv_item.prompt = "Select item"
    ws.add_data_validation(dv_item)
    dv_item.add(f"D5:D{TRACKER_END}")

    # F: Test Name KR dropdown (from TBL_TestCatalog)
    dv_test = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$9:$A$26", allow_blank=True)
    dv_test.prompt = "Select test"
    ws.add_data_validation(dv_test)
    dv_test.add(f"F5:F{TRACKER_END}")

    # L: Review Result dropdown
    dv_review = DataValidation(type="list", formula1='"Under Review,Approved,Revision Needed"', allow_blank=True)
    dv_review.prompt = "Select review result"
    ws.add_data_validation(dv_review)
    dv_review.add(f"L5:L{TRACKER_END}")

    # U: Override dropdown
    dv_override = DataValidation(
        type="list",
        formula1='"06-BIW중단,07-미인수,12-조건부승인"',
        allow_blank=True
    )
    dv_override.prompt = "Select override status"
    ws.add_data_validation(dv_override)
    dv_override.add(f"U5:U{TRACKER_END}")

    # -- Conditional Formatting --

    # V (Status) — 12 status color rules
    status_range = f"V5:V{TRACKER_END}"
    status_colors = [
        ("01-미제출", C_GRAY_LIGHT, C_CHARCOAL),
        ("02-제출_미검토", C_LIGHT_BLUE, C_CHARCOAL),
        ("03-검토중", C_LIGHT_BLUE, C_MID_BLUE),
        ("04-검토중_보완", C_ORANGE_LIGHT, C_CHARCOAL),
        ("05-계획서승인", C_GREEN_LIGHT, C_CHARCOAL),
        ("06-BIW중단", C_RED_LIGHT, C_RED),
        ("07-미인수", C_ORANGE_LIGHT, C_ORANGE),
        ("08-단품준비완료", C_GREEN_LIGHT, C_GREEN),
        ("09-시험준비중", C_GREEN_LIGHT, C_GREEN),
        ("10-시험완료", C_LIGHT_GREEN, C_CHARCOAL),
        ("11-결과등록완료", C_GREEN, C_WHITE),
        ("12-조건부승인", C_YELLOW_LIGHT, C_CHARCOAL),
    ]
    for status_code, bg_color, font_color in status_colors:
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status_code}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Malgun Gothic", size=10, bold=True, color=font_color),
            )
        )

    # Y (Risk) — 7 risk color rules
    risk_range = f"Y5:Y{TRACKER_END}"
    risk_colors = [
        ("DONE", C_GREEN_LIGHT, C_GREEN),
        ("BLOCKED", C_RED_LIGHT, C_RED),
        ("CRITICAL", C_RED_LIGHT, C_RED),
        ("OVERDUE", C_RED_LIGHT, C_RED),
        ("WARNING", C_YELLOW_LIGHT, C_ORANGE),
        ("ACTION", C_ORANGE_LIGHT, C_ORANGE),
        ("ON TRACK", C_GREEN_LIGHT, C_GREEN),
    ]
    for risk_code, bg_color, font_color in risk_colors:
        ws.conditional_formatting.add(
            risk_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{risk_code}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Malgun Gothic", size=10, bold=True, color=font_color),
            )
        )

    # W (Progress) — DataBar
    progress_range = f"W5:W{TRACKER_END}"
    ws.conditional_formatting.add(
        progress_range,
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color=C_MID_BLUE,
        )
    )

    # X (Days to Deadline) — <=0 red, <=14 yellow
    days_range = f"X5:X{TRACKER_END}"
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["14"],
            fill=PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_ORANGE),
        )
    )

    # Z (Days Since Activity) — >14 red, >7 orange
    activity_range = f"Z5:Z{TRACKER_END}"
    ws.conditional_formatting.add(
        activity_range,
        CellIsRule(
            operator="greaterThan",
            formula=["14"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        activity_range,
        CellIsRule(
            operator="greaterThan",
            formula=["7"],
            fill=PatternFill(start_color=C_ORANGE_LIGHT, end_color=C_ORANGE_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_ORANGE),
        )
    )

    # Date plan/actual pairs: if plan < TODAY and actual is empty -> red highlight
    # Plan/Actual pairs: (J,K), (M,N), (O,P), (Q,R), (S,T)
    plan_actual_pairs = [
        ("J", "K"), ("M", "N"), ("O", "P"), ("Q", "R"), ("S", "T"),
    ]
    for plan_col, actual_col in plan_actual_pairs:
        pa_range = f"{plan_col}5:{plan_col}{TRACKER_END}"
        ws.conditional_formatting.add(
            pa_range,
            FormulaRule(
                formula=[f'AND({plan_col}5<TODAY(),{actual_col}5="")'],
                fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            )
        )

    # -- Auto-filter --
    ws.auto_filter.ref = f"A4:AB{TRACKER_END}"

    # -- Freeze panes at A5 --
    ws.freeze_panes = "A5"

    return ws, data_start, data_end + 1, num_items


# ================================================================
# Sheet 4: Dashboard
# ================================================================

def build_dashboard(wb, tracker_sheet_name, data_start, data_end, num_items):
    ws = wb.create_sheet("Dashboard 대시보드")
    ws.sheet_properties.tabColor = C_GREEN

    ts = f"'{tracker_sheet_name}'"
    cfg = "$C$3"  # vehicle selector (Dashboard's own dropdown cell)

    # Column ranges on tracker (V4 columns)
    b_range = f"{ts}!$B${data_start}:$B${data_end}"   # Vehicle
    d_range = f"{ts}!$D${data_start}:$D${data_end}"   # Item name
    v_range = f"{ts}!$V${data_start}:$V${data_end}"   # Status
    w_range = f"{ts}!$W${data_start}:$W${data_end}"   # Progress
    x_range = f"{ts}!$X${data_start}:$X${data_end}"   # Days to deadline
    y_range = f"{ts}!$Y${data_start}:$Y${data_end}"   # Risk flag
    z_range = f"{ts}!$Z${data_start}:$Z${data_end}"   # Days since activity

    set_col_widths(ws, {
        "A": 4, "B": 24, "C": 14, "D": 14, "E": 14,
        "F": 14, "G": 14, "H": 14,
    })

    # -- Row 1-2: Title --
    merge_and_style(ws, 1, 1, 2, 8,
                    "ESIR Dashboard V4",
                    FONT_TITLE, FILL_NAVY)

    # -- Row 3: Vehicle selector + Key dates --
    styled_cell(ws, 3, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 3, 2, "Selected Vehicle:", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)

    # C3: vehicle selector (plain value — dropdown overwrites directly)
    styled_cell(ws, 3, 3, "AE_PE", FONT_BODY_BOLD, FILL_MANUAL, ALIGN_CENTER, MEDIUM_BORDER)

    # D3: T0 date
    styled_cell(ws, 3, 4, "T0:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    t0_cell = styled_cell(ws, 3, 5,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},3,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    t0_cell.number_format = "YYYY-MM-DD"

    # F3: Proto
    styled_cell(ws, 3, 6, "Proto:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    proto_cell = styled_cell(ws, 3, 7,
                             f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},4,FALSE),"")',
                             FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    proto_cell.number_format = "YYYY-MM-DD"

    # H3
    styled_cell(ws, 3, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Row 4: More dates + D-Day
    styled_cell(ws, 4, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 4, 2, "D-Day:", FONT_RED_BOLD, FILL_WHITE, ALIGN_RIGHT, THIN_BORDER)

    # C4: D-Day countdown = TestDeadline - TODAY()
    f_dday = f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},6,FALSE)-TODAY(),"")'
    styled_cell(ws, 4, 3, f_dday, FONT_RED_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    styled_cell(ws, 4, 4, "MC:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    mc_cell = styled_cell(ws, 4, 5,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},5,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    mc_cell.number_format = "YYYY-MM-DD"

    styled_cell(ws, 4, 6, "Deadline:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    dl_cell = styled_cell(ws, 4, 7,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},6,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    dl_cell.number_format = "YYYY-MM-DD"

    styled_cell(ws, 4, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Vehicle dropdown on C3
    dv_vehicle = DataValidation(
        type="list",
        formula1=f"'Reference Data 기준데이터'!$A$3:$A$5",
        allow_blank=False
    )
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add("C3")

    # -- Row 5: blank separator --
    for col in range(1, 9):
        styled_cell(ws, 5, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # ========================================
    # Section 1: KPI Summary (rows 6-8)
    # ========================================
    merge_and_style(ws, 6, 1, 6, 8,
                    "KPI Summary",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 7: KPI labels --
    kpi_labels = ["", "Total", "Complete", "In Progress", "At Risk", "Avg Progress", "Stagnant", ""]
    for i, label in enumerate(kpi_labels):
        styled_cell(ws, 7, i + 1, label, FONT_KPI_LABEL, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)

    # -- Row 8: KPI values --
    styled_cell(ws, 8, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # B8: Total = COUNTIF(B_range, vehicle)
    f_total = f'=COUNTIF({b_range},{cfg})'
    styled_cell(ws, 8, 2, f_total, FONT_KPI_BIG, FILL_WHITE, ALIGN_CENTER, MEDIUM_BORDER)

    # C8: Complete = COUNTIFS(B, vehicle, V, "11-*")
    f_complete = f'=COUNTIFS({b_range},{cfg},{v_range},"11-*")'
    styled_cell(ws, 8, 3, f_complete, FONT_KPI_BIG, FILL_GREEN_LIGHT, ALIGN_CENTER, MEDIUM_BORDER)

    # D8: In Progress = Total - Complete - NotStarted(01)
    f_notstarted = f'COUNTIFS({b_range},{cfg},{v_range},"01-*")'
    f_inprogress = f'=B8-C8-{f_notstarted}'
    styled_cell(ws, 8, 4, f_inprogress, FONT_KPI_BIG, FILL_LIGHT_BLUE, ALIGN_CENTER, MEDIUM_BORDER)

    # E8: At Risk = CRITICAL + OVERDUE + WARNING
    f_risk = (
        f'=COUNTIFS({b_range},{cfg},{y_range},"CRITICAL")'
        f'+COUNTIFS({b_range},{cfg},{y_range},"OVERDUE")'
        f'+COUNTIFS({b_range},{cfg},{y_range},"WARNING")'
    )
    styled_cell(ws, 8, 5, f_risk, FONT_KPI_BIG, FILL_RED_LIGHT, ALIGN_CENTER, MEDIUM_BORDER)

    # F8: Avg Progress
    f_avg = f'=IFERROR(AVERAGEIFS({w_range},{b_range},{cfg}),0)'
    styled_cell(ws, 8, 6, f_avg, FONT_KPI_BIG, FILL_WHITE, ALIGN_CENTER, MEDIUM_BORDER, number_format="0%")

    # G8: Stagnant = items with days_since_activity > 7, not completed, numeric only (exclude "—")
    f_stagnant = f'=SUMPRODUCT(({b_range}={cfg})*ISNUMBER({z_range})*({z_range}>7)*({v_range}<>"11-결과등록완료"))'
    styled_cell(ws, 8, 7, f_stagnant, FONT_KPI_BIG, FILL_ORANGE_LIGHT, ALIGN_CENTER, MEDIUM_BORDER)

    styled_cell(ws, 8, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # -- Row 9: blank --
    for col in range(1, 9):
        styled_cell(ws, 9, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # ========================================
    # Section 2: Item Progress (rows 10-15) — 4 demo items
    # ========================================
    merge_and_style(ws, 10, 1, 10, 8,
                    "Item Progress 아이템별 현황 (Selected Vehicle)",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 11: Headers --
    item_headers = ["", "아이템명", "시험수", "완료", "진행중", "지연", "진행률", "최소잔여일"]
    for i, h in enumerate(item_headers):
        styled_cell(ws, 11, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 12-15: One row per demo item --
    for pi, iname in enumerate(DEMO_ITEMS):
        r = 12 + pi
        styled_cell(ws, r, 1, pi + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, iname, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # C: 시험수 = COUNTIFS(B, vehicle, D, item)
        f_cnt = f'=COUNTIFS({b_range},{cfg},{d_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # D: 완료 = COUNTIFS(B, vehicle, D, item, V, "11-*")
        f_done = f'=COUNTIFS({b_range},{cfg},{d_range},B{r},{v_range},"11-*")'
        styled_cell(ws, r, 4, f_done, FONT_BODY, FILL_GREEN_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # F: 지연 = CRITICAL+OVERDUE+WARNING
        f_delay = (
            f'=COUNTIFS({b_range},{cfg},{d_range},B{r},{y_range},"CRITICAL")'
            f'+COUNTIFS({b_range},{cfg},{d_range},B{r},{y_range},"OVERDUE")'
            f'+COUNTIFS({b_range},{cfg},{d_range},B{r},{y_range},"WARNING")'
        )
        styled_cell(ws, r, 6, f_delay, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # E: 진행중 = Total - 완료 - 지연
        f_prog = f'=C{r}-D{r}-F{r}'
        styled_cell(ws, r, 5, f_prog, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # G: 진행률 = AVERAGEIFS
        f_pct = f'=IFERROR(AVERAGEIFS({w_range},{b_range},{cfg},{d_range},B{r}),0)'
        styled_cell(ws, r, 7, f_pct, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # H: 최소잔여일
        f_min = f'=IF(COUNTIFS({b_range},{cfg},{d_range},B{r},{y_range},"<>DONE")=0,"—",IFERROR(MINIFS({x_range},{b_range},{cfg},{d_range},B{r},{y_range},"<>DONE"),"—"))'
        styled_cell(ws, r, 8, f_min, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Conditional formatting on Item Progress
    ws.conditional_formatting.add(
        "F12:F15",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        "H12:H15",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        "H12:H15",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["14"],
            fill=PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_ORANGE),
        )
    )

    # -- Row 16: blank --
    for col in range(1, 9):
        styled_cell(ws, 16, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # ========================================
    # Section 3: Cross-Vehicle Comparison (rows 17-22)
    # ========================================
    merge_and_style(ws, 17, 1, 17, 8,
                    "Cross-Vehicle Comparison 차종간 비교",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 18: Headers --
    xv_headers = ["", "차종", "시험수", "완료", "지연", "진행률", "", ""]
    for i, h in enumerate(xv_headers):
        styled_cell(ws, 18, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 19-21: One row per vehicle --
    for vi, vcode in enumerate(VEHICLE_ORDER):
        r = 19 + vi
        styled_cell(ws, r, 1, vi + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, vcode, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # C: 시험수
        f_cnt = f'=COUNTIF({b_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # D: 완료
        f_done = f'=COUNTIFS({b_range},B{r},{v_range},"11-*")'
        styled_cell(ws, r, 4, f_done, FONT_BODY, FILL_GREEN_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # E: 지연
        f_delay = (
            f'=COUNTIFS({b_range},B{r},{y_range},"CRITICAL")'
            f'+COUNTIFS({b_range},B{r},{y_range},"OVERDUE")'
            f'+COUNTIFS({b_range},B{r},{y_range},"WARNING")'
        )
        styled_cell(ws, r, 5, f_delay, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # F: 진행률
        f_pct = f'=IFERROR(AVERAGEIFS({w_range},{b_range},B{r}),0)'
        styled_cell(ws, r, 6, f_pct, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        for col in [7, 8]:
            styled_cell(ws, r, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Row 22: TOTAL row
    r = 22
    styled_cell(ws, r, 1, "", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 2, "TOTAL", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 3, "=SUM(C19:C21)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 4, "=SUM(D19:D21)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 5, "=SUM(E19:E21)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 6, f'=IFERROR(AVERAGEIFS({w_range},{b_range},"<>"),0)',
                FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER, number_format="0%")
    for col in [7, 8]:
        styled_cell(ws, r, col, "", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)

    # Conditional formatting on 지연 column
    ws.conditional_formatting.add(
        "E19:E22",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )

    # -- Row 23: blank --
    for col in range(1, 9):
        styled_cell(ws, 23, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # ========================================
    # Section 4: Status Distribution (rows 24-37) — 12 status codes
    # ========================================
    merge_and_style(ws, 24, 1, 24, 8,
                    "Status Distribution 상태 분포 (Selected Vehicle)",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 25: Headers --
    sd_headers = ["", "상태", "건수", "", "", "", "", ""]
    for i, h in enumerate(sd_headers):
        styled_cell(ws, 25, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 26-37: 12 status rows --
    status_codes = [
        "01-미제출",
        "02-제출_미검토",
        "03-검토중",
        "04-검토중_보완",
        "05-계획서승인",
        "06-BIW중단",
        "07-미인수",
        "08-단품준비완료",
        "09-시험준비중",
        "10-시험완료",
        "11-결과등록완료",
        "12-조건부승인",
    ]
    for si, scode in enumerate(status_codes):
        r = 26 + si
        styled_cell(ws, r, 1, si + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, scode, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        f_cnt = f'=COUNTIFS({b_range},{cfg},{v_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        for col in range(4, 9):
            styled_cell(ws, r, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # -- Row 38: blank --
    for col in range(1, 9):
        styled_cell(ws, 38, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # ========================================
    # Section 5: Risk Distribution (rows 39-47) — 7 risk flags
    # ========================================
    merge_and_style(ws, 39, 1, 39, 8,
                    "Risk Distribution 리스크 분포 (Selected Vehicle)",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 40: Headers --
    rd_headers = ["", "리스크", "건수", "", "", "", "", ""]
    for i, h in enumerate(rd_headers):
        styled_cell(ws, 40, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 41-47: 7 risk flags --
    risk_flags = [
        "DONE",
        "ON TRACK",
        "WARNING",
        "OVERDUE",
        "CRITICAL",
        "BLOCKED",
        "ACTION",
    ]
    for ri, rflag in enumerate(risk_flags):
        r = 41 + ri
        styled_cell(ws, r, 1, ri + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, rflag, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        f_cnt = f'=COUNTIFS({b_range},{cfg},{y_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        for col in range(4, 9):
            styled_cell(ws, r, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Conditional formatting for risk distribution count
    # Highlight CRITICAL/OVERDUE/BLOCKED rows with red if count > 0
    for risk_row in [43, 44, 46]:  # OVERDUE=43, CRITICAL=44, BLOCKED=46
        ws.conditional_formatting.add(
            f"C{risk_row}:C{risk_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
                font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
            )
        )

    return ws


# ================================================================
# Sheet 5: Supplier Checklist
# ================================================================

def build_supplier_checklist(wb, tracker_sheet_name, data_start, data_end):
    ws = wb.create_sheet("Supplier Checklist 협력사체크리스트")
    ws.sheet_properties.tabColor = C_ORANGE

    ts = f"'{tracker_sheet_name}'"

    set_col_widths(ws, {
        "A": 4, "B": 18, "C": 18, "D": 14, "E": 20,
        "F": 22, "G": 14, "H": 14, "I": 30,
    })

    # Title
    merge_and_style(ws, 1, 1, 2, 9,
                    "Supplier Checklist 협력사 체크리스트",
                    FONT_TITLE, FILL_NAVY)

    # Headers (9 cols: No, Supplier, Item, Vehicle, Test, Status, Progress, Risk, NextAction)
    sc_headers = ["No.", "Supplier", "Item Name", "Vehicle", "Test Name",
                  "Status", "Progress", "Risk", "Next Action"]
    write_table_header(ws, 3, 1, sc_headers)

    # Placeholder rows referencing tracker data
    for i in range(20):
        r = 4 + i
        tr = data_start + i
        if tr > data_end:
            break

        styled_cell(ws, r, 1, i + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # E (col 5 in tracker) = Supplier
        f_sup = f"={ts}!E{tr}"
        styled_cell(ws, r, 2, f_sup, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # D (col 4 in tracker) = Item Name
        f_item = f"={ts}!D{tr}"
        styled_cell(ws, r, 3, f_item, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # B (col 2 in tracker) = Vehicle
        f_veh = f"={ts}!B{tr}"
        styled_cell(ws, r, 4, f_veh, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # F (col 6 in tracker) = Test Name
        f_test = f"={ts}!F{tr}"
        styled_cell(ws, r, 5, f_test, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # V (col 22) = Status
        f_status = f"={ts}!V{tr}"
        styled_cell(ws, r, 6, f_status, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # W (col 23) = Progress
        f_prog = f"={ts}!W{tr}"
        styled_cell(ws, r, 7, f_prog, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # Y (col 25) = Risk
        f_risk = f"={ts}!Y{tr}"
        styled_cell(ws, r, 8, f_risk, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # AA (col 27) = Next Action
        f_next = f"={ts}!AA{tr}"
        styled_cell(ws, r, 9, f_next, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

    return ws


# ================================================================
# Main
# ================================================================

def main():
    wb = Workbook()
    build_config(wb)
    build_reference_data(wb)
    ws_tracker, data_start, data_end, num_items = build_master_tracker(wb)
    build_dashboard(wb, ws_tracker.title, data_start, TRACKER_END, num_items)
    build_supplier_checklist(wb, ws_tracker.title, data_start, TRACKER_END)
    wb.active = wb.sheetnames.index("Dashboard 대시보드")
    wb.save(OUTPUT)
    print(f"V4 Generated: {OUTPUT}")
    print(f"  - {num_items} tracker rows (data rows {data_start}-{data_start + num_items - 1})")
    print(f"  - Empty formula rows {data_start + num_items}-{TRACKER_END}")
    print(f"  - 28 columns (A-AB), 5 sheets")
    print(f"  - 12 items, 18 tests, 4 reference tables")


if __name__ == "__main__":
    main()
