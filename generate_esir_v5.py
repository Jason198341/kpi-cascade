"""
ESIR Master Tracker V5 — Per-Vehicle Excel Generator
One file per vehicle. 2 Reference sheets + 1 Tracker.
Simplified flat structure, no cross-vehicle dashboard.

Usage:
  python generate_esir_v5.py           # generates all 3 vehicles
  python generate_esir_v5.py AE_PE     # generates one vehicle
"""

import sys
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Constants ──
OUT_DIR = r"C:\Users\USER\kpi-cascade"
TODAY = datetime.date(2026, 2, 27)
TRACKER_END = 80  # max rows for formulas/CF

# ── Styles ──
THIN = Side(style="thin", color="BDC3C7")
MED = Side(style="medium", color="1B2A4A")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER = Border(left=MED, right=MED, top=MED, bottom=MED)

FONT_TITLE    = Font(name="Malgun Gothic", size=16, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
FONT_HEADER   = Font(name="Malgun Gothic", size=9, bold=True, color="FFFFFF")
FONT_KR       = Font(name="Malgun Gothic", size=8, bold=True, color="1B2A4A")
FONT_BODY     = Font(name="Malgun Gothic", size=10)
FONT_BOLD     = Font(name="Malgun Gothic", size=10, bold=True)
FONT_SMALL    = Font(name="Malgun Gothic", size=9)
FONT_KPI      = Font(name="Malgun Gothic", size=20, bold=True, color="1B2A4A")
FONT_RED      = Font(name="Malgun Gothic", size=10, bold=True, color="E74C3C")

FILL_NAVY     = PatternFill("solid", fgColor="1B2A4A")
FILL_BLUE     = PatternFill("solid", fgColor="D6E4F0")
FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")
FILL_MANUAL   = PatternFill("solid", fgColor="FFF9C4")
FILL_AUTO     = PatternFill("solid", fgColor="F0F4F8")
FILL_GREEN    = PatternFill("solid", fgColor="D5F5E3")
FILL_RED      = PatternFill("solid", fgColor="FADBD8")
FILL_ORANGE   = PatternFill("solid", fgColor="FDEBD0")
FILL_GRAY     = PatternFill("solid", fgColor="F2F3F4")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def sc(ws, r, c, val, font=FONT_BODY, fill=FILL_WHITE, align=ALIGN_C, border=THIN_BORDER, fmt=None):
    """Style cell helper."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def merge_title(ws, r1, c1, r2, c2, val, font, fill):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font = font
    cell.fill = fill
    cell.alignment = ALIGN_C
    cell.border = MED_BORDER
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = MED_BORDER


# ── Vehicle Data ──
VEHICLES = {
    "AE_PE": {
        "name": "AE PE (Creta EV)",
        "t0": datetime.date(2026, 1, 15),
        "proto": datetime.date(2026, 4, 15),
        "mc": datetime.date(2026, 6, 15),
        "lt_plan": 14, "lt_parts": 49, "lt_test": 63, "lt_report": 98,
    },
    "SU2i": {
        "name": "SU2i (Venue Next)",
        "t0": datetime.date(2026, 3, 1),
        "proto": datetime.date(2026, 6, 1),
        "mc": datetime.date(2026, 8, 15),
        "lt_plan": 14, "lt_parts": 49, "lt_test": 63, "lt_report": 98,
    },
    "EN_SUV": {
        "name": "EN SUV",
        "t0": datetime.date(2026, 2, 1),
        "proto": datetime.date(2026, 5, 1),
        "mc": datetime.date(2026, 7, 15),
        "lt_plan": 14, "lt_parts": 49, "lt_test": 63, "lt_report": 98,
    },
}

# Part-Test master data: (System, PartName, ESSpec, SpecTitle, Cat, Duration,
#                          Supplier, SupplierContact, SurrSupplier, SurrContact, BIWContact, Budget)
PART_TESTS = [
    ("Console", "센터콘솔",   "ES 95400-05", "EMC 시험",     "B", 14, "HS R&A",          "박민수 010-1234-0001", "Seoyon",       "김태호 010-2345-0001", "이정우 010-3456-0001", "✓"),
    ("Console", "센터콘솔",   "MS-300-57",   "VOC 방출 시험", "C", 14, "HS R&A",          "박민수 010-1234-0001", "",              "",                     "",                     "✓"),
    ("Console", "센터콘솔",   "ES-90500",    "BSR 시험",     "A", 14, "HS R&A",          "박민수 010-1234-0001", "Hyundai Mobis", "장현석 010-2345-0002", "이정우 010-3456-0001", "✓"),
    ("Console", "센터콘솔",   "ES-90101",    "치수 검사",    "C",  3, "HS R&A",          "박민수 010-1234-0001", "",              "",                     "",                     "✓"),
    ("Console", "콘솔리드",   "FMVSS 302",   "연소성 시험",  "C",  3, "HS R&A",          "박민수 010-1234-0001", "",              "",                     "",                     "✓"),
    ("Console", "콘솔리드",   "ES-92030",    "힌지 내구 시험","C", 10, "HS R&A",          "박민수 010-1234-0001", "",              "",                     "",                     "미처리"),
    ("Console", "콘솔리드",   "ES-91300",    "조작력 시험",  "C",  5, "HS R&A",          "박민수 010-1234-0001", "",              "",                     "",                     "✓"),
    ("Cockpit", "칵핏모듈",   "ES 95400-05", "EMC 시험",     "A", 14, "Hyundai Mobis",   "장현석 010-2345-0002", "HS R&A",       "박민수 010-1234-0001", "이정우 010-3456-0001", "✓"),
    ("Cockpit", "칵핏모듈",   "ES-90500",    "BSR 시험",     "A", 14, "Hyundai Mobis",   "장현석 010-2345-0002", "Faurecia",     "David 010-2345-0003",  "이정우 010-3456-0001", "✓"),
    ("Trim",    "도어트림(앞)","MS-300-57",   "VOC 방출 시험","C", 14, "Faurecia",        "David 010-2345-0003",  "",              "",                     "",                     "✓"),
    ("Trim",    "도어트림(앞)","ES-90101",    "치수 검사",   "B",  3, "Faurecia",        "David 010-2345-0003",  "Seoyon",       "김태호 010-2345-0001", "",                     "✓"),
    ("Trim",    "도어트림(앞)","ES-91300",    "조작력 시험", "B",  5, "Faurecia",        "David 010-2345-0003",  "Seoyon",       "김태호 010-2345-0001", "",                     "미처리"),
    ("Trim",    "도어트림(뒤)","MS-300-57",   "VOC 방출 시험","C", 14, "Faurecia",        "David 010-2345-0003",  "",              "",                     "",                     "✓"),
    ("Trim",    "도어트림(뒤)","ES-90101",    "치수 검사",   "B",  3, "Faurecia",        "David 010-2345-0003",  "Seoyon",       "김태호 010-2345-0001", "",                     "✓"),
    ("Trim",    "필러트림 A", "MS-300-34",   "내후성 시험",  "C", 21, "Seoyon",          "김태호 010-2345-0001", "",              "",                     "",                     "✓"),
    ("Trim",    "필러트림 B", "MS-300-34",   "내후성 시험",  "C", 21, "Seoyon",          "김태호 010-2345-0001", "",              "",                     "",                     "✓"),
    ("Trim",    "필러트림 C", "MS-300-34",   "내후성 시험",  "C", 21, "Seoyon",          "김태호 010-2345-0001", "",              "",                     "",                     "✓"),
    ("Trim",    "패키지트레이","ECE R17",     "수하물 충격 시험","A", 7,"HS R&A",          "박민수 010-1234-0001", "Faurecia",     "David 010-2345-0003",  "이정우 010-3456-0001", "✓"),
    ("Trim",    "헤드라이너", "ES 80062",    "포깅 시험",    "C",  7, "Seoyon E-Hwa",    "유동관 010-2345-0004", "",              "",                     "",                     "✓"),
    ("Seat",    "시트(앞)",   "ES-93100",    "시트 내구 시험","A", 21, "Hyundai Transys",  "최진호 010-2345-0005", "HS R&A",       "박민수 010-1234-0001", "이정우 010-3456-0001", "✓"),
    ("Seat",    "시트(앞)",   "ES-93000",    "H-Point 검증", "A", 14, "Hyundai Transys",  "최진호 010-2345-0005", "Hyundai Mobis","장현석 010-2345-0002", "이정우 010-3456-0001", "미처리"),
    ("Seat",    "시트(앞)",   "FMVSS 210",   "안전벨트 앵커리지","A",14,"Hyundai Transys", "최진호 010-2345-0005", "Hyundai Mobis","장현석 010-2345-0002", "이정우 010-3456-0001", "✓"),
    ("Seat",    "시트(뒤)",   "ECE R14",     "ISOFIX 강도 시험","A",10,"Hyundai Transys",  "최진호 010-2345-0005", "Hyundai Mobis","장현석 010-2345-0002", "이정우 010-3456-0001", "✓"),
    ("Seat",    "시트(뒤)",   "MS-300-55",   "냄새 시험",    "C",  5, "Hyundai Transys",  "최진호 010-2345-0005", "",              "",                     "",                     "✓"),
]

# Sim status patterns (cycle through for demo data)
SIM_STATES = [
    # (k_sub, l_rev, p_act, q_act, r_act, s_sub, override)
    (None,  None,    None, None, None, None, None),                     # 01-미제출
    ("D",   None,    None, None, None, None, None),                     # 02-제출_미검토
    ("D",   "Under Review", None, None, None, None, None),              # 03-검토중
    ("D",   "Revision Needed", None, None, None, None, None),           # 04-검토중_보완
    ("D",   "Approved", None, None, None, None, None),                  # 05-계획서승인
    ("D",   "Approved", "D", None, None, None, None),                   # 08-단품준비완료
    ("D",   "Approved", "D", "D", None, None, None),                    # 09-시험준비중
    ("D",   "Approved", "D", "D", "D", None, None),                     # 10-시험완료
    ("D",   "Approved", "D", "D", "D", "D", None),                      # 11-결과등록완료
    ("D",   "Approved", "D", None, None, None, "06-BIW중단"),           # 06-BIW중단
    ("D",   "Approved", "D", None, None, None, "07-미인수"),            # 07-미인수
    ("D",   "Approved", "D", "D", "D", "D", "12-조건부승인"),           # 12-조건부승인
]


def build_ref_parttest(wb):
    """Sheet 1: Ref_PartTest — static part/test master data."""
    ws = wb.create_sheet("Ref_PartTest", 0)
    ws.sheet_properties.tabColor = "3498DB"

    widths = {"A": 12, "B": 16, "C": 16, "D": 18, "E": 6, "F": 8,
              "G": 18, "H": 24, "I": 16, "J": 24, "K": 24, "L": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header
    headers = ["System", "Part Name", "ES Spec No.", "Spec Title",
               "Cat", "Duration\n(days)", "담당 협력사", "협력사 담당자",
               "주변부품 협력사", "주변부품 담당자", "BIW 담당자", "예산처리"]
    for i, h in enumerate(headers):
        sc(ws, 1, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_C, MED_BORDER)

    # Data
    for i, pt in enumerate(PART_TESTS):
        r = i + 2
        for j, val in enumerate(pt):
            fill = FILL_WHITE
            if j == 4:  # Category
                fill = {"A": FILL_RED, "B": FILL_ORANGE, "C": FILL_GREEN}.get(val, FILL_WHITE)
            sc(ws, r, j + 1, val, FONT_BODY, fill,
               ALIGN_L if j in (1, 3, 7, 9, 10) else ALIGN_C, THIN_BORDER)

    ws.auto_filter.ref = f"A1:L{len(PART_TESTS) + 1}"
    ws.freeze_panes = "A2"
    return ws


def build_ref_schedule(wb, vcode, vdata):
    """Sheet 2: Ref_Schedule — milestone dates (one set for entire vehicle)."""
    ws = wb.create_sheet("Ref_Schedule", 1)
    ws.sheet_properties.tabColor = "27AE60"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30

    rows = [
        ("Vehicle Code", vcode, ""),
        ("Vehicle Name", vdata["name"], ""),
        ("T0 도면배포", vdata["t0"], "기준일"),
        ("계획서 기한", f"=B4+{vdata['lt_plan']}", f"T0 + {vdata['lt_plan']}일"),
        ("부품준비 기한", f"=B4+{vdata['lt_parts']}", f"T0 + {vdata['lt_parts']}일"),
        ("시험시작 기한", f"=B4+{vdata['lt_test']}", f"T0 + {vdata['lt_test']}일"),
        ("성적서 기한", f"=B4+{vdata['lt_report']}", f"T0 + {vdata['lt_report']}일"),
        ("Proto Build", vdata["proto"], ""),
        ("Master Car", vdata["mc"], ""),
        ("Test Deadline", "=B9-30", "MC - 30일"),
    ]

    sc(ws, 1, 1, "Milestone", FONT_HEADER, FILL_NAVY, ALIGN_C, MED_BORDER)
    sc(ws, 1, 2, "Date", FONT_HEADER, FILL_NAVY, ALIGN_C, MED_BORDER)
    sc(ws, 1, 3, "Note", FONT_HEADER, FILL_NAVY, ALIGN_C, MED_BORDER)

    for i, (label, val, note) in enumerate(rows):
        r = i + 2
        sc(ws, r, 1, label, FONT_BOLD, FILL_BLUE, ALIGN_L, THIN_BORDER)
        c = sc(ws, r, 2, val, FONT_BOLD, FILL_WHITE, ALIGN_C, THIN_BORDER)
        if isinstance(val, datetime.date):
            c.number_format = "YYYY-MM-DD"
        elif isinstance(val, str) and val.startswith("="):
            c.number_format = "YYYY-MM-DD"
        sc(ws, r, 3, note, FONT_SMALL, FILL_GRAY, ALIGN_L, THIN_BORDER)

    return ws


def build_tracker(wb, vcode, vdata):
    """Sheet 3: Tracker — main tracking with auto+manual columns."""
    ws = wb.create_sheet("Tracker", 2)
    ws.sheet_properties.tabColor = "E74C3C"

    REF = "'Ref_PartTest'"
    SCH = "'Ref_Schedule'"
    num_items = len(PART_TESTS)

    # Column widths (A-W = 23 cols)
    widths = {
        "A": 5, "B": 10, "C": 14, "D": 14, "E": 16, "F": 5, "G": 14, "H": 7,
        "I": 12, "J": 12, "K": 12, "L": 12, "M": 12,
        "N": 12, "O": 14, "P": 12, "Q": 12, "R": 12, "S": 12,
        "T": 18, "U": 7, "V": 8, "W": 22,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row 1-2: Title
    merge_title(ws, 1, 1, 2, 23, f"ESIR Tracker — {vcode} {vdata['name']}", FONT_TITLE, FILL_NAVY)

    # Row 3: EN headers
    en_h = [
        "No", "System", "Part Name", "ES Spec", "Spec Title", "Cat", "Supplier", "Budget",
        "Plan\nDue", "Parts\nDue", "Test\nStart Due", "Test\nEnd Due", "Report\nDue",
        "Plan\n\u2713", "Review\nResult", "Parts\n\u2713", "Test Start\n\u2713", "Test End\n\u2713", "Report\n\u2713",
        "Status", "Progress", "D-Day", "Next Action",
    ]
    # Manual columns (0-indexed): N(13), O(14), P(15), Q(16), R(17), S(18)
    manual_cols = {13, 14, 15, 16, 17, 18}

    for i, h in enumerate(en_h):
        f = FONT_HEADER
        fill = FILL_NAVY
        if i in manual_cols:
            fill = PatternFill("solid", fgColor="B7950B")
        sc(ws, 3, i + 1, h, f, fill, ALIGN_C, MED_BORDER)

    # Row 4: KR headers
    kr_h = [
        "번호", "시스템", "파트명", "ES 스펙", "시험명", "등급", "협력사", "예산",
        "계획서\n기한", "부품준비\n기한", "시험시작\n기한", "시험완료\n기한", "성적서\n기한",
        "계획서\n\u2713", "검토\n결과", "부품\n\u2713", "시험시작\n\u2713", "시험완료\n\u2713", "성적서\n\u2713",
        "상태", "진행률", "잔여일", "다음조치",
    ]
    for i, h in enumerate(kr_h):
        fill = FILL_BLUE
        if i in manual_cols:
            fill = PatternFill("solid", fgColor="FFF9C4")
            f = Font(name="Malgun Gothic", size=8, bold=True, color="7D6608")
        else:
            f = FONT_KR
        sc(ws, 4, i + 1, h, f, fill, ALIGN_C, THIN_BORDER)

    # Data rows (5 ~ 5+num_items-1)
    data_start = 5
    for idx in range(num_items):
        r = data_start + idx
        ref_r = idx + 2  # Ref_PartTest data starts at row 2
        sim = SIM_STATES[idx % len(SIM_STATES)]

        # A: No
        sc(ws, r, 1, idx + 1, FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # B-H: Auto from Ref_PartTest (INDEX)
        auto_cols = [
            (2, 1, ALIGN_C),   # B: System = Ref col A
            (3, 2, ALIGN_L),   # C: Part Name = Ref col B
            (4, 3, ALIGN_C),   # D: ES Spec = Ref col C
            (5, 4, ALIGN_L),   # E: Spec Title = Ref col D
            (6, 5, ALIGN_C),   # F: Category = Ref col E
            (7, 7, ALIGN_C),   # G: Supplier = Ref col G
            (8, 12, ALIGN_C),  # H: Budget = Ref col L
        ]
        for col, ref_col, al in auto_cols:
            f = f"=INDEX({REF}!{get_column_letter(ref_col)}:{get_column_letter(ref_col)},{ref_r})"
            sc(ws, r, col, f, FONT_BODY, FILL_AUTO, al, THIN_BORDER)

        # I-M: Schedule dates (auto from Ref_Schedule)
        # I: Plan Due = Ref_Schedule B5 (계획서 기한)
        sc(ws, r, 9, f"={SCH}!$B$5", FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")
        # J: Parts Due = Ref_Schedule B6 (부품준비 기한)
        sc(ws, r, 10, f"={SCH}!$B$6", FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")
        # K: Test Start Due = Ref_Schedule B7 (시험시작 기한)
        sc(ws, r, 11, f"={SCH}!$B$7", FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")
        # L: Test End Due = Test Start Due + Duration
        dur_ref = f"INDEX({REF}!F:F,{ref_r})"
        sc(ws, r, 12, f"=K{r}+{dur_ref}", FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")
        # M: Report Due = Ref_Schedule B8 (성적서 기한)
        sc(ws, r, 13, f"={SCH}!$B$8", FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")

        # N-S: Manual inputs — checkbox style (✓ or empty)
        k_sub, l_rev, p_act, q_act, r_act, s_sub, override = sim

        n_val = "\u2713" if k_sub == "D" else None
        o_val = l_rev                   # O: Review Result (text dropdown)
        p_val = "\u2713" if p_act == "D" else None
        q_val = "\u2713" if q_act == "D" else None
        r_val = "\u2713" if r_act == "D" else None
        s_val = "\u2713" if s_sub == "D" else None

        for col, val in [(14, n_val), (16, p_val), (17, q_val), (18, r_val), (19, s_val)]:
            sc(ws, r, col, val, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)

        # O: Review Result (text)
        sc(ws, r, 15, o_val, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)

        # Override (hidden column — we embed in status formula via a helper)
        # For sim, if override exists, we put it in a note or handle in formula
        # Simpler: add a hidden col or just hardcode override in status for demo
        # Let's keep it simple: no override column. Status formula handles 06/07/12 via review+actuals pattern

        # T: Status (AUTO) — simplified: no override column, pure date-driven
        if override:
            # For demo: directly write the override status
            sc(ws, r, 20, override, FONT_BOLD, FILL_WHITE, ALIGN_C, THIN_BORDER)
        else:
            f_t = (
                f'=IF(AND(S{r}<>"",R{r}<>""),"11-결과등록완료",'
                f'IF(AND(R{r}<>"",S{r}=""),"10-시험완료",'
                f'IF(AND(Q{r}<>"",R{r}=""),"09-시험준비중",'
                f'IF(AND(P{r}<>"",Q{r}=""),"08-단품준비완료",'
                f'IF(O{r}="Approved","05-계획서승인",'
                f'IF(O{r}="Revision Needed","04-검토중_보완",'
                f'IF(O{r}="Under Review","03-검토중",'
                f'IF(AND(N{r}<>"",O{r}=""),"02-제출_미검토",'
                f'"01-미제출"))))))))'
            )
            sc(ws, r, 20, f_t, FONT_BOLD, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # U: Progress % (AUTO) — nested IF (IFS not supported in Excel 2016)
        f_u = (
            f'=IF(T{r}="11-결과등록완료",1,'
            f'IF(T{r}="12-조건부승인",0.9,'
            f'IF(T{r}="10-시험완료",0.85,'
            f'IF(T{r}="09-시험준비중",0.55,'
            f'IF(T{r}="08-단품준비완료",0.4,'
            f'IF(T{r}="07-미인수",0.3,'
            f'IF(T{r}="05-계획서승인",0.2,'
            f'IF(T{r}="06-BIW중단",0.15,'
            f'IF(T{r}="03-검토중",0.1,'
            f'IF(T{r}="02-제출_미검토",0.05,'
            f'IF(T{r}="04-검토중_보완",0.05,'
            f'0)))))))))))'
        )
        sc(ws, r, 21, f_u, FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER, "0%")

        # V: D-Day (AUTO) — days to report due
        f_v = f'=IF(S{r}<>"","—",IF(M{r}="","",M{r}-TODAY()))'
        sc(ws, r, 22, f_v, FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # W: Next Action (AUTO)
        f_w = (
            f'=IF(T{r}="11-결과등록완료","— 완료",'
            f'IF(T{r}="01-미제출",IF(I{r}<TODAY(),"계획서 제출 독촉 (OVERDUE)","계획서 제출 대기"),'
            f'IF(T{r}="02-제출_미검토","검토 착수 필요",'
            f'IF(T{r}="03-검토중","승인 처리 대기",'
            f'IF(T{r}="04-검토중_보완","보완 → 재제출 독촉",'
            f'IF(T{r}="05-계획서승인","부품 준비 현황 확인",'
            f'IF(T{r}="06-BIW중단","BIW팀 확인 → 일정 재수립",'
            f'IF(T{r}="07-미인수","협력사 인수 독촉",'
            f'IF(T{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(T{r}="09-시험준비중","시험 모니터링",'
            f'IF(T{r}="10-시험완료","성적서 등록 독촉",'
            f'IF(T{r}="12-조건부승인","조건부 승인 판단","—"))))))))))))'
        )
        sc(ws, r, 23, f_w, FONT_SMALL, FILL_WHITE, ALIGN_L, THIN_BORDER)

    # Empty rows with formulas (for future entries)
    data_end = data_start + num_items - 1
    for r in range(data_end + 1, TRACKER_END + 1):
        # A: No
        sc(ws, r, 1, f'=IF(C{r}="","",ROW()-{data_start - 1})', FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # B-H: blank manual or formula
        for col in [2, 3, 4, 5]:
            sc(ws, r, col, None, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)
        # F, G, H: auto if C has value
        sc(ws, r, 6, None, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)
        sc(ws, r, 7, None, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)
        sc(ws, r, 8, None, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)

        # I-M: Schedule (same for all)
        for col, sch_row in [(9, 5), (10, 6), (11, 7), (13, 8)]:
            c = sc(ws, r, col, f'=IF(C{r}="","",{SCH}!$B${sch_row})', FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER)
            c.number_format = "YYYY-MM-DD"
        # L: Test End = K + duration (need manual duration input for new rows)
        sc(ws, r, 12, f'=IF(C{r}="","",K{r}+14)', FONT_BODY, FILL_AUTO, ALIGN_C, THIN_BORDER, "YYYY-MM-DD")

        # N-S: Manual blanks (checkbox style, no date format)
        for col in [14, 15, 16, 17, 18, 19]:
            sc(ws, r, col, None, FONT_BODY, FILL_MANUAL, ALIGN_C, THIN_BORDER)

        # T: Status
        f_t = (
            f'=IF(C{r}="","",'
            f'IF(AND(S{r}<>"",R{r}<>""),"11-결과등록완료",'
            f'IF(AND(R{r}<>"",S{r}=""),"10-시험완료",'
            f'IF(AND(Q{r}<>"",R{r}=""),"09-시험준비중",'
            f'IF(AND(P{r}<>"",Q{r}=""),"08-단품준비완료",'
            f'IF(O{r}="Approved","05-계획서승인",'
            f'IF(O{r}="Revision Needed","04-검토중_보완",'
            f'IF(O{r}="Under Review","03-검토중",'
            f'IF(AND(N{r}<>"",O{r}=""),"02-제출_미검토",'
            f'"01-미제출")))))))))'
        )
        sc(ws, r, 20, f_t, FONT_BOLD, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # U: Progress — nested IF (IFS not supported in Excel 2016)
        f_u = (
            f'=IF(C{r}="","",'
            f'IF(T{r}="11-결과등록완료",1,'
            f'IF(T{r}="12-조건부승인",0.9,'
            f'IF(T{r}="10-시험완료",0.85,'
            f'IF(T{r}="09-시험준비중",0.55,'
            f'IF(T{r}="08-단품준비완료",0.4,'
            f'IF(T{r}="07-미인수",0.3,'
            f'IF(T{r}="05-계획서승인",0.2,'
            f'IF(T{r}="06-BIW중단",0.15,'
            f'IF(T{r}="03-검토중",0.1,'
            f'IF(T{r}="02-제출_미검토",0.05,'
            f'IF(T{r}="04-검토중_보완",0.05,'
            f'0))))))))))))'
        )
        sc(ws, r, 21, f_u, FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER, "0%")

        # V: D-Day
        sc(ws, r, 22, f'=IF(C{r}="","",IF(S{r}<>"","—",IF(M{r}="","",M{r}-TODAY())))',
           FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER)

        # W: Next Action
        f_w = (
            f'=IF(C{r}="","",'
            f'IF(T{r}="11-결과등록완료","— 완료",'
            f'IF(T{r}="01-미제출",IF(I{r}<TODAY(),"계획서 제출 독촉","계획서 제출 대기"),'
            f'IF(T{r}="02-제출_미검토","검토 착수 필요",'
            f'IF(T{r}="03-검토중","승인 처리 대기",'
            f'IF(T{r}="04-검토중_보완","보완 → 재제출 독촉",'
            f'IF(T{r}="05-계획서승인","부품 준비 현황 확인",'
            f'IF(T{r}="06-BIW중단","BIW팀 확인",'
            f'IF(T{r}="07-미인수","협력사 인수 독촉",'
            f'IF(T{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(T{r}="09-시험준비중","시험 모니터링",'
            f'IF(T{r}="10-시험완료","성적서 등록 독촉",'
            f'IF(T{r}="12-조건부승인","조건부 승인 판단","—")))))))))))))'
        )
        sc(ws, r, 23, f_w, FONT_SMALL, FILL_WHITE, ALIGN_L, THIN_BORDER)

    # ── Data Validations ──
    dv_review = DataValidation(type="list", formula1='"Under Review,Approved,Revision Needed"', allow_blank=True)
    ws.add_data_validation(dv_review)
    dv_review.add(f"O5:O{TRACKER_END}")

    # Checkbox ✓ dropdown for N, P, Q, R, S
    dv_check = DataValidation(type="list", formula1='"\u2713"', allow_blank=True)
    dv_check.prompt = "Click to mark as done"
    dv_check.promptTitle = "\u2713"
    ws.add_data_validation(dv_check)
    for check_col in ["N", "P", "Q", "R", "S"]:
        dv_check.add(f"{check_col}5:{check_col}{TRACKER_END}")

    # ── Conditional Formatting ──

    # OVERDUE: due date < TODAY() and check cell empty → red fill
    overdue_pairs = [
        ("N", "I"),  # Plan ✓ vs Plan Due
        ("P", "J"),  # Parts ✓ vs Parts Due
        ("Q", "K"),  # Test Start ✓ vs Test Start Due
        ("R", "L"),  # Test End ✓ vs Test End Due
        ("S", "M"),  # Report ✓ vs Report Due
    ]
    for check_col, due_col in overdue_pairs:
        rng = f"{check_col}5:{check_col}{TRACKER_END}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({check_col}5="",{due_col}5<TODAY(),{due_col}5<>"")'],
            fill=PatternFill(bgColor="E74C3C"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF"),
        ))

    sr = f"T5:T{TRACKER_END}"
    status_colors = [
        ("01-미제출",     "BDC3C7", "2C3E50"),
        ("02-제출_미검토", "D6E4F0", "2C3E50"),
        ("03-검토중",     "D6E4F0", "2C3E50"),
        ("04-검토중_보완", "FADBD8", "C0392B"),
        ("05-계획서승인",  "FEF9E7", "B7950B"),
        ("06-BIW중단",    "E74C3C", "FFFFFF"),
        ("07-미인수",     "FDEBD0", "E67E22"),
        ("08-단품준비완료", "FEF9E7", "B7950B"),
        ("09-시험준비중",  "AED6F1", "2471A3"),
        ("10-시험완료",   "ABEBC6", "27AE60"),
        ("11-결과등록완료","27AE60", "FFFFFF"),
        ("12-조건부승인",  "FDEBD0", "E67E22"),
    ]
    for code, bg, fg in status_colors:
        ws.conditional_formatting.add(sr, CellIsRule(
            operator="equal", formula=[f'"{code}"'],
            fill=PatternFill(bgColor=bg),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=fg),
        ))

    # Progress data bar
    ws.conditional_formatting.add(f"U5:U{TRACKER_END}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1,
        color="3498DB"
    ))

    # D-Day red/yellow
    ws.conditional_formatting.add(f"V5:V{TRACKER_END}", CellIsRule(
        operator="lessThanOrEqual", formula=["0"],
        fill=PatternFill(bgColor="FADBD8"),
        font=Font(name="Malgun Gothic", size=10, bold=True, color="E74C3C"),
    ))
    ws.conditional_formatting.add(f"V5:V{TRACKER_END}", CellIsRule(
        operator="lessThanOrEqual", formula=["14"],
        fill=PatternFill(bgColor="FEF9E7"),
        font=Font(name="Malgun Gothic", size=10, bold=True, color="B7950B"),
    ))

    # Category coloring
    for cat, color in [("A", "FADBD8"), ("B", "FDEBD0"), ("C", "D5F5E3")]:
        ws.conditional_formatting.add(f"F5:F{TRACKER_END}", CellIsRule(
            operator="equal", formula=[f'"{cat}"'],
            fill=PatternFill(bgColor=color),
            font=Font(name="Malgun Gothic", size=10, bold=True),
        ))

    # ── Summary row (below all tracker rows) ──
    kr = TRACKER_END + 2
    sc(ws, kr, 1, "", FONT_BODY, FILL_WHITE, ALIGN_C)
    kr += 1
    merge_title(ws, kr, 1, kr, 23, "KPI Summary", FONT_SUBTITLE, FILL_NAVY)
    kr += 1
    labels = ["Total", "완료", "진행중", "미착수", "지연(D-Day≤0)", "평균 진행률"]
    for i, lb in enumerate(labels):
        sc(ws, kr, 1 + i * 2, lb, FONT_BOLD, FILL_BLUE, ALIGN_C, THIN_BORDER)

    kr += 1
    ds, de = data_start, data_end
    t_rng = f"T{ds}:T{de}"
    v_rng = f"V{ds}:V{de}"
    u_rng = f"U{ds}:U{de}"

    formulas = [
        f"=COUNTA(T{ds}:T{de})-COUNTBLANK(T{ds}:T{de})",
        f'=COUNTIF({t_rng},"11-결과등록완료")',
        f'=COUNTA({t_rng})-COUNTBLANK({t_rng})-COUNTIF({t_rng},"11-결과등록완료")-COUNTIF({t_rng},"01-미제출")',
        f'=COUNTIF({t_rng},"01-미제출")',
        f'=SUMPRODUCT((ISNUMBER({v_rng}))*({v_rng}<=0))',
        f'=IFERROR(AVERAGE({u_rng}),0)',
    ]
    fmts = [None, None, None, None, None, "0%"]

    for i, (fml, fmt) in enumerate(zip(formulas, fmts)):
        sc(ws, kr, 1 + i * 2, fml, FONT_KPI, FILL_WHITE, ALIGN_C, MED_BORDER, fmt)

    # ── Freeze & Filter ──
    ws.auto_filter.ref = f"A4:W{TRACKER_END}"
    ws.freeze_panes = "A5"

    return ws


def generate_vehicle(vcode):
    vdata = VEHICLES[vcode]
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_ref_parttest(wb)
    build_ref_schedule(wb, vcode, vdata)
    build_tracker(wb, vcode, vdata)

    # Set Tracker as active
    wb.active = wb.sheetnames.index("Tracker")

    path = f"{OUT_DIR}\\ESIR_{vcode}.xlsx"
    wb.save(path)
    print(f"  ✓ {path} ({len(PART_TESTS)} items)")
    return path


def main():
    targets = sys.argv[1:] if len(sys.argv) > 1 else list(VEHICLES.keys())
    print(f"ESIR V5 Generator — {len(targets)} vehicle(s)")
    for vcode in targets:
        if vcode not in VEHICLES:
            print(f"  ✗ Unknown vehicle: {vcode}")
            continue
        generate_vehicle(vcode)
    print("Done.")


if __name__ == "__main__":
    main()
