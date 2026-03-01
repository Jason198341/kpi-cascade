"""
ESIR Master Tracker V4 — Validation Script
Validates the generated Excel file for:
  1. Sheet structure (5 sheets, 4 reference tables)
  2. Column headers match V4 spec (28 cols A-AB)
  3. Status simulation: compute expected status from manual fields, verify formula logic
  4. Formula syntax (parentheses balance) for V, Y, AA columns
  5. Conditional formatting rule count
  6. Per-row simulation report: expected Status, Risk, NextAction
  7. Alarm gap detection (formula structure check)
"""

import sys
import io
import datetime
from openpyxl import load_workbook

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

FILE = r"C:\Users\USER\kpi-cascade\ESIR_Master_Tracker_v4.xlsx"
TODAY = datetime.date(2026, 2, 27)

ALL_STATES = [
    "01-미제출", "02-제출_미검토", "03-검토중",
    "04-검토중_보완", "05-계획서승인", "06-BIW중단",
    "07-미인수", "08-단품준비완료", "09-시험준비중",
    "10-시험완료", "11-결과등록완료", "12-조건부승인",
]

# V4 Column mapping (0-indexed for iter_rows, 28 cols A-AB)
COL = {
    "A_no": 0,              # A: 번호
    "B_vehicle": 1,         # B: 차종
    "C_partcat": 2,         # C: 파트구분 (AUTO)
    "D_item": 3,            # D: 아이템명
    "E_supplier": 4,        # E: 협력사 (AUTO)
    "F_test": 5,            # F: 시험명
    "G_testtype": 6,        # G: 시험유형 (AUTO)
    "H_scope": 7,           # H: 범위 (AUTO)
    "I_t0": 8,              # I: T0 도면배포 (AUTO)
    "J_plandue": 9,         # J: 계획서기한 (AUTO)
    "K_plansub": 10,        # K: 계획서제출 (MANUAL)
    "L_review": 11,         # L: 검토결과 (MANUAL)
    "M_partsplan": 12,      # M: 부품준비(계획) (AUTO)
    "N_partsact": 13,       # N: 부품준비(실적) (MANUAL)
    "O_teststartplan": 14,  # O: 시험시작(계획) (AUTO)
    "P_teststartact": 15,   # P: 시험시작(실적) (MANUAL)
    "Q_testcompplan": 16,   # Q: 시험완료(계획) (AUTO)
    "R_testcompact": 17,    # R: 시험완료(실적) (MANUAL)
    "S_reportdue": 18,      # S: 성적서기한 (AUTO)
    "T_reportsub": 19,      # T: 성적서제출 (MANUAL)
    "U_override": 20,       # U: 오버라이드 (MANUAL)
    "V_status": 21,         # V: 상태 (AUTO)
    "W_progress": 22,       # W: 진행률% (AUTO)
    "X_days": 23,           # X: 잔여일 (AUTO)
    "Y_risk": 24,           # Y: 리스크 (AUTO)
    "Z_stagnant": 25,       # Z: 정체일수 (AUTO)
    "AA_action": 26,        # AA: 다음조치 (AUTO)
    "AB_note": 27,          # AB: 비고 (MANUAL)
}

EXPECTED_HEADERS_KR = [
    "번호", "차종", "파트구분", "아이템명", "협력사", "시험명", "시험유형", "범위",
    "T₀ 도면배포", "계획서기한", "계획서제출", "검토결과",
    "부품준비(계획)", "부품준비(실적)", "시험시작(계획)", "시험시작(실적)",
    "시험완료(계획)", "시험완료(실적)", "성적서기한", "성적서제출",
    "오버라이드", "상태", "진행률%", "잔여일", "리스크", "정체일", "다음조치", "비고",
]

EXPECTED_TABLES = [
    "TBL_Vehicle", "TBL_TestCatalog", "TBL_ItemMaster", "TBL_ItemTests",
]


def check_parens(formula_str):
    if not formula_str or not isinstance(formula_str, str):
        return True
    return formula_str.count("(") == formula_str.count(")")


def compute_expected_status(k_sub, l_rev, n_act, p_act, r_act, t_sub, u_ovr):
    """Simulate the V (Status) formula from manual field values."""
    if u_ovr and isinstance(u_ovr, str) and u_ovr.strip():
        return u_ovr
    if t_sub and r_act:
        return "11-결과등록완료"
    if r_act and not t_sub:
        return "10-시험완료"
    if p_act and not r_act:
        return "09-시험준비중"
    if n_act and not p_act:
        return "08-단품준비완료"
    if l_rev == "Approved":
        return "05-계획서승인"
    if l_rev == "Revision Needed":
        return "04-검토중_보완"
    if l_rev == "Under Review":
        return "03-검토중"
    if k_sub and (not l_rev or l_rev == ""):
        return "02-제출_미검토"
    return "01-미제출"


def compute_expected_risk(status, plan_due, days_to_deadline, u_ovr):
    """Simulate the Y (Risk) formula."""
    if status == "11-결과등록완료":
        return "DONE"
    if u_ovr == "06-BIW중단":
        return "BLOCKED"
    if status == "01-미제출" and plan_due and plan_due < TODAY:
        return "CRITICAL"
    if days_to_deadline is not None and isinstance(days_to_deadline, (int, float)):
        if days_to_deadline <= 0:
            return "OVERDUE"
        if days_to_deadline <= 14:
            return "WARNING"
    if u_ovr and isinstance(u_ovr, str) and u_ovr.strip():
        return "ACTION"
    return "ON TRACK"


def compute_expected_action(status, plan_due):
    """Simulate the AA (NextAction) formula."""
    actions = {
        "11-결과등록완료": "— 완료",
        "02-제출_미검토": "설계자: 계획서 검토 착수",
        "03-검토중": "설계자: 승인 처리",
        "04-검토중_보완": "보완 요청 → 협력사 재제출 독촉",
        "05-계획서승인": "시험편 준비 현황 확인",
        "06-BIW중단": "BIW 생산팀 확인 → 일정 재수립",
        "07-미인수": "협력사 인수 독촉",
        "08-단품준비완료": "시험 일정 확인",
        "09-시험준비중": "시험 진행 모니터링",
        "10-시험완료": "결과 보고서 전산 등록 독촉",
        "12-조건부승인": "조건부 승인 여부 판단 → 관련부서 협의",
    }
    if status == "01-미제출":
        if plan_due and plan_due < TODAY:
            return "계획서 제출 독촉 (OVERDUE)"
        return "계획서 제출 대기"
    return actions.get(status, "—")


def to_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def is_filled(val):
    """Check if a cell has a real value (not None, not empty string, not formula)."""
    if val is None:
        return False
    if isinstance(val, str):
        if val.strip() == "" or val.startswith("="):
            return False
        return True
    return True  # date, number, etc.


def main():
    print(f"{'='*80}")
    print("ESIR Master Tracker V4 — Validation Report")
    print(f"File: {FILE}")
    print(f"{'='*80}\n")

    wb = load_workbook(FILE, data_only=False)

    # ── 0. Sheet structure ─────────────────────────────────────────────
    print("=" * 60)
    print("CHECK 0: Sheet structure")
    print("=" * 60)

    print(f"  Sheets found: {wb.sheetnames}")
    expected_sheets = ["Config", "Reference Data 기준데이터", "Master Tracker 마스터추적표",
                       "Dashboard 대시보드", "Supplier Checklist 협력사체크리스트"]
    for es in expected_sheets:
        found = es in wb.sheetnames
        print(f"  {'✓' if found else '✗'} {es}")

    # Check reference tables
    ref_ws = wb["Reference Data 기준데이터"]
    tables_found = [t.name for t in ref_ws.tables.values()]
    print(f"\n  Reference tables: {tables_found}")
    tables_ok = True
    for et in EXPECTED_TABLES:
        found = et in tables_found
        if not found:
            tables_ok = False
        print(f"  {'✓' if found else '✗'} {et}")

    # Find tracker sheet
    tracker_name = "Master Tracker 마스터추적표"
    if tracker_name not in wb.sheetnames:
        print("[FAIL] Could not find Master Tracker sheet!")
        return

    ws = wb[tracker_name]
    print(f"\n  Tracker dimensions: {ws.dimensions}")

    # ── 1. Column headers ──────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 1: Column headers (row 4 KR, 28 cols)")
    print("=" * 60)

    headers_ok = True
    for i, expected in enumerate(EXPECTED_HEADERS_KR):
        actual = ws.cell(row=4, column=i + 1).value
        match = actual == expected
        if not match:
            headers_ok = False
            print(f"  ✗ Col {chr(65+i) if i < 26 else 'A'+chr(65+i-26)}: expected '{expected}', got '{actual}'")
    if headers_ok:
        print(f"  [PASS] All 28 column headers match V4 spec")
    else:
        print(f"  [FAIL] Some headers don't match")

    # ── 2. Read data rows and compute expected statuses ────────────────
    print(f"\n{'='*60}")
    print("CHECK 2: Status simulation (12 states from manual fields)")
    print("=" * 60)

    DATA_START = 5
    rows_data = []
    statuses_found = set()

    for row in ws.iter_rows(min_row=DATA_START, max_col=28, values_only=False):
        a_val = row[COL["A_no"]].value
        # Stop at first truly empty row (no number in A)
        if a_val is None:
            break
        # Skip formula-only rows (empty rows with =IF(B...))
        if isinstance(a_val, str) and a_val.startswith("="):
            break

        vehicle = row[COL["B_vehicle"]].value
        item = row[COL["D_item"]].value
        test = row[COL["F_test"]].value
        k_sub = row[COL["K_plansub"]].value
        l_rev = row[COL["L_review"]].value
        n_act = row[COL["N_partsact"]].value
        p_act = row[COL["P_teststartact"]].value
        r_act = row[COL["R_testcompact"]].value
        t_sub = row[COL["T_reportsub"]].value
        u_ovr = row[COL["U_override"]].value

        # Compute expected status
        exp_status = compute_expected_status(
            k_sub if is_filled(k_sub) else None,
            l_rev if is_filled(l_rev) else None,
            n_act if is_filled(n_act) else None,
            p_act if is_filled(p_act) else None,
            r_act if is_filled(r_act) else None,
            t_sub if is_filled(t_sub) else None,
            u_ovr if is_filled(u_ovr) else None,
        )
        statuses_found.add(exp_status)

        # days_to_deadline and plan_due can't be computed without formula eval
        days_val = None
        plan_due = None

        rows_data.append({
            "row": row[0].row,
            "row_num": a_val,
            "vehicle": vehicle,
            "item": item,
            "test": test,
            "exp_status": exp_status,
            "plan_due": plan_due,
            "days_to_deadline": days_val,
            "u_ovr": u_ovr if is_filled(u_ovr) else None,
            "status_formula": row[COL["V_status"]].value,
            "risk_formula": row[COL["Y_risk"]].value,
            "action_formula": row[COL["AA_action"]].value,
        })

    # State coverage
    missing_states = set(ALL_STATES) - statuses_found
    state_counts = {s: sum(1 for rd in rows_data if rd["exp_status"] == s) for s in ALL_STATES}

    for s in ALL_STATES:
        marker = "✓" if s in statuses_found else "✗ MISSING"
        print(f"  {s:25s} count={state_counts.get(s,0):2d}  {marker}")

    if missing_states:
        print(f"\n  [FAIL] Missing states: {missing_states}")
    else:
        print(f"\n  [PASS] All 12 states covered by simulation data!")

    print(f"  Total data rows: {len(rows_data)}")

    # Vehicle distribution
    v_counts = {}
    for rd in rows_data:
        v = rd["vehicle"]
        v_counts[v] = v_counts.get(v, 0) + 1
    print(f"  Vehicle distribution: {v_counts}")

    # ── 3. Formula syntax ──────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 3: Formula syntax (parentheses balance)")
    print("=" * 60)

    formula_errors = 0
    formulas_checked = 0
    for rd in rows_data:
        for fname, fval in [("V Status", rd["status_formula"]),
                            ("Y Risk", rd["risk_formula"]),
                            ("AA Action", rd["action_formula"])]:
            if fval and isinstance(fval, str) and fval.startswith("="):
                formulas_checked += 1
                if not check_parens(fval):
                    print(f"  [FAIL] Row {rd['row']}: {fname} — unbalanced parens")
                    formula_errors += 1

    if formula_errors == 0:
        print(f"  [PASS] All {formulas_checked} formulas have balanced parentheses")
    else:
        print(f"  [FAIL] {formula_errors} formulas with unbalanced parentheses")

    # ── 4. Formula structure spot-check ────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 4: Formula structure spot-check (row 5)")
    print("=" * 60)

    structure_ok = True
    if rows_data:
        rd0 = rows_data[0]
        v_f = rd0["status_formula"] or ""
        y_f = rd0["risk_formula"] or ""
        aa_f = rd0["action_formula"] or ""

        # V (Status) formula checks — should reference U (override), T, R, P, N, L, K
        v_checks = [
            ("Override ref (U5)", "U5" in v_f),
            ("11-결과등록완료 literal", "11-결과등록완료" in v_f),
            ("01-미제출 literal", "01-미제출" in v_f),
            ("Approved check", "Approved" in v_f),
        ]
        for label, passed in v_checks:
            if not passed:
                structure_ok = False
            print(f"  {'✓' if passed else '✗'} V formula: {label}")

        # Y (Risk) formula checks
        y_checks = [
            ("BLOCKED literal", "BLOCKED" in y_f),
            ("DONE literal", "DONE" in y_f),
            ("CRITICAL literal", "CRITICAL" in y_f),
            ("OVERDUE literal", "OVERDUE" in y_f),
            ("WARNING literal", "WARNING" in y_f),
            ("ON TRACK literal", "ON TRACK" in y_f),
        ]
        for label, passed in y_checks:
            if not passed:
                structure_ok = False
            print(f"  {'✓' if passed else '✗'} Y formula: {label}")

        # AA (NextAction) formula checks
        aa_checks = [
            ("— 완료 literal", "완료" in aa_f),
            ("계획서 제출 대기", "대기" in aa_f),
            ("12-조건부승인 literal", "12-조건부승인" in aa_f),
        ]
        for label, passed in aa_checks:
            if not passed:
                structure_ok = False
            print(f"  {'✓' if passed else '✗'} AA formula: {label}")

    if structure_ok:
        print(f"\n  [PASS] All formula structure checks passed")
    else:
        print(f"\n  [WARN] Some formula structure checks failed — review formulas")

    # ── 5. Conditional formatting ──────────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 5: Conditional formatting rules")
    print("=" * 60)

    cf_rules = ws.conditional_formatting
    rule_count = len(list(cf_rules))
    print(f"  Tracker CF rules: {rule_count}")
    for cf in cf_rules:
        range_str = str(cf)
        rules_in_range = len(cf.rules)
        print(f"    {range_str:35s}  Rules: {rules_in_range}")

    # Dashboard CF
    dash_ws = wb["Dashboard 대시보드"] if "Dashboard 대시보드" in wb.sheetnames else None
    if dash_ws:
        dash_cf = len(list(dash_ws.conditional_formatting))
        print(f"  Dashboard CF rules: {dash_cf}")

    # ── 6. Per-row simulation report ───────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 6: Per-row simulation report (first 30 rows)")
    print("=" * 60)
    print(f"  {'Row':>4s} | {'Vehicle':8s} | {'Item':18s} | {'ExpStatus':22s} | {'ExpRisk':10s}")
    print(f"  {'-'*4} | {'-'*8} | {'-'*18} | {'-'*22} | {'-'*10}")

    for rd in rows_data[:30]:
        exp_risk = compute_expected_risk(
            rd["exp_status"], rd["plan_due"], rd["days_to_deadline"], rd["u_ovr"])
        print(f"  {rd['row_num']:>4} | {str(rd['vehicle'])[:8]:8s} | "
              f"{str(rd['item'])[:18]:18s} | {rd['exp_status']:22s} | {exp_risk}")

    if len(rows_data) > 30:
        print(f"  ... ({len(rows_data) - 30} more rows)")

    # ── 7. Alarm gap detection ─────────────────────────────────────────
    print(f"\n{'='*60}")
    print("CHECK 7: Alarm gap detection (formula content)")
    print("=" * 60)

    alarm_gaps = []
    for rd in rows_data:
        formula = rd["risk_formula"]
        if not formula or not isinstance(formula, str):
            continue

        # BLOCKED check (06-BIW중단)
        if rd["u_ovr"] == "06-BIW중단" and "BLOCKED" not in formula:
            alarm_gaps.append((rd["row"], rd["exp_status"], "Risk formula missing BLOCKED for BIW"))

        # CRITICAL check (01-미제출 with overdue plan)
        if rd["exp_status"] == "01-미제출" and "CRITICAL" not in formula:
            alarm_gaps.append((rd["row"], rd["exp_status"], "Risk formula missing CRITICAL check"))

    should_alarm = [rd for rd in rows_data
                    if compute_expected_risk(
                        rd["exp_status"], rd["plan_due"], rd["days_to_deadline"], rd["u_ovr"])
                    in ("CRITICAL", "OVERDUE", "ACTION", "BLOCKED")]
    for rd in should_alarm:
        exp_risk = compute_expected_risk(
            rd["exp_status"], rd["plan_due"], rd["days_to_deadline"], rd["u_ovr"])
        print(f"  Row {rd['row']:2d} | {rd['exp_status']:22s} | {exp_risk:10s} | ALARM")

    if alarm_gaps:
        print(f"\n  [WARN] Potential alarm gaps:")
        for row, status, msg in alarm_gaps:
            print(f"    Row {row}: {status} — {msg}")
    else:
        print(f"\n  [PASS] No alarm gaps detected")

    # ── Summary ────────────────────────────────────────────────────────
    print(f"\n{'='*80}")
    print("VALIDATION SUMMARY")
    print(f"{'='*80}")

    checks = [
        ("Sheet structure (5 sheets, 4 tables)", len(wb.sheetnames) == 5 and tables_ok),
        ("Column headers (28 KR)", headers_ok),
        (f"All 12 states in sim data", not missing_states),
        (f"Formula syntax ({formulas_checked} checked)", formula_errors == 0),
        ("Formula structure", structure_ok),
        (f"Data rows ({len(rows_data)})", len(rows_data) == 69),
        ("No alarm gaps", len(alarm_gaps) == 0),
    ]

    all_pass = True
    for label, passed in checks:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {label}")

    if all_pass:
        print(f"\n  >>> ALL CHECKS PASSED <<<")
    else:
        print(f"\n  >>> SOME CHECKS FAILED — review above <<<")


if __name__ == "__main__":
    main()
