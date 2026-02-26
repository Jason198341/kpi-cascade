"""
ESIR Master Tracker V3.1 -- Excel Generator
Automotive Interior Parts (Console / Trim / Seat)
Hyundai Motor India Engineering Sample / Initial Run Test Preparation

V3.2 Architecture (25 cols A-Y, 3 reference tables, no supplier):
  Sheet 1: Config (hidden helper)
  Sheet 2: Reference Data (SSOT: Vehicle + TestCatalog + PartMap)
  Sheet 3: Master Tracker (25 cols, formula-driven, G=manual)
  Sheet 4: Dashboard (COUNTIFS + vehicle filter)
  Sheet 5: Supplier Checklist

Reference Data (3 tables, 32 rows):
  - TBL_Vehicle (10 cols, 3 rows): Vehicle schedule + LeadTime offsets
  - TBL_TestCatalog (4 cols, 15 rows): Unique tests with type/spec/duration
  - TBL_PartMap (2 cols, 6 rows): Part→Category mapping
  Removed: TBL_VehicleParts (supplier info → col G manual)
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# == Constants ==
OUTPUT = r"C:\Users\USER\kpi-cascade\ESIR_Master_Tracker.xlsx"
TODAY = datetime.date(2026, 2, 26)
TRACKER_END = 200  # max data row for formulas/validations

# == Explicit cross-sheet range references for VLOOKUPs ==
REF = "'Reference Data 기준데이터'"
RNG_VEHICLE     = f"{REF}!$A$2:$J$5"       # 10 cols (schedule + 4 LeadTime offsets)
RNG_TESTCATALOG = f"{REF}!$A$8:$D$23"      # 15 unique tests, 4 cols
RNG_PARTMAP     = f"{REF}!$A$26:$B$32"     # 6 parts → category

# == Mapping Data (SSOT for table generation + sim data) ==
VEHICLE_ORDER = ["AE_PE", "SU2i", "EN_SUV"]
PART_NAMES = [
    "Center Console Assy", "Console Lid Assy", "Door Trim LH",
    "Pillar Trim C", "Front Seat Assy LH", "Rear Seat Assy",
]

T0_DATES = {
    "AE_PE":  datetime.date(2026, 1, 15),
    "SU2i":   datetime.date(2026, 3, 1),
    "EN_SUV": datetime.date(2026, 2, 1),
}

# Part -> Tests (fixed test set per part, human-readable names)
PART_TESTS = {
    "Center Console Assy": ["VOC 방출 시험", "내열 시험", "마모 시험", "치수 검사"],
    "Console Lid Assy": ["VOC 방출 시험", "힌지 내구 시험", "외관 검사"],
    "Door Trim LH": ["VOC 방출 시험", "내후성 시험", "연소성 시험", "외관 검사"],
    "Pillar Trim C": ["내후성 시험", "연소성 시험", "치수 검사"],
    "Front Seat Assy LH": ["시트 내구 시험", "안전벨트 앵커리지", "H-Point 검증", "BSR 시험"],
    "Rear Seat Assy": ["시트 내구 시험", "ISOFIX 강도 시험", "수하물 충격 시험", "외관 검사"],
}

# Test catalog: TestNameKR -> (TestType, Spec, DurationDays)
TEST_CATALOG = {
    "VOC 방출 시험": ("단품 Component", "MS-300-57", 14),
    "내후성 시험": ("단품 Component", "MS-300-34", 21),
    "내열 시험": ("단품 Component", "MS-300-31", 14),
    "마모 시험": ("단품 Component", "MS-300-35", 7),
    "힌지 내구 시험": ("단품 Component", "ES-92030", 10),
    "시트 내구 시험": ("단품 Component", "ES-93100", 21),
    "연소성 시험": ("단품 Component", "FMVSS 302", 3),
    "안전벨트 앵커리지": ("차량장착 Vehicle", "FMVSS 210", 14),
    "ISOFIX 강도 시험": ("차량장착 Vehicle", "ECE R14", 10),
    "암레스트 하중": ("단품 Component", "ES-91200", 7),
    "치수 검사": ("단품 Component", "ES-90101", 3),
    "H-Point 검증": ("차량장착 Vehicle", "ES-93000", 14),
    "BSR 시험": ("차량장착 Vehicle", "ES-90500", 14),
    "외관 검사": ("단품 Component", "ES-90100", 3),
    "수하물 충격 시험": ("차량장착 Vehicle", "ECE R17", 7),
}

# Part category mapping
PART_CATEGORIES = {
    "Center Console Assy": "Console",
    "Console Lid Assy": "Console",
    "Door Trim LH": "Trim",
    "Pillar Trim C": "Trim",
    "Front Seat Assy LH": "Seat",
    "Rear Seat Assy": "Seat",
}

# 22 target states per vehicle (matches 22 Part x Test combos, all 12 states covered)
STATE_PATTERN = [
    1, 5, 9, 11,       # Center Console Assy (4 tests)
    2, 8, 11,           # Console Lid Assy (3 tests)
    3, 10, 6, 11,       # Door Trim LH (4 tests)
    4, 9, 11,           # Pillar Trim C (3 tests)
    7, 5, 9, 10,        # Front Seat Assy LH (4 tests)
    8, 12, 11, 11,      # Rear Seat Assy (4 tests)
]

# == Color Palette ==
C_DARK_NAVY    = "1B2A4A"
C_NAVY         = "2C3E6B"
C_MID_BLUE     = "3B5998"
C_LIGHT_BLUE   = "D6E4F0"
C_VERY_LIGHT   = "EBF1F8"
C_WHITE        = "FFFFFF"
C_RED          = "E74C3C"
C_RED_LIGHT    = "FADBD8"
C_ORANGE       = "F39C12"
C_ORANGE_LIGHT = "FDEBD0"
C_YELLOW       = "F1C40F"
C_YELLOW_LIGHT = "FEF9E7"
C_GREEN        = "27AE60"
C_GREEN_LIGHT  = "D5F5E3"
C_GRAY         = "BDC3C7"
C_GRAY_LIGHT   = "F2F3F4"
C_CHARCOAL     = "2C3E50"
C_BLACK        = "000000"
C_LIGHT_GREEN  = "ABEBC6"

C_MANUAL_YELLOW = "FFF9C4"

# == Reusable Styles ==
THIN_BORDER = Border(
    left=Side(style="thin", color=C_GRAY),
    right=Side(style="thin", color=C_GRAY),
    top=Side(style="thin", color=C_GRAY),
    bottom=Side(style="thin", color=C_GRAY),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color=C_NAVY),
    right=Side(style="medium", color=C_NAVY),
    top=Side(style="medium", color=C_NAVY),
    bottom=Side(style="medium", color=C_NAVY),
)

FONT_TITLE      = Font(name="Malgun Gothic", size=18, bold=True, color=C_WHITE)
FONT_SUBTITLE   = Font(name="Malgun Gothic", size=13, bold=True, color=C_WHITE)
FONT_HEADER     = Font(name="Malgun Gothic", size=10, bold=True, color=C_WHITE)
FONT_HEADER_KR  = Font(name="Malgun Gothic", size=9, bold=True, color=C_DARK_NAVY)
FONT_BODY       = Font(name="Malgun Gothic", size=10, color=C_CHARCOAL)
FONT_BODY_BOLD  = Font(name="Malgun Gothic", size=10, bold=True, color=C_CHARCOAL)
FONT_SMALL      = Font(name="Malgun Gothic", size=9, color=C_CHARCOAL)
FONT_KPI_BIG    = Font(name="Malgun Gothic", size=28, bold=True, color=C_DARK_NAVY)
FONT_KPI_LABEL  = Font(name="Malgun Gothic", size=10, color=C_MID_BLUE)
FONT_RED_BOLD   = Font(name="Malgun Gothic", size=11, bold=True, color=C_RED)
FONT_GREEN_BOLD = Font(name="Malgun Gothic", size=11, bold=True, color=C_GREEN)
FONT_LINK       = Font(name="Malgun Gothic", size=10, color="2980B9", underline="single")
FONT_REF_TITLE  = Font(name="Malgun Gothic", size=12, bold=True, color=C_DARK_NAVY)

FILL_NAVY        = PatternFill(start_color=C_DARK_NAVY, end_color=C_DARK_NAVY, fill_type="solid")
FILL_MID_BLUE    = PatternFill(start_color=C_MID_BLUE, end_color=C_MID_BLUE, fill_type="solid")
FILL_LIGHT_BLUE  = PatternFill(start_color=C_LIGHT_BLUE, end_color=C_LIGHT_BLUE, fill_type="solid")
FILL_VERY_LIGHT  = PatternFill(start_color=C_VERY_LIGHT, end_color=C_VERY_LIGHT, fill_type="solid")
FILL_WHITE       = PatternFill(start_color=C_WHITE, end_color=C_WHITE, fill_type="solid")
FILL_RED         = PatternFill(start_color=C_RED, end_color=C_RED, fill_type="solid")
FILL_RED_LIGHT   = PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid")
FILL_ORANGE_LIGHT= PatternFill(start_color=C_ORANGE_LIGHT, end_color=C_ORANGE_LIGHT, fill_type="solid")
FILL_YELLOW_LIGHT= PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid")
FILL_GREEN_LIGHT = PatternFill(start_color=C_GREEN_LIGHT, end_color=C_GREEN_LIGHT, fill_type="solid")
FILL_GREEN       = PatternFill(start_color=C_GREEN, end_color=C_GREEN, fill_type="solid")
FILL_GRAY_LIGHT  = PatternFill(start_color=C_GRAY_LIGHT, end_color=C_GRAY_LIGHT, fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color=C_LIGHT_GREEN, end_color=C_LIGHT_GREEN, fill_type="solid")
FILL_MANUAL      = PatternFill(start_color=C_MANUAL_YELLOW, end_color=C_MANUAL_YELLOW, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")


# ================================================================
# Helpers
# ================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def write_row(ws, row, col_start, values, font=FONT_BODY, fill=None, alignment=ALIGN_CENTER, border=THIN_BORDER):
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        apply_style(c, font=font, fill=fill, alignment=alignment, border=border)


def merge_and_style(ws, r1, c1, r2, c2, value, font, fill, alignment=ALIGN_CENTER, border=MEDIUM_BORDER):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    apply_style(cell, font=font, fill=fill, alignment=alignment, border=border)
    for r in range(r1, r2 + 1):
        for ci in range(c1, c2 + 1):
            ws.cell(row=r, column=ci).border = border


def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def styled_cell(ws, row, col, value, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    apply_style(c, font=font, fill=fill, alignment=alignment, border=border, number_format=number_format)
    return c


def write_table_header(ws, row, col_start, headers, font=FONT_HEADER, fill=FILL_NAVY, alignment=ALIGN_CENTER, border=THIN_BORDER):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        apply_style(c, font=font, fill=fill, alignment=alignment, border=border)


def add_openpyxl_table(ws, name, ref, style_name="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style_name, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


# ================================================================
# Simulation Data
# ================================================================

def generate_sim_data():
    """Generate 66 rows (3 vehicles x 22 part-test combos), all 12 states covered."""
    d = datetime.timedelta
    sim = []
    no = 0
    for v in VEHICLE_ORDER:
        t0 = T0_DATES[v]
        si = 0
        for p in PART_NAMES:
            for t in PART_TESTS[p]:
                no += 1
                st = STATE_PATTERN[si]; si += 1
                pd = t0 + d(days=10 + no % 5)
                sd = t0 + d(days=45 + no % 8)
                ud = t0 + d(days=62 + no % 6)
                wd = t0 + d(days=85 + no % 7)
                yd = t0 + d(days=93 + no % 4)
                fields = {
                    1:  (None, None, None, None, None, None, None),
                    2:  (pd, None, None, None, None, None, None),
                    3:  (pd, "Under Review", None, None, None, None, None),
                    4:  (pd, "Revision Needed", None, None, None, None, None),
                    5:  (pd, "Approved", None, None, None, None, None),
                    6:  (pd, "Approved", None, None, None, None, "06-BIW생산중단"),
                    7:  (pd, "Approved", sd, None, None, None, "07-생산완료_미인수"),
                    8:  (pd, "Approved", sd, None, None, None, None),
                    9:  (pd, "Approved", sd, ud, None, None, None),
                    10: (pd, "Approved", sd, ud, wd, None, None),
                    11: (pd, "Approved", sd, ud, wd, yd, None),
                    12: (pd, "Approved", sd, ud, wd, yd, "12-조건부승인검토"),
                }[st]
                # Tuple: (no, vehicle, partname, testname, j_sub, k_rev, m_act, o_act, q_act, s_sub, t_ovr)
                sim.append((no, v, p, t, *fields))
    return sim


# ================================================================
# Sheet 1: Config (hidden helper)
# ================================================================

def build_config(wb):
    ws = wb.active
    ws.title = "Config"
    ws.sheet_properties.tabColor = C_GRAY

    # A1: Selected Vehicle (default first vehicle)
    ws.cell(row=1, column=1, value="AE_PE")
    apply_style(ws.cell(row=1, column=1), font=FONT_BODY_BOLD, fill=FILL_LIGHT_BLUE,
                alignment=ALIGN_CENTER, border=THIN_BORDER)

    # Label
    ws.cell(row=1, column=2, value="<-- Select Vehicle Code")
    apply_style(ws.cell(row=1, column=2), font=FONT_SMALL, alignment=ALIGN_LEFT)

    # Vehicle dropdown from Reference Data
    dv = DataValidation(
        type="list",
        formula1="'Reference Data 기준데이터'!$A$3:$A$20",
        allow_blank=False
    )
    dv.prompt = "Select a vehicle code"
    dv.promptTitle = "Vehicle"
    ws.add_data_validation(dv)
    dv.add("A1")

    set_col_widths(ws, {"A": 16, "B": 30})
    ws.sheet_state = "hidden"
    return ws


# ================================================================
# Sheet 2: Reference Data (SSOT - 3 tables, 32 rows)
# ================================================================

def build_reference_data(wb):
    ws = wb.create_sheet("Reference Data 기준데이터")
    ws.sheet_properties.tabColor = C_MID_BLUE

    set_col_widths(ws, {
        "A": 30, "B": 28, "C": 18, "D": 18, "E": 18, "F": 18,
        "G": 16, "H": 16, "I": 16, "J": 16,
    })

    # ── TBL_Vehicle (A1:J5) — 10 cols, 3 vehicles ──
    r = 1
    styled_cell(ws, r, 1, "TBL_Vehicle (차종 + 일정 + LeadTime)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 2
    vh = ["VehicleCode", "VehicleName", "T0_DrawingRelease", "ProtoBuild",
          "MasterCar", "TestDeadline", "LT_PlanDue", "LT_PartsReady",
          "LT_TestStart", "LT_ReportDue"]
    write_table_header(ws, r, 1, vh)
    vehicles = [
        ("AE_PE", "AE PE (Creta EV)", datetime.date(2026, 1, 15),
         datetime.date(2026, 4, 15), datetime.date(2026, 6, 15), None,
         14, 49, 63, 98),
        ("SU2i", "SU2i (Venue Next)", datetime.date(2026, 3, 1),
         datetime.date(2026, 6, 1), datetime.date(2026, 8, 15), None,
         14, 49, 63, 98),
        ("EN_SUV", "EN SUV", datetime.date(2026, 2, 1),
         datetime.date(2026, 5, 1), datetime.date(2026, 7, 15), None,
         14, 49, 63, 98),
    ]
    for i, v in enumerate(vehicles):
        row = r + 1 + i
        for j, val in enumerate(v):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
            if isinstance(val, datetime.date):
                c.number_format = "YYYY-MM-DD"
        # TestDeadline (col F=6) = MasterCar (col E=5) - 30
        td_cell = ws.cell(row=row, column=6)
        td_cell.value = f"=E{row}-30"
        apply_style(td_cell, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)
        td_cell.number_format = "YYYY-MM-DD"

    add_openpyxl_table(ws, "TBL_Vehicle", f"A2:J{r + len(vehicles)}", "TableStyleMedium2")

    # ── TBL_TestCatalog (A7:D23) — 15 unique tests ──
    r = 7
    styled_cell(ws, r, 1, "TBL_TestCatalog (시험 카탈로그)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 8
    th = ["TestNameKR", "TestType", "Spec", "DurationDays"]
    write_table_header(ws, r, 1, th)
    unique_tests = list(TEST_CATALOG.items())  # 15 unique tests
    for i, (tname, (ttype, spec, dur)) in enumerate(unique_tests):
        row = r + 1 + i
        for j, val in enumerate([tname, ttype, spec, dur]):
            c = ws.cell(row=row, column=1 + j, value=val)
            apply_style(c, font=FONT_BODY, fill=FILL_WHITE,
                        alignment=ALIGN_LEFT if j == 0 else ALIGN_CENTER, border=THIN_BORDER)

    add_openpyxl_table(ws, "TBL_TestCatalog", f"A8:D{r + len(unique_tests)}", "TableStyleMedium6")

    # ── TBL_PartMap (A25:B32) — 6 parts ──
    r = 25
    styled_cell(ws, r, 1, "TBL_PartMap (부품→구분)", FONT_REF_TITLE, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    r = 26
    ph = ["PartName", "PartCategory"]
    write_table_header(ws, r, 1, ph)
    for i, pname in enumerate(PART_NAMES):
        row = r + 1 + i
        cat = PART_CATEGORIES[pname]
        c1 = ws.cell(row=row, column=1, value=pname)
        apply_style(c1, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER)
        c2 = ws.cell(row=row, column=2, value=cat)
        apply_style(c2, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_CENTER, border=THIN_BORDER)

    add_openpyxl_table(ws, "TBL_PartMap", f"A26:B{r + len(PART_NAMES)}", "TableStyleMedium5")

    return ws


# ================================================================
# Sheet 3: Master Tracker (25 cols, A-Y)
# ================================================================

# V3.2 Column mapping (1-based): A=1 .. Y=25
# A:No, B:VehicleCode, C:PartName, D:PartCategory(auto), E:TestNameKR,
# F:TestType(auto), G:Supplier(MANUAL), H:T0(auto),
# I:PlanDue(auto), J:PlanSubmitted(manual), K:ReviewResult(manual),
# L:PartsReadyPlan(auto), M:PartsReadyActual(manual),
# N:TestStartPlan(auto), O:TestStartActual(manual),
# P:TestCompletePlan(auto), Q:TestCompleteActual(manual),
# R:ReportDue(auto), S:ReportSubmitted(manual),
# T:Override(manual), U:Status(auto), V:Progress%(auto),
# W:DaysToDeadline(auto), X:RiskFlag(auto), Y:NextAction(auto)

# Manual columns (1-based): B(2), C(3), E(5), G(7), J(10), K(11), M(13), O(15), Q(17), S(19), T(20)
MANUAL_COLS = {2, 3, 5, 7, 10, 11, 13, 15, 17, 19, 20}

def build_master_tracker(wb):
    ws = wb.create_sheet("Master Tracker 마스터추적표")
    ws.sheet_properties.tabColor = C_MID_BLUE

    # -- English headers (row 3) --
    headers_en = [
        "No.",                          # A  (1)
        "Vehicle Code",                 # B  (2)  MANUAL
        "Part Name",                    # C  (3)  MANUAL
        "Part Category",                # D  (4)  AUTO
        "Test Name KR",                 # E  (5)  MANUAL
        "Test Type",                    # F  (6)  AUTO
        "Supplier",                     # G  (7)  AUTO
        "T\u2080 Drawing\nRelease",    # H  (8)  AUTO
        "Plan Due\nDate",              # I  (9)  AUTO
        "Plan Submitted\nDate",        # J  (10) MANUAL
        "Review\nResult",             # K  (11) MANUAL
        "Parts Ready\n(Plan)",         # L  (12) AUTO
        "Parts Ready\n(Actual)",       # M  (13) MANUAL
        "Test Start\n(Plan)",          # N  (14) AUTO
        "Test Start\n(Actual)",        # O  (15) MANUAL
        "Test Complete\n(Plan)",       # P  (16) AUTO
        "Test Complete\n(Actual)",     # Q  (17) MANUAL
        "Report Due\nDate",            # R  (18) AUTO
        "Report Submitted\nDate",      # S  (19) MANUAL
        "Override",                     # T  (20) MANUAL
        "Status",                       # U  (21) AUTO
        "Progress\n%",                 # V  (22) AUTO
        "Days to\nDeadline",          # W  (23) AUTO
        "Risk\nFlag",                  # X  (24) AUTO
        "Next Action",                  # Y  (25) AUTO
    ]

    # -- Korean headers (row 4) --
    headers_kr = [
        "번호",              # A
        "차종",              # B
        "부품명",            # C
        "파트구분",          # D
        "스펙명",            # E
        "시험유형",          # F
        "협력사",            # G
        "T₀ 도면배포",      # H
        "계획서기한",        # I
        "계획서제출",        # J
        "검토결과",          # K
        "부품준비(계획)",    # L
        "부품준비(실적)",    # M
        "시험시작(계획)",    # N
        "시험시작(실적)",    # O
        "시험완료(계획)",    # P
        "시험완료(실적)",    # Q
        "성적서기한",        # R
        "성적서제출",        # S
        "오버라이드",        # T
        "상태",              # U
        "진행률%",          # V
        "잔여일",            # W
        "리스크",            # X
        "다음조치",          # Y
    ]

    num_cols = len(headers_en)  # 25

    # -- Title rows (1-2) --
    merge_and_style(ws, 1, 1, 2, num_cols,
                    "ESIR Master Tracker V3 — Engineering Sample / Initial Run Test",
                    FONT_TITLE, FILL_NAVY)

    # -- Header rows (3-4) --
    for i, h in enumerate(headers_en):
        col = i + 1
        c = ws.cell(row=3, column=col, value=h)
        if col in MANUAL_COLS:
            apply_style(c, font=FONT_HEADER, fill=FILL_MID_BLUE, alignment=ALIGN_CENTER, border=THIN_BORDER)
        else:
            apply_style(c, font=FONT_HEADER, fill=FILL_NAVY, alignment=ALIGN_CENTER, border=THIN_BORDER)

    for i, h in enumerate(headers_kr):
        col = i + 1
        c = ws.cell(row=4, column=col, value=h)
        if col in MANUAL_COLS:
            apply_style(c, font=FONT_HEADER_KR, fill=FILL_MANUAL, alignment=ALIGN_CENTER, border=THIN_BORDER)
        else:
            apply_style(c, font=FONT_HEADER_KR, fill=FILL_LIGHT_BLUE, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # -- Column widths --
    col_widths = {
        "A": 6, "B": 12, "C": 24, "D": 12, "E": 20,
        "F": 16, "G": 14, "H": 14, "I": 14, "J": 14,
        "K": 16, "L": 14, "M": 14, "N": 14, "O": 14,
        "P": 14, "Q": 14, "R": 14, "S": 14, "T": 18,
        "U": 22, "V": 10, "W": 10, "X": 12, "Y": 30,
    }
    set_col_widths(ws, col_widths)

    # -- Row height for headers --
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 24

    # -- Generate sim data and write --
    sim = generate_sim_data()
    data_start = 5
    num_items = len(sim)  # 66

    for idx, item in enumerate(sim):
        r = data_start + idx
        no, vehicle, partname, testname, j_sub, k_rev, m_act, o_act, q_act, s_sub, t_ovr = item

        # A: No. (static)
        styled_cell(ws, r, 1, no, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # B: Vehicle Code (MANUAL)
        styled_cell(ws, r, 2, vehicle, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # C: Part Name (MANUAL)
        styled_cell(ws, r, 3, partname, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # D: Part Category (AUTO) = VLOOKUP(C, RNG_PARTMAP, 2, FALSE)
        f_d = f'=IFERROR(VLOOKUP(C{r},{RNG_PARTMAP},2,FALSE),"")'
        styled_cell(ws, r, 4, f_d, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # E: Test Name KR (MANUAL)
        styled_cell(ws, r, 5, testname, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # F: Test Type (AUTO) = VLOOKUP(E, RNG_TESTCATALOG, 2, FALSE)
        f_f = f'=IFERROR(VLOOKUP(E{r},{RNG_TESTCATALOG},2,FALSE),"")'
        styled_cell(ws, r, 6, f_f, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # G: Supplier (MANUAL — no auto-lookup)
        styled_cell(ws, r, 7, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # H: T0 Drawing Release (AUTO) = VLOOKUP(B, RNG_VEHICLE, 3, FALSE)
        f_h = f'=IFERROR(VLOOKUP(B{r},{RNG_VEHICLE},3,FALSE),"")'
        c_h = styled_cell(ws, r, 8, f_h, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_h.number_format = "YYYY-MM-DD"

        # I: Plan Due (AUTO) = H + VLOOKUP(B, VEHICLE, 7) [LT_PlanDue]
        f_i = f'=IFERROR(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},7,FALSE),"")'
        c_i = styled_cell(ws, r, 9, f_i, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_i.number_format = "YYYY-MM-DD"

        # J: Plan Submitted (MANUAL)
        c_j = styled_cell(ws, r, 10, j_sub, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(j_sub, datetime.date):
            c_j.number_format = "YYYY-MM-DD"

        # K: Review Result (MANUAL)
        styled_cell(ws, r, 11, k_rev, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # L: Parts Ready Plan (AUTO) = H + VLOOKUP(B, VEHICLE, 8) [LT_PartsReady]
        f_l = f'=IFERROR(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},8,FALSE),"")'
        c_l = styled_cell(ws, r, 12, f_l, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_l.number_format = "YYYY-MM-DD"

        # M: Parts Ready Actual (MANUAL)
        c_m = styled_cell(ws, r, 13, m_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(m_act, datetime.date):
            c_m.number_format = "YYYY-MM-DD"

        # N: Test Start Plan (AUTO) — vehicle-mounted tests use MAX with ProtoBuild
        f_n = (
            f'=IFERROR(IF(ISNUMBER(SEARCH("Vehicle",F{r})),'
            f'MAX(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},4,FALSE)),'
            f'H{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE)),"")'
        )
        c_n = styled_cell(ws, r, 14, f_n, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_n.number_format = "YYYY-MM-DD"

        # O: Test Start Actual (MANUAL)
        c_o = styled_cell(ws, r, 15, o_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(o_act, datetime.date):
            c_o.number_format = "YYYY-MM-DD"

        # P: Test Complete Plan (AUTO) = N + VLOOKUP(E, TestCatalog, DurationDays)
        f_p = f'=IFERROR(N{r}+VLOOKUP(E{r},{RNG_TESTCATALOG},4,FALSE),"")'
        c_p = styled_cell(ws, r, 16, f_p, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_p.number_format = "YYYY-MM-DD"

        # Q: Test Complete Actual (MANUAL)
        c_q = styled_cell(ws, r, 17, q_act, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(q_act, datetime.date):
            c_q.number_format = "YYYY-MM-DD"

        # R: Report Due (AUTO) = MIN(H+VLOOKUP(B,VEHICLE,10), TestDeadline)
        f_r = (
            f'=IFERROR(MIN(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},10,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},6,FALSE)),"")'
        )
        c_r = styled_cell(ws, r, 18, f_r, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_r.number_format = "YYYY-MM-DD"

        # S: Report Submitted (MANUAL)
        c_s = styled_cell(ws, r, 19, s_sub, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        if isinstance(s_sub, datetime.date):
            c_s.number_format = "YYYY-MM-DD"

        # T: Override (MANUAL)
        styled_cell(ws, r, 20, t_ovr, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # U: Status (AUTO) — 12-state formula
        f_u = (
            f'=IF(T{r}<>"",T{r},'
            f'IF(AND(S{r}<>"",Q{r}<>""),"11-결과등록완료",'
            f'IF(AND(Q{r}<>"",S{r}=""),"10-시험완료",'
            f'IF(AND(O{r}<>"",Q{r}=""),"09-시험준비중",'
            f'IF(AND(M{r}<>"",O{r}=""),"08-단품준비완료",'
            f'IF(K{r}="Approved","05-계획서승인",'
            f'IF(K{r}="Revision Needed","04-검토중_조건불만족",'
            f'IF(K{r}="Under Review","03-검토중_조건충족",'
            f'IF(AND(J{r}<>"",K{r}=""),"02-계획서제출_미검토",'
            f'"01-계획서미제출")))))))))'
        )
        styled_cell(ws, r, 21, f_u, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # V: Progress % (AUTO) = IFS on Status (replaces TBL_Status VLOOKUP)
        f_v = (
            f'=IFS(U{r}="01-계획서미제출",0,U{r}="02-계획서제출_미검토",0.05,'
            f'U{r}="03-검토중_조건충족",0.1,U{r}="04-검토중_조건불만족",0.05,'
            f'U{r}="05-계획서승인",0.2,U{r}="06-BIW생산중단",0.15,'
            f'U{r}="07-생산완료_미인수",0.3,U{r}="08-단품준비완료",0.4,'
            f'U{r}="09-시험준비중",0.55,U{r}="10-시험완료",0.85,'
            f'U{r}="11-결과등록완료",1,U{r}="12-조건부승인검토",0.9,TRUE,0)'
        )
        styled_cell(ws, r, 22, f_v, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # W: Days to Deadline (AUTO)
        f_w = f'=IF(S{r}<>"","-",IF(R{r}="","N/A",R{r}-TODAY()))'
        styled_cell(ws, r, 23, f_w, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # X: Risk Flag (AUTO)
        f_x = (
            f'=IF(U{r}="11-결과등록완료","DONE",'
            f'IF(OR(U{r}="06-BIW생산중단",AND(U{r}="01-계획서미제출",I{r}<TODAY())),"CRITICAL",'
            f'IF(OR(U{r}="04-검토중_조건불만족",U{r}="07-생산완료_미인수",U{r}="12-조건부승인검토"),"ACTION",'
            f'IF(AND(ISNUMBER(W{r}),W{r}<=0),"OVERDUE",'
            f'IF(AND(ISNUMBER(W{r}),W{r}<=14),"WARNING","ON TRACK")))))'
        )
        styled_cell(ws, r, 24, f_x, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Y: Next Action (AUTO) — 12-branch IF
        f_y = (
            f'=IF(U{r}="11-결과등록완료","— 완료",'
            f'IF(U{r}="01-계획서미제출",IF(I{r}<TODAY(),"계획서 제출 독촉 (OVERDUE)","계획서 제출 대기"),'
            f'IF(U{r}="02-계획서제출_미검토","설계자: 계획서 검토 착수",'
            f'IF(U{r}="03-검토중_조건충족","설계자: 승인 처리",'
            f'IF(U{r}="04-검토중_조건불만족","보완 요청 → 협력사 재제출 독촉",'
            f'IF(U{r}="05-계획서승인","시험편 준비 현황 확인",'
            f'IF(U{r}="06-BIW생산중단","BIW 생산팀 확인 → 일정 재수립",'
            f'IF(U{r}="07-생산완료_미인수","협력사 인수 독촉",'
            f'IF(U{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(U{r}="09-시험준비중","시험 진행 모니터링",'
            f'IF(U{r}="10-시험완료","결과 보고서 전산 등록 독촉",'
            f'IF(U{r}="12-조건부승인검토","조건부 승인 여부 판단 → 관련부서 협의","—"))))))))))))'
        )
        styled_cell(ws, r, 25, f_y, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

    # -- Empty rows (after sim data to row TRACKER_END) with wrapped formulas --
    data_end = data_start + num_items - 1  # last sim row
    for r in range(data_end + 1, TRACKER_END + 1):
        # A: No. — auto-increment if B has value
        f_a = f'=IF(B{r}="","",ROW()-{data_start - 1})'
        styled_cell(ws, r, 1, f_a, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # B: Vehicle Code (MANUAL - blank)
        styled_cell(ws, r, 2, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # C: Part Name (MANUAL - blank)
        styled_cell(ws, r, 3, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # D: Part Category (AUTO, wrapped)
        f_d = f'=IF(B{r}="","",IFERROR(VLOOKUP(C{r},{RNG_PARTMAP},2,FALSE),""))'
        styled_cell(ws, r, 4, f_d, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # E: Test Name KR (MANUAL - blank)
        styled_cell(ws, r, 5, None, FONT_BODY, FILL_MANUAL, ALIGN_LEFT, THIN_BORDER)

        # F: Test Type (AUTO, wrapped)
        f_f = f'=IF(B{r}="","",IFERROR(VLOOKUP(E{r},{RNG_TESTCATALOG},2,FALSE),""))'
        styled_cell(ws, r, 6, f_f, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # G: Supplier (MANUAL - blank)
        styled_cell(ws, r, 7, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # H: T0 (AUTO, wrapped)
        f_h = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},{RNG_VEHICLE},3,FALSE),""))'
        c_h = styled_cell(ws, r, 8, f_h, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_h.number_format = "YYYY-MM-DD"

        # I: Plan Due (AUTO, wrapped)
        f_i = f'=IF(B{r}="","",IFERROR(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},7,FALSE),""))'
        c_i = styled_cell(ws, r, 9, f_i, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_i.number_format = "YYYY-MM-DD"

        # J: Plan Submitted (MANUAL - blank)
        c_j = styled_cell(ws, r, 10, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_j.number_format = "YYYY-MM-DD"

        # K: Review Result (MANUAL - blank)
        styled_cell(ws, r, 11, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # L: Parts Ready Plan (AUTO, wrapped)
        f_l = f'=IF(B{r}="","",IFERROR(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},8,FALSE),""))'
        c_l = styled_cell(ws, r, 12, f_l, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_l.number_format = "YYYY-MM-DD"

        # M: Parts Ready Actual (MANUAL - blank)
        c_m = styled_cell(ws, r, 13, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_m.number_format = "YYYY-MM-DD"

        # N: Test Start Plan (AUTO, wrapped)
        f_n = (
            f'=IF(B{r}="","",IFERROR(IF(ISNUMBER(SEARCH("Vehicle",F{r})),'
            f'MAX(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},4,FALSE)),'
            f'H{r}+VLOOKUP(B{r},{RNG_VEHICLE},9,FALSE)),""))'
        )
        c_n = styled_cell(ws, r, 14, f_n, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_n.number_format = "YYYY-MM-DD"

        # O: Test Start Actual (MANUAL - blank)
        c_o = styled_cell(ws, r, 15, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_o.number_format = "YYYY-MM-DD"

        # P: Test Complete Plan (AUTO, wrapped)
        f_p = f'=IF(B{r}="","",IFERROR(N{r}+VLOOKUP(E{r},{RNG_TESTCATALOG},4,FALSE),""))'
        c_p = styled_cell(ws, r, 16, f_p, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_p.number_format = "YYYY-MM-DD"

        # Q: Test Complete Actual (MANUAL - blank)
        c_q = styled_cell(ws, r, 17, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_q.number_format = "YYYY-MM-DD"

        # R: Report Due (AUTO, wrapped)
        f_r = (
            f'=IF(B{r}="","",IFERROR(MIN(H{r}+VLOOKUP(B{r},{RNG_VEHICLE},10,FALSE),'
            f'VLOOKUP(B{r},{RNG_VEHICLE},6,FALSE)),""))'
        )
        c_r = styled_cell(ws, r, 18, f_r, FONT_BODY, FILL_VERY_LIGHT, ALIGN_CENTER, THIN_BORDER)
        c_r.number_format = "YYYY-MM-DD"

        # S: Report Submitted (MANUAL - blank)
        c_s = styled_cell(ws, r, 19, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)
        c_s.number_format = "YYYY-MM-DD"

        # T: Override (MANUAL - blank)
        styled_cell(ws, r, 20, None, FONT_BODY, FILL_MANUAL, ALIGN_CENTER, THIN_BORDER)

        # U: Status (AUTO, wrapped)
        f_u = (
            f'=IF(B{r}="","",'
            f'IF(T{r}<>"",T{r},'
            f'IF(AND(S{r}<>"",Q{r}<>""),"11-결과등록완료",'
            f'IF(AND(Q{r}<>"",S{r}=""),"10-시험완료",'
            f'IF(AND(O{r}<>"",Q{r}=""),"09-시험준비중",'
            f'IF(AND(M{r}<>"",O{r}=""),"08-단품준비완료",'
            f'IF(K{r}="Approved","05-계획서승인",'
            f'IF(K{r}="Revision Needed","04-검토중_조건불만족",'
            f'IF(K{r}="Under Review","03-검토중_조건충족",'
            f'IF(AND(J{r}<>"",K{r}=""),"02-계획서제출_미검토",'
            f'"01-계획서미제출"))))))))))'
        )
        styled_cell(ws, r, 21, f_u, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # V: Progress % (AUTO, wrapped with IFS)
        f_v = (
            f'=IF(B{r}="","",IFS(U{r}="01-계획서미제출",0,U{r}="02-계획서제출_미검토",0.05,'
            f'U{r}="03-검토중_조건충족",0.1,U{r}="04-검토중_조건불만족",0.05,'
            f'U{r}="05-계획서승인",0.2,U{r}="06-BIW생산중단",0.15,'
            f'U{r}="07-생산완료_미인수",0.3,U{r}="08-단품준비완료",0.4,'
            f'U{r}="09-시험준비중",0.55,U{r}="10-시험완료",0.85,'
            f'U{r}="11-결과등록완료",1,U{r}="12-조건부승인검토",0.9,TRUE,0))'
        )
        styled_cell(ws, r, 22, f_v, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # W: Days to Deadline (AUTO, wrapped)
        f_w = f'=IF(B{r}="","",IF(S{r}<>"","-",IF(R{r}="","N/A",R{r}-TODAY())))'
        styled_cell(ws, r, 23, f_w, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # X: Risk Flag (AUTO, wrapped)
        f_x = (
            f'=IF(B{r}="","",'
            f'IF(U{r}="11-결과등록완료","DONE",'
            f'IF(OR(U{r}="06-BIW생산중단",AND(U{r}="01-계획서미제출",I{r}<TODAY())),"CRITICAL",'
            f'IF(OR(U{r}="04-검토중_조건불만족",U{r}="07-생산완료_미인수",U{r}="12-조건부승인검토"),"ACTION",'
            f'IF(AND(ISNUMBER(W{r}),W{r}<=0),"OVERDUE",'
            f'IF(AND(ISNUMBER(W{r}),W{r}<=14),"WARNING","ON TRACK"))))))'
        )
        styled_cell(ws, r, 24, f_x, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Y: Next Action (AUTO, wrapped)
        f_y = (
            f'=IF(B{r}="","",'
            f'IF(U{r}="11-결과등록완료","— 완료",'
            f'IF(U{r}="01-계획서미제출",IF(I{r}<TODAY(),"계획서 제출 독촉 (OVERDUE)","계획서 제출 대기"),'
            f'IF(U{r}="02-계획서제출_미검토","설계자: 계획서 검토 착수",'
            f'IF(U{r}="03-검토중_조건충족","설계자: 승인 처리",'
            f'IF(U{r}="04-검토중_조건불만족","보완 요청 → 협력사 재제출 독촉",'
            f'IF(U{r}="05-계획서승인","시험편 준비 현황 확인",'
            f'IF(U{r}="06-BIW생산중단","BIW 생산팀 확인 → 일정 재수립",'
            f'IF(U{r}="07-생산완료_미인수","협력사 인수 독촉",'
            f'IF(U{r}="08-단품준비완료","시험 일정 확인",'
            f'IF(U{r}="09-시험준비중","시험 진행 모니터링",'
            f'IF(U{r}="10-시험완료","결과 보고서 전산 등록 독촉",'
            f'IF(U{r}="12-조건부승인검토","조건부 승인 여부 판단 → 관련부서 협의","—")))))))))))))'
        )
        styled_cell(ws, r, 25, f_y, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

    # -- Data Validations (rows 5:TRACKER_END) --
    # B: Vehicle Code dropdown
    dv_vehicle = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$3:$A$5", allow_blank=True)
    dv_vehicle.prompt = "Select vehicle"
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add(f"B5:B{TRACKER_END}")

    # C: Part Name dropdown (from TBL_PartMap)
    dv_part = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$27:$A$32", allow_blank=True)
    dv_part.prompt = "Select part"
    ws.add_data_validation(dv_part)
    dv_part.add(f"C5:C{TRACKER_END}")

    # E: Test Name KR dropdown (from TBL_TestCatalog)
    dv_test = DataValidation(type="list", formula1=f"'Reference Data 기준데이터'!$A$9:$A$23", allow_blank=True)
    dv_test.prompt = "Select test"
    ws.add_data_validation(dv_test)
    dv_test.add(f"E5:E{TRACKER_END}")

    # K: Review Result dropdown
    dv_review = DataValidation(type="list", formula1='"Under Review,Approved,Revision Needed"', allow_blank=True)
    dv_review.prompt = "Select review result"
    ws.add_data_validation(dv_review)
    dv_review.add(f"K5:K{TRACKER_END}")

    # T: Override dropdown
    dv_override = DataValidation(
        type="list",
        formula1='"06-BIW생산중단,07-생산완료_미인수,12-조건부승인검토"',
        allow_blank=True
    )
    dv_override.prompt = "Select override status"
    ws.add_data_validation(dv_override)
    dv_override.add(f"T5:T{TRACKER_END}")

    # -- Conditional Formatting --
    range_str = f"U5:U{TRACKER_END}"

    # U (Status) - 12 status color rules
    status_colors = [
        ("01-계획서미제출", C_GRAY_LIGHT, C_CHARCOAL),
        ("02-계획서제출_미검토", C_LIGHT_BLUE, C_CHARCOAL),
        ("03-검토중_조건충족", C_LIGHT_BLUE, C_MID_BLUE),
        ("04-검토중_조건불만족", C_ORANGE_LIGHT, C_CHARCOAL),
        ("05-계획서승인", C_GREEN_LIGHT, C_CHARCOAL),
        ("06-BIW생산중단", C_RED_LIGHT, C_RED),
        ("07-생산완료_미인수", C_ORANGE_LIGHT, C_ORANGE),
        ("08-단품준비완료", C_GREEN_LIGHT, C_GREEN),
        ("09-시험준비중", C_GREEN_LIGHT, C_GREEN),
        ("10-시험완료", C_LIGHT_GREEN, C_CHARCOAL),
        ("11-결과등록완료", C_GREEN, C_WHITE),
        ("12-조건부승인검토", C_YELLOW_LIGHT, C_CHARCOAL),
    ]
    for status_code, bg_color, font_color in status_colors:
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(
                operator="equal",
                formula=[f'"{status_code}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Malgun Gothic", size=10, bold=True, color=font_color),
            )
        )

    # X (Risk) - 6 risk color rules
    risk_range = f"X5:X{TRACKER_END}"
    risk_colors = [
        ("DONE", C_GREEN_LIGHT, C_GREEN),
        ("CRITICAL", C_RED_LIGHT, C_RED),
        ("OVERDUE", C_RED_LIGHT, C_RED),
        ("WARNING", C_YELLOW_LIGHT, C_ORANGE),
        ("ACTION", C_ORANGE_LIGHT, C_ORANGE),
        ("ON TRACK", C_GREEN_LIGHT, C_GREEN),
    ]
    for risk_code, bg_color, font_color in risk_colors:
        ws.conditional_formatting.add(
            risk_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{risk_code}"'],
                fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
                font=Font(name="Malgun Gothic", size=10, bold=True, color=font_color),
            )
        )

    # W (Days to Deadline) - <=0 red, <=14 yellow
    days_range = f"W5:W{TRACKER_END}"
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        days_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["14"],
            fill=PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_ORANGE),
        )
    )

    # V (Progress) - DataBar
    progress_range = f"V5:V{TRACKER_END}"
    ws.conditional_formatting.add(
        progress_range,
        DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color=C_MID_BLUE,
        )
    )

    # Date plan/actual pairs: if plan < TODAY and actual is empty -> red highlight
    # Plan/Actual pairs: (I,J), (L,M), (N,O), (P,Q), (R,S) — columns 9/10, 12/13, 14/15, 16/17, 18/19
    plan_actual_pairs = [
        ("I", "J"), ("L", "M"), ("N", "O"), ("P", "Q"), ("R", "S"),
    ]
    for plan_col, actual_col in plan_actual_pairs:
        pa_range = f"{plan_col}5:{plan_col}{TRACKER_END}"
        ws.conditional_formatting.add(
            pa_range,
            FormulaRule(
                formula=[f'AND({plan_col}5<TODAY(),{actual_col}5="")'],
                fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            )
        )

    # -- Auto-filter --
    ws.auto_filter.ref = f"A4:Y{TRACKER_END}"

    # -- Freeze panes at A5 --
    ws.freeze_panes = "A5"

    return ws, data_start, data_end + 1, num_items


# ================================================================
# Sheet 4: Dashboard
# ================================================================

def build_dashboard(wb, tracker_sheet_name, data_start, data_end, num_items):
    ws = wb.create_sheet("Dashboard 대시보드")
    ws.sheet_properties.tabColor = C_GREEN

    ts = f"'{tracker_sheet_name}'"
    cfg = "Config!$A$1"  # vehicle selector

    # Column ranges on tracker
    b_range = f"{ts}!$B${data_start}:$B${data_end}"
    c_range = f"{ts}!$C${data_start}:$C${data_end}"
    u_range = f"{ts}!$U${data_start}:$U${data_end}"
    v_range = f"{ts}!$V${data_start}:$V${data_end}"
    w_range = f"{ts}!$W${data_start}:$W${data_end}"
    x_range = f"{ts}!$X${data_start}:$X${data_end}"

    set_col_widths(ws, {
        "A": 4, "B": 24, "C": 14, "D": 14, "E": 14,
        "F": 14, "G": 14, "H": 14,
    })

    # -- Row 1-2: Title --
    merge_and_style(ws, 1, 1, 2, 8,
                    "ESIR Dashboard V3",
                    FONT_TITLE, FILL_NAVY)

    # -- Row 3: Vehicle selector + Key dates --
    styled_cell(ws, 3, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 3, 2, "Selected Vehicle:", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)

    # C3: vehicle selector (linked to Config!A1)
    c3 = styled_cell(ws, 3, 3, f"={cfg}", FONT_BODY_BOLD, FILL_MANUAL, ALIGN_CENTER, MEDIUM_BORDER)

    # D3: T0 date
    styled_cell(ws, 3, 4, "T0:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    t0_cell = styled_cell(ws, 3, 5,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},3,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    t0_cell.number_format = "YYYY-MM-DD"

    # F3: Proto
    styled_cell(ws, 3, 6, "Proto:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    proto_cell = styled_cell(ws, 3, 7,
                             f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},4,FALSE),"")',
                             FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    proto_cell.number_format = "YYYY-MM-DD"

    # H3: MC
    styled_cell(ws, 3, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Row 4: More dates + D-Day
    styled_cell(ws, 4, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 4, 2, "D-Day:", FONT_RED_BOLD, FILL_WHITE, ALIGN_RIGHT, THIN_BORDER)

    # C4: D-Day countdown = TestDeadline - TODAY()
    f_dday = f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},6,FALSE)-TODAY(),"")'
    styled_cell(ws, 4, 3, f_dday, FONT_RED_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    styled_cell(ws, 4, 4, "MC:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    mc_cell = styled_cell(ws, 4, 5,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},5,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    mc_cell.number_format = "YYYY-MM-DD"

    styled_cell(ws, 4, 6, "Deadline:", FONT_SMALL, FILL_LIGHT_BLUE, ALIGN_RIGHT, THIN_BORDER)
    dl_cell = styled_cell(ws, 4, 7,
                          f'=IFERROR(VLOOKUP(C3,{RNG_VEHICLE},6,FALSE),"")',
                          FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    dl_cell.number_format = "YYYY-MM-DD"

    styled_cell(ws, 4, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Vehicle dropdown on C3
    dv_vehicle = DataValidation(
        type="list",
        formula1=f"'Reference Data 기준데이터'!$A$3:$A$5",
        allow_blank=False
    )
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add("C3")

    # -- Row 5: blank separator --
    for col in range(1, 9):
        styled_cell(ws, 5, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # -- Row 6: Section header: "KPI Summary" --
    merge_and_style(ws, 6, 1, 6, 8,
                    "KPI Summary",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 7: KPI labels --
    kpi_labels = ["", "Total", "Complete 완료", "In Progress 진행중", "At Risk 지연", "Avg Progress", "", ""]
    for i, label in enumerate(kpi_labels):
        styled_cell(ws, 7, i + 1, label, FONT_KPI_LABEL, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)

    # -- Row 8: KPI values --
    # B8: Total = COUNTIF(B_range, vehicle)
    f_total = f'=COUNTIF({b_range},{cfg})'
    styled_cell(ws, 8, 2, f_total, FONT_KPI_BIG, FILL_WHITE, ALIGN_CENTER, MEDIUM_BORDER)

    # C8: Complete = COUNTIFS(B, vehicle, U, "11-*")
    f_complete = f'=COUNTIFS({b_range},{cfg},{u_range},"11-*")'
    styled_cell(ws, 8, 3, f_complete, FONT_KPI_BIG, FILL_GREEN_LIGHT, ALIGN_CENTER, MEDIUM_BORDER)

    # D8: In Progress = Total - Complete - NotStarted(01)
    f_notstarted = f'COUNTIFS({b_range},{cfg},{u_range},"01-*")'
    f_inprogress = f'=B8-C8-{f_notstarted}'
    styled_cell(ws, 8, 4, f_inprogress, FONT_KPI_BIG, FILL_LIGHT_BLUE, ALIGN_CENTER, MEDIUM_BORDER)

    # E8: At Risk = CRITICAL + OVERDUE + WARNING
    f_risk = (
        f'=COUNTIFS({b_range},{cfg},{x_range},"CRITICAL")'
        f'+COUNTIFS({b_range},{cfg},{x_range},"OVERDUE")'
        f'+COUNTIFS({b_range},{cfg},{x_range},"WARNING")'
    )
    styled_cell(ws, 8, 5, f_risk, FONT_KPI_BIG, FILL_RED_LIGHT, ALIGN_CENTER, MEDIUM_BORDER)

    # F8: Avg Progress
    f_avg = f'=IFERROR(AVERAGEIFS({v_range},{b_range},{cfg}),0)'
    styled_cell(ws, 8, 6, f_avg, FONT_KPI_BIG, FILL_WHITE, ALIGN_CENTER, MEDIUM_BORDER, number_format="0%")

    styled_cell(ws, 8, 1, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 8, 7, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, 8, 8, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # -- Row 9: blank --
    for col in range(1, 9):
        styled_cell(ws, 9, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # -- Row 10: Section: "Part Progress 부품별 현황 (Selected Vehicle)" --
    merge_and_style(ws, 10, 1, 10, 8,
                    "Part Progress 부품별 현황 (Selected Vehicle)",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 11: Headers --
    part_headers = ["", "부품명", "시험수", "완료", "진행중", "지연", "진행률", "최소잔여일"]
    for i, h in enumerate(part_headers):
        styled_cell(ws, 11, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 12-17: One row per part --
    for pi, pname in enumerate(PART_NAMES):
        r = 12 + pi
        styled_cell(ws, r, 1, pi + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, pname, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # C: 시험수 = COUNTIFS(B, vehicle, C, partname)
        f_cnt = f'=COUNTIFS({b_range},{cfg},{c_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # D: 완료 = COUNTIFS(B, vehicle, C, partname, U, "11-*")
        f_done = f'=COUNTIFS({b_range},{cfg},{c_range},B{r},{u_range},"11-*")'
        styled_cell(ws, r, 4, f_done, FONT_BODY, FILL_GREEN_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # F: 지연 = CRITICAL+OVERDUE+WARNING
        f_delay = (
            f'=COUNTIFS({b_range},{cfg},{c_range},B{r},{x_range},"CRITICAL")'
            f'+COUNTIFS({b_range},{cfg},{c_range},B{r},{x_range},"OVERDUE")'
            f'+COUNTIFS({b_range},{cfg},{c_range},B{r},{x_range},"WARNING")'
        )
        styled_cell(ws, r, 6, f_delay, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # E: 진행중 = Total - 완료 - 지연 (note: this is approximate but matches spec)
        f_prog = f'=C{r}-D{r}-F{r}'
        styled_cell(ws, r, 5, f_prog, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # G: 진행률 = AVERAGEIFS
        f_pct = f'=IFERROR(AVERAGEIFS({v_range},{b_range},{cfg},{c_range},B{r}),0)'
        styled_cell(ws, r, 7, f_pct, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # H: 최소잔여일
        f_min = f'=IFERROR(MINIFS({w_range},{b_range},{cfg},{c_range},B{r},{x_range},"<>DONE"),"—")'
        styled_cell(ws, r, 8, f_min, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Conditional formatting on Part Progress
    # 지연 column (F12:F17): >0 -> red fill
    ws.conditional_formatting.add(
        "F12:F17",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    # 최소잔여일 (H12:H17): <=0 -> red, <=14 -> yellow
    ws.conditional_formatting.add(
        "H12:H17",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )
    ws.conditional_formatting.add(
        "H12:H17",
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["14"],
            fill=PatternFill(start_color=C_YELLOW_LIGHT, end_color=C_YELLOW_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_ORANGE),
        )
    )

    # -- Row 18: blank --
    for col in range(1, 9):
        styled_cell(ws, 18, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # -- Row 19: Section: "Cross-Vehicle Comparison 차종간 비교" --
    merge_and_style(ws, 19, 1, 19, 8,
                    "Cross-Vehicle Comparison 차종간 비교",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 20: Headers --
    xv_headers = ["", "차종", "시험수", "완료", "지연", "진행률", "", ""]
    for i, h in enumerate(xv_headers):
        styled_cell(ws, 20, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 21-23: One row per vehicle --
    for vi, vcode in enumerate(VEHICLE_ORDER):
        r = 21 + vi
        styled_cell(ws, r, 1, vi + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, vcode, FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # C: 시험수
        f_cnt = f'=COUNTIF({b_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # D: 완료
        f_done = f'=COUNTIFS({b_range},B{r},{u_range},"11-*")'
        styled_cell(ws, r, 4, f_done, FONT_BODY, FILL_GREEN_LIGHT, ALIGN_CENTER, THIN_BORDER)

        # E: 지연
        f_delay = (
            f'=COUNTIFS({b_range},B{r},{x_range},"CRITICAL")'
            f'+COUNTIFS({b_range},B{r},{x_range},"OVERDUE")'
            f'+COUNTIFS({b_range},B{r},{x_range},"WARNING")'
        )
        styled_cell(ws, r, 5, f_delay, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # F: 진행률
        f_pct = f'=IFERROR(AVERAGEIFS({v_range},{b_range},B{r}),0)'
        styled_cell(ws, r, 6, f_pct, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        for col in [7, 8]:
            styled_cell(ws, r, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    # Row 24: TOTAL row
    r = 24
    styled_cell(ws, r, 1, "", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 2, "TOTAL", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 3, "=SUM(C21:C23)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 4, "=SUM(D21:D23)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 5, "=SUM(E21:E23)", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    styled_cell(ws, r, 6, f'=IFERROR(AVERAGEIFS({v_range},{b_range},"<>"),0)',
                FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER, number_format="0%")
    for col in [7, 8]:
        styled_cell(ws, r, col, "", FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)

    # Conditional formatting on 지연 column (E21:E24)
    ws.conditional_formatting.add(
        "E21:E24",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"),
            font=Font(name="Malgun Gothic", size=10, bold=True, color=C_RED),
        )
    )

    # -- Row 25: blank --
    for col in range(1, 9):
        styled_cell(ws, 25, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER)

    # -- Row 26: Section: "Status Distribution 상태 분포 (Selected Vehicle)" --
    merge_and_style(ws, 26, 1, 26, 8,
                    "Status Distribution 상태 분포 (Selected Vehicle)",
                    FONT_SUBTITLE, FILL_MID_BLUE)

    # -- Row 27: Headers --
    sd_headers = ["", "상태", "건수", "", "", "", "", ""]
    for i, h in enumerate(sd_headers):
        styled_cell(ws, 27, i + 1, h, FONT_HEADER, FILL_NAVY, ALIGN_CENTER, THIN_BORDER)

    # -- Rows 28-39: 12 status rows --
    status_codes = [
        "01-계획서미제출",
        "02-계획서제출_미검토",
        "03-검토중_조건충족",
        "04-검토중_조건불만족",
        "05-계획서승인",
        "06-BIW생산중단",
        "07-생산완료_미인수",
        "08-단품준비완료",
        "09-시험준비중",
        "10-시험완료",
        "11-결과등록완료",
        "12-조건부승인검토",
    ]
    for si, scode in enumerate(status_codes):
        r = 28 + si
        styled_cell(ws, r, 1, si + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
        styled_cell(ws, r, 2, scode, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        f_cnt = f'=COUNTIFS({b_range},{cfg},{u_range},B{r})'
        styled_cell(ws, r, 3, f_cnt, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        for col in range(4, 9):
            styled_cell(ws, r, col, "", FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    return ws


# ================================================================
# Sheet 5: Supplier Checklist
# ================================================================

def build_supplier_checklist(wb, tracker_sheet_name, data_start, data_end):
    ws = wb.create_sheet("Supplier Checklist 협력사체크리스트")
    ws.sheet_properties.tabColor = C_ORANGE

    ts = f"'{tracker_sheet_name}'"

    set_col_widths(ws, {
        "A": 4, "B": 14, "C": 24, "D": 18, "E": 22,
        "F": 14, "G": 14, "H": 30,
    })

    # Title
    merge_and_style(ws, 1, 1, 2, 8,
                    "Supplier Checklist 협력사 체크리스트",
                    FONT_TITLE, FILL_NAVY)

    # Headers
    sc_headers = ["No.", "Supplier", "Part Name", "Vehicle", "Status",
                  "Progress", "Risk", "Next Action"]
    write_table_header(ws, 3, 1, sc_headers)

    # Placeholder rows referencing tracker data
    for i in range(20):
        r = 4 + i
        tr = data_start + i
        if tr > data_end:
            break

        styled_cell(ws, r, 1, i + 1, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # G (col 7 in tracker) = Supplier
        f_sup = f"={ts}!G{tr}"
        styled_cell(ws, r, 2, f_sup, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # C (col 3 in tracker) = Part Name
        f_part = f"={ts}!C{tr}"
        styled_cell(ws, r, 3, f_part, FONT_BODY, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

        # B (col 2 in tracker) = Vehicle
        f_veh = f"={ts}!B{tr}"
        styled_cell(ws, r, 4, f_veh, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # U (col 21) = Status
        f_status = f"={ts}!U{tr}"
        styled_cell(ws, r, 5, f_status, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # V (col 22) = Progress
        f_prog = f"={ts}!V{tr}"
        styled_cell(ws, r, 6, f_prog, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER, number_format="0%")

        # X (col 24) = Risk
        f_risk = f"={ts}!X{tr}"
        styled_cell(ws, r, 7, f_risk, FONT_BODY, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

        # Y (col 25) = Next Action
        f_next = f"={ts}!Y{tr}"
        styled_cell(ws, r, 8, f_next, FONT_SMALL, FILL_WHITE, ALIGN_LEFT, THIN_BORDER)

    return ws


# ================================================================
# Main
# ================================================================

def main():
    wb = Workbook()
    build_config(wb)
    build_reference_data(wb)
    ws_tracker, data_start, data_end, num_items = build_master_tracker(wb)
    build_dashboard(wb, ws_tracker.title, data_start, TRACKER_END, num_items)
    build_supplier_checklist(wb, ws_tracker.title, data_start, TRACKER_END)
    wb.active = wb.sheetnames.index("Dashboard 대시보드")
    wb.save(OUTPUT)
    print(f"V3.2 Generated: {OUTPUT}")
    print(f"  - {num_items} tracker rows (data rows {data_start}-{data_start + num_items - 1})")
    print(f"  - Empty formula rows {data_start + num_items}-{TRACKER_END}")
    print(f"  - 25 columns (A-Y), 5 sheets")


if __name__ == "__main__":
    main()
