"""
Render ESIR Master Tracker V4 as HTML preview.
Computes formula values in Python (openpyxl can't evaluate).
V4: 28-column layout (A-AB), 3 vehicles, 12 items, 18 tests.
"""
import sys, io, datetime, html
from collections import Counter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook

FILE = r"C:\Users\USER\kpi-cascade\ESIR_Master_Tracker_v4.xlsx"
OUT  = r"C:\Users\USER\kpi-cascade\esir_preview.html"
TODAY = datetime.date(2026, 2, 27)

# ── Reference Data Loaders (V4 layout) ──

def load_ref_tables(wb):
    ws = None
    for name in wb.sheetnames:
        if "Reference Data" in name:
            ws = wb[name]
            break
    if not ws:
        return {}, {}, {}

    # TBL_Vehicle: header row 2, data rows 3-5, cols A-J (10 cols)
    vehicles = {}
    for r in range(3, 6):
        code = ws.cell(row=r, column=1).value
        if not code or not isinstance(code, str):
            break
        t0 = ws.cell(row=r, column=3).value
        proto = ws.cell(row=r, column=4).value
        mc = ws.cell(row=r, column=5).value
        lt_plan = ws.cell(row=r, column=7).value or 14
        lt_parts = ws.cell(row=r, column=8).value or 49
        lt_test = ws.cell(row=r, column=9).value or 63
        lt_report = ws.cell(row=r, column=10).value or 98

        for key in ['t0', 'proto', 'mc']:
            val = locals()[key]
            if hasattr(val, 'date'):
                locals()[key] = val.date()
        if hasattr(t0, 'date'): t0 = t0.date()
        if hasattr(proto, 'date'): proto = proto.date()
        if hasattr(mc, 'date'): mc = mc.date()

        # TestDeadline = MasterCar - 30
        deadline = (mc - datetime.timedelta(days=30)) if isinstance(mc, datetime.date) else None

        vehicles[code] = {
            'name': ws.cell(row=r, column=2).value,
            't0': t0, 'proto': proto, 'mc': mc, 'deadline': deadline,
            'lt_plan': lt_plan, 'lt_parts': lt_parts,
            'lt_test': lt_test, 'lt_report': lt_report,
        }

    # TBL_ItemMaster: header row 29, data rows 30-41, cols A-D
    items = {}
    for r in range(30, 42):
        name = ws.cell(row=r, column=1).value
        if not name or not isinstance(name, str):
            break
        items[name] = {
            'category': ws.cell(row=r, column=2).value,
            'supplier': ws.cell(row=r, column=3).value,
            'drawing': ws.cell(row=r, column=4).value,
        }

    # TBL_TestCatalog: header row 8, data rows 9-26, cols A-E
    tests = {}
    for r in range(9, 27):
        name = ws.cell(row=r, column=1).value
        if not name or not isinstance(name, str):
            break
        tests[name] = {
            'type': ws.cell(row=r, column=2).value,
            'spec': ws.cell(row=r, column=3).value,
            'duration': ws.cell(row=r, column=4).value or 14,
            'scope': ws.cell(row=r, column=5).value,
        }

    return vehicles, items, tests


# ── Computation Logic (V4) ──

def compute_status(k_sub, l_rev, n_act, p_act, r_act, t_sub, u_ovr):
    """V4 Status: 12-state formula (matches V column)."""
    if u_ovr: return u_ovr
    if t_sub and r_act: return "11-결과등록완료"
    if r_act and not t_sub: return "10-시험완료"
    if p_act and not r_act: return "09-시험준비중"
    if n_act and not p_act: return "08-단품준비완료"
    if l_rev == "Approved": return "05-계획서승인"
    if l_rev == "Revision Needed": return "04-검토중_보완"
    if l_rev == "Under Review": return "03-검토중"
    if k_sub and not l_rev: return "02-제출_미검토"
    return "01-미제출"

PROGRESS = {
    "01-미제출": 0, "02-제출_미검토": 5, "03-검토중": 10,
    "04-검토중_보완": 5, "05-계획서승인": 20, "06-BIW중단": 15,
    "07-미인수": 30, "08-단품준비완료": 40, "09-시험준비중": 55,
    "10-시험완료": 85, "11-결과등록완료": 100, "12-조건부승인": 90,
}

def compute_risk(status, plan_due, report_due, t_sub, u_ovr):
    """V4 Risk: 7-state (DONE, BLOCKED, CRITICAL, OVERDUE, WARNING, ACTION, ON TRACK)."""
    if status == "11-결과등록완료": return "DONE"
    if u_ovr == "06-BIW중단": return "BLOCKED"
    if status == "01-미제출" and plan_due and plan_due < TODAY: return "CRITICAL"
    if report_due and not t_sub:
        days = (report_due - TODAY).days
        if days <= 0: return "OVERDUE"
        if days <= 14: return "WARNING"
    if u_ovr: return "ACTION"
    return "ON TRACK"

ACTIONS = {
    "11-결과등록완료": "— 완료",
    "02-제출_미검토": "설계자: 계획서 검토 착수",
    "03-검토중": "설계자: 승인 처리",
    "04-검토중_보완": "보완 요청 → 협력사 재제출 독촉",
    "05-계획서승인": "시험편 준비 현황 확인",
    "06-BIW중단": "BIW 생산팀 확인 → 일정 재수립",
    "07-미인수": "협력사 인수 독촉",
    "08-단품준비완료": "시험 일정 확인",
    "09-시험준비중": "시험 진행 모니터링",
    "10-시험완료": "결과 보고서 전산 등록 독촉",
    "12-조건부승인": "조건부 승인 여부 판단 → 관련부서 협의",
}

RISK_STYLE = {
    "DONE":     ("#D5F5E3", "#27AE60", "✓ DONE"),
    "BLOCKED":  ("#D5D8DC", "#2C3E50", "⬛ BLOCKED"),
    "CRITICAL": ("#E74C3C", "#FFF",    "🔴 CRITICAL"),
    "OVERDUE":  ("#FADBD8", "#C0392B", "🔴 OVERDUE"),
    "ACTION":   ("#FDEBD0", "#E67E22", "🟠 ACTION"),
    "WARNING":  ("#FEF9E7", "#B7950B", "🟡 WARNING"),
    "ON TRACK": ("#EBF1F8", "#2C3E50", "🟢 ON TRACK"),
}

STATUS_BG = {
    "01-미제출": "#F2F3F4", "02-제출_미검토": "#F2F3F4",
    "03-검토중": "#F2F3F4", "04-검토중_보완": "#FADBD8",
    "05-계획서승인": "#FEF9E7", "06-BIW중단": "#E74C3C",
    "07-미인수": "#FDEBD0", "08-단품준비완료": "#FEF9E7",
    "09-시험준비중": "#FEF9E7", "10-시험완료": "#ABEBC6",
    "11-결과등록완료": "#27AE60", "12-조건부승인": "#FDEBD0",
}


def fmt_date(v):
    if v is None: return ""
    if hasattr(v, 'strftime'): return v.strftime('%m/%d')
    return str(v)[:10]

def to_date(v):
    if v is None: return None
    if hasattr(v, 'date'): return v.date()
    if isinstance(v, datetime.date): return v
    return None


def main():
    wb = load_workbook(FILE, data_only=False)
    vehicles, items_tbl, tests_tbl = load_ref_tables(wb)

    ws = None
    for name in wb.sheetnames:
        if "Master Tracker" in name:
            ws = wb[name]
            break

    # ── Read & compute all rows (V4: 28 cols) ──
    rows = []
    for row_idx in range(5, 250):
        # B (col 2): VehicleCode
        vehicle_raw = ws.cell(row=row_idx, column=2).value
        if vehicle_raw is None:
            break
        # Skip formula-only empty rows
        if isinstance(vehicle_raw, str) and vehicle_raw.startswith("="):
            break
        vehicle = str(vehicle_raw)

        # A (col 1): No
        no_raw = ws.cell(row=row_idx, column=1).value
        no = (row_idx - 4) if isinstance(no_raw, str) and no_raw.startswith("=") else no_raw

        # D (col 4): ItemName
        item_name = ws.cell(row=row_idx, column=4).value or ""
        # F (col 6): TestNameKR
        test_name = ws.cell(row=row_idx, column=6).value or ""

        # Manual inputs (V4 columns)
        k_sub = to_date(ws.cell(row=row_idx, column=11).value)    # K: PlanSubmitted
        l_rev = ws.cell(row=row_idx, column=12).value             # L: ReviewResult
        n_act = to_date(ws.cell(row=row_idx, column=14).value)    # N: PartsReadyActual
        p_act = to_date(ws.cell(row=row_idx, column=16).value)    # P: TestStartActual
        r_act = to_date(ws.cell(row=row_idx, column=18).value)    # R: TestCompleteActual
        t_sub = to_date(ws.cell(row=row_idx, column=20).value)    # T: ReportSubmitted
        u_ovr = ws.cell(row=row_idx, column=21).value             # U: Override
        if isinstance(u_ovr, str) and (u_ovr.startswith("=") or u_ovr.strip() == ""):
            u_ovr = None

        # Auto-derive from reference tables
        item_info = items_tbl.get(str(item_name), {})
        test_info = tests_tbl.get(str(test_name), {})
        veh_info = vehicles.get(vehicle, {})

        part_cat = item_info.get('category', '')
        supplier = item_info.get('supplier', '')
        test_type = test_info.get('type', '')
        scope = test_info.get('scope', '')
        spec = test_info.get('spec', '')
        duration = test_info.get('duration', 14)

        t0 = veh_info.get('t0')
        proto = veh_info.get('proto')
        mc = veh_info.get('mc')
        deadline = veh_info.get('deadline')
        lt_plan = veh_info.get('lt_plan', 14)
        lt_parts = veh_info.get('lt_parts', 49)
        lt_test = veh_info.get('lt_test', 63)
        lt_report = veh_info.get('lt_report', 98)

        # Compute plan dates
        plan_due = (t0 + datetime.timedelta(days=lt_plan)) if t0 else None
        parts_plan = (t0 + datetime.timedelta(days=lt_parts)) if t0 else None
        test_start_plan = (t0 + datetime.timedelta(days=lt_test)) if t0 else None
        if test_type and "차량장착" in str(test_type) and proto and test_start_plan:
            test_start_plan = max(test_start_plan, proto)
        test_end_plan = (test_start_plan + datetime.timedelta(days=duration)) if test_start_plan else None
        report_due = (t0 + datetime.timedelta(days=lt_report)) if t0 else None
        if report_due and deadline:
            report_due = min(report_due, deadline)

        # Status, Progress, Risk, Action
        status = compute_status(k_sub, l_rev, n_act, p_act, r_act, t_sub, u_ovr)
        progress = PROGRESS.get(status, 0)
        risk = compute_risk(status, plan_due, report_due, t_sub, u_ovr)
        action = ACTIONS.get(status, "—")
        if status == "01-미제출" and plan_due and plan_due < TODAY:
            action = "계획서 제출 독촉 (OVERDUE)"

        days_val = "—" if t_sub else ((report_due - TODAY).days if report_due else "N/A")

        # Days since activity
        activity_dates = [d for d in [t_sub, r_act, p_act, n_act, k_sub] if d]
        stagnant = (TODAY - max(activity_dates)).days if activity_dates else "—"

        rows.append({
            'no': no, 'vehicle': vehicle, 'item_name': item_name,
            'part_cat': part_cat, 'supplier': supplier,
            'test_name': test_name, 'test_type': test_type,
            'scope': scope, 'spec': spec,
            't0': t0, 'plan_due': plan_due, 'k_sub': k_sub, 'l_rev': l_rev,
            'parts_plan': parts_plan, 'n_act': n_act,
            'test_start_plan': test_start_plan, 'p_act': p_act,
            'test_end_plan': test_end_plan, 'r_act': r_act,
            'report_due': report_due, 't_sub': t_sub,
            'u_ovr': u_ovr, 'status': status, 'progress': progress,
            'risk': risk, 'action': action, 'days': days_val,
            'stagnant': stagnant,
        })

    # ── Stats ──
    total = len(rows)
    done = sum(1 for r in rows if r['status'] == '11-결과등록완료')
    critical = sum(1 for r in rows if r['risk'] in ('CRITICAL', 'OVERDUE'))
    action_cnt = sum(1 for r in rows if r['risk'] == 'ACTION')
    warning_cnt = sum(1 for r in rows if r['risk'] == 'WARNING')
    blocked_cnt = sum(1 for r in rows if r['risk'] == 'BLOCKED')
    avg_progress = sum(r['progress'] for r in rows) / total if total else 0

    # Per-vehicle stats
    veh_stats = {}
    for v in vehicles:
        vrows = [r for r in rows if r['vehicle'] == v]
        if not vrows:
            continue
        veh_stats[v] = {
            'total': len(vrows),
            'done': sum(1 for r in vrows if r['status'] == '11-결과등록완료'),
            'critical': sum(1 for r in vrows if r['risk'] in ('CRITICAL', 'OVERDUE')),
            'warning': sum(1 for r in vrows if r['risk'] == 'WARNING'),
            'blocked': sum(1 for r in vrows if r['risk'] == 'BLOCKED'),
            'avg': sum(r['progress'] for r in vrows) / len(vrows),
            'mc': vehicles[v].get('mc'),
            'mc_days': (vehicles[v]['mc'] - TODAY).days if vehicles[v].get('mc') else 0,
        }

    # ── Build HTML ──
    h = []
    h.append("""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{box-sizing:border-box}
body{font-family:'Malgun Gothic','Segoe UI',sans-serif;background:#f0f2f5;margin:0;padding:20px;color:#2C3E50}
.container{max-width:1800px;margin:0 auto}
h1{color:#1B2A4A;font-size:22px;margin:0 0 4px}
.subtitle{color:#666;font-size:12px;margin-bottom:20px}

/* Dashboard Cards */
.dash-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:10px;margin-bottom:20px}
.card{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.08)}
.card-num{font-size:28px;font-weight:bold;color:#1B2A4A}
.card-label{font-size:11px;color:#7f8c8d;margin-top:4px}
.card-red .card-num{color:#E74C3C}
.card-green .card-num{color:#27AE60}
.card-orange .card-num{color:#F39C12}
.card-gray .card-num{color:#7f8c8d}

/* Vehicle comparison */
.veh-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:20px}
.veh-card{background:#fff;border-radius:10px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.08)}
.veh-title{font-size:14px;font-weight:bold;color:#1B2A4A;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.veh-mc{font-size:11px;padding:3px 8px;border-radius:12px;font-weight:bold}
.veh-mc.green{background:#D5F5E3;color:#27AE60}
.veh-mc.yellow{background:#FEF9E7;color:#B7950B}
.veh-mc.red{background:#FADBD8;color:#C0392B}
.veh-bar-wrap{background:#ecf0f1;border-radius:6px;height:20px;margin:8px 0}
.veh-bar{height:20px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:bold;color:#fff;min-width:30px}
.veh-stats{display:flex;gap:12px;font-size:11px;color:#7f8c8d}
.veh-stats span{font-weight:bold;color:#2C3E50}

/* Section headers */
.section-hdr{background:#1B2A4A;color:#fff;padding:8px 14px;border-radius:8px 8px 0 0;font-size:14px;font-weight:bold;margin-top:20px}

/* Tracker Table */
.tracker-wrap{overflow-x:auto;margin-bottom:30px;border-radius:0 0 8px 8px;box-shadow:0 2px 8px rgba(0,0,0,.08)}
table{border-collapse:collapse;font-size:10px;width:100%;background:#fff}
th{background:#1B2A4A;color:#fff;padding:6px 4px;border:1px solid #2C3E6B;text-align:center;font-size:9px;white-space:nowrap}
th.kr{background:#D6E4F0;color:#1B2A4A;font-size:8px;padding:3px}
th.manual{background:#B7950B;color:#fff}
th.manual-kr{background:#FFF9C4;color:#7D6608;font-size:8px;padding:3px}
td{padding:4px 3px;border:1px solid #ddd;text-align:center;white-space:nowrap;font-size:10px}
td.manual-cell{background:#FFF9C4 !important}
td.left{text-align:left}
tr:nth-child(even) td:not(.manual-cell){background:#F8FAFC}
.progress-bar{width:60px;height:12px;background:#ecf0f1;border-radius:6px;display:inline-block;vertical-align:middle}
.progress-fill{height:12px;border-radius:6px}

/* Legend */
.legend{display:flex;gap:12px;flex-wrap:wrap;margin:12px 0;font-size:11px}
.legend-item{display:flex;align-items:center;gap:4px}
.leg-dot{width:12px;height:12px;border-radius:3px;display:inline-block}
.manual-legend{background:#FFF9C4;border:1px solid #B7950B;padding:4px 10px;border-radius:4px;font-size:11px;color:#7D6608;font-weight:bold;margin-bottom:10px}

/* State dist */
.state-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:12px 0}
.state-chip{padding:6px 10px;border-radius:6px;font-size:10px;font-weight:bold;text-align:center}
</style></head><body>
<div class="container">
""")

    # Title
    h.append(f'<h1>ESIR Test Master Tracker V4</h1>')
    h.append(f'<div class="subtitle">Generated: {TODAY} | 3 Vehicles | {total} Test Items | 28 Columns | 12-State Lifecycle + BLOCKED</div>')

    # Dashboard cards (7 cards)
    h.append('<div class="dash-grid">')
    cards = [
        (total, "Total Items", ""),
        (done, "Complete", "card-green"),
        (total - done, "Remaining", ""),
        (critical, "Critical/Overdue", "card-red"),
        (blocked_cnt, "Blocked", "card-gray"),
        (action_cnt + warning_cnt, "Attention", "card-orange"),
        (f"{avg_progress:.0f}%", "Avg Progress", ""),
    ]
    for val, label, cls in cards:
        h.append(f'<div class="card {cls}"><div class="card-num">{val}</div><div class="card-label">{label}</div></div>')
    h.append('</div>')

    # Vehicle comparison
    h.append('<div class="veh-grid">')
    for v, vs in veh_stats.items():
        mc_cls = "green" if vs['mc_days'] > 90 else ("yellow" if vs['mc_days'] > 30 else "red")
        bar_pct = vs['avg']
        bar_color = "#27AE60" if bar_pct >= 70 else ("#3B5998" if bar_pct >= 30 else "#F39C12")
        comp_rate = (vs['done'] / vs['total'] * 100) if vs['total'] else 0

        h.append(f'<div class="veh-card">')
        h.append(f'<div class="veh-title">{v} <span class="veh-mc {mc_cls}">MC D-{vs["mc_days"]}</span></div>')
        h.append(f'<div class="veh-bar-wrap"><div class="veh-bar" style="width:{max(bar_pct,5)}%;background:{bar_color}">{bar_pct:.0f}%</div></div>')
        h.append(f'<div class="veh-stats">Total <span>{vs["total"]}</span> | Done <span>{vs["done"]}</span> ({comp_rate:.0f}%) | Critical <span style="color:#E74C3C">{vs["critical"]}</span> | Blocked <span>{vs["blocked"]}</span></div>')
        h.append('</div>')
    h.append('</div>')

    # Legend
    h.append('<div class="manual-legend">🟨 노란색 셀 = 수동 입력 (Manual Input)</div>')
    h.append('<div class="legend">')
    for key, (bg, fg, label) in RISK_STYLE.items():
        h.append(f'<div class="legend-item"><span class="leg-dot" style="background:{bg};border:1px solid {fg}"></span>{label}</div>')
    h.append('</div>')

    # Master Tracker table (V4: 28 data cols + computed cols)
    h.append('<div class="section-hdr">Master Tracker 마스터추적표 (V4)</div>')
    h.append('<div class="tracker-wrap"><table>')

    # EN header row (28 cols)
    en_h = [
        "No", "Vehicle", "Part\nCategory", "Item Name", "Supplier",
        "Test Name", "Test\nType", "Scope",
        "T₀", "Plan\nDue", "Plan\nSubmit", "Review\nResult",
        "Parts\n(Plan)", "Parts\n(Actual)", "Test Start\n(Plan)", "Test Start\n(Actual)",
        "Test End\n(Plan)", "Test End\n(Actual)", "Report\nDue", "Report\nSubmit",
        "Override", "Status", "Progress", "D-Day", "Risk", "Stagnant", "Next Action", "Note",
    ]
    # 0-indexed manual columns: B(1), D(3), F(5), K(10), L(11), N(13), P(15), R(17), T(19), U(20), AB(27)
    manual_en = {1, 3, 5, 10, 11, 13, 15, 17, 19, 20, 27}

    h.append('<tr>')
    for i, hdr in enumerate(en_h):
        cls = ' class="manual"' if i in manual_en else ''
        h.append(f'<th{cls}>{html.escape(hdr).replace(chr(10),"<br>")}</th>')
    h.append('</tr>')

    # KR header row
    kr_h = [
        "번호", "차종", "파트구분", "아이템명", "협력사",
        "시험명", "시험유형", "범위",
        "도면배포", "계획기한", "계획제출", "검토결과",
        "부품(계획)", "부품(실적)", "시험시작(계획)", "시험시작(실적)",
        "시험완료(계획)", "시험완료(실적)", "성적서기한", "성적서제출",
        "오버라이드", "상태", "진행률", "잔여일", "리스크", "정체일수", "다음조치", "비고",
    ]

    h.append('<tr>')
    for i, hdr in enumerate(kr_h):
        cls = ' class="manual-kr"' if i in manual_en else ' class="kr"'
        h.append(f'<th{cls}>{html.escape(hdr)}</th>')
    h.append('</tr>')

    # Data rows
    for r in rows:
        h.append('<tr>')
        cells = [
            (r['no'], False),
            (r['vehicle'], True),
            (r['part_cat'], False),
            (r['item_name'], True),
            (r['supplier'], False),
            (r['test_name'], True),
            (r['test_type'], False),
            (r['scope'], False),
            (fmt_date(r['t0']), False),
            (fmt_date(r['plan_due']), False),
            (fmt_date(r['k_sub']), True),
            (r['l_rev'] or "", True),
            (fmt_date(r['parts_plan']), False),
            (fmt_date(r['n_act']), True),
            (fmt_date(r['test_start_plan']), False),
            (fmt_date(r['p_act']), True),
            (fmt_date(r['test_end_plan']), False),
            (fmt_date(r['r_act']), True),
            (fmt_date(r['report_due']), False),
            (fmt_date(r['t_sub']), True),
            (r['u_ovr'] or "", True),
        ]

        for val, is_manual in cells:
            mc = ' class="manual-cell"' if is_manual else ''
            txt = html.escape(str(val or ""))
            h.append(f'<td{mc}>{txt}</td>')

        # Status
        st = r['status']
        st_bg = STATUS_BG.get(st, "#FFF")
        st_fg = "#FFF" if st in ("06-BIW중단", "11-결과등록완료") else "#2C3E50"
        h.append(f'<td style="background:{st_bg};color:{st_fg};font-weight:bold;font-size:9px">{html.escape(st)}</td>')

        # Progress
        p = r['progress']
        bar_c = "#27AE60" if p >= 80 else ("#3B5998" if p >= 40 else ("#F39C12" if p > 0 else "#BDC3C7"))
        h.append(f'<td><div class="progress-bar"><div class="progress-fill" style="width:{p}%;background:{bar_c}"></div></div> {p}%</td>')

        # Days
        d = r['days']
        ds = ""
        if isinstance(d, (int, float)):
            if d <= 0: ds = "color:#E74C3C;font-weight:bold"
            elif d <= 14: ds = "color:#B7950B;font-weight:bold"
        h.append(f'<td style="{ds}">{d}</td>')

        # Risk
        rbg, rfg, rlabel = RISK_STYLE.get(r['risk'], ("#FFF", "#000", "?"))
        h.append(f'<td style="background:{rbg};color:{rfg};font-weight:bold;font-size:9px">{rlabel}</td>')

        # Stagnant days
        sg = r['stagnant']
        sg_style = "color:#E74C3C;font-weight:bold" if isinstance(sg, int) and sg > 7 else ""
        h.append(f'<td style="{sg_style}">{sg}</td>')

        # Action
        h.append(f'<td class="left" style="font-size:9px">{html.escape(r["action"])}</td>')

        # Note (empty)
        h.append(f'<td class="manual-cell"></td>')

        h.append('</tr>')

    h.append('</table></div>')

    # State distribution
    h.append('<div class="section-hdr">State Distribution 상태별 분포</div>')
    state_counts = Counter(r['status'] for r in rows)
    h.append('<div class="state-grid">')
    for st in sorted(state_counts.keys()):
        bg = STATUS_BG.get(st, "#FFF")
        fg = "#FFF" if st in ("06-BIW중단", "11-결과등록완료") else "#2C3E50"
        h.append(f'<div class="state-chip" style="background:{bg};color:{fg}">{html.escape(st)}: {state_counts[st]}</div>')
    h.append('</div>')

    # Item × Vehicle summary
    h.append('<div class="section-hdr">Item Progress 아이템별 현황</div>')
    h.append('<div style="background:#fff;padding:14px;border-radius:0 0 8px 8px;box-shadow:0 2px 8px rgba(0,0,0,.08)">')
    h.append('<table style="width:auto"><tr><th>Vehicle</th><th>Item</th><th>Tests</th><th>Done</th><th>Avg Progress</th></tr>')

    item_groups = {}
    for r in rows:
        key = f"{r['vehicle']}:{r['item_name']}"
        if key not in item_groups:
            item_groups[key] = {'vehicle': r['vehicle'], 'item': r['item_name'], 'rows': []}
        item_groups[key]['rows'].append(r)

    for key in sorted(item_groups.keys()):
        info = item_groups[key]
        cnt = len(info['rows'])
        done_cnt = sum(1 for r in info['rows'] if r['status'] == '11-결과등록완료')
        avg_p = sum(r['progress'] for r in info['rows']) / cnt if cnt else 0
        bg = "#D5F5E3" if avg_p >= 80 else ("#FEF9E7" if avg_p >= 30 else "#FFF")
        h.append(f'<tr><td style="font-weight:bold">{html.escape(info["vehicle"])}</td>'
                 f'<td>{html.escape(info["item"])}</td>'
                 f'<td style="text-align:center">{cnt}</td>'
                 f'<td style="text-align:center">{done_cnt}</td>'
                 f'<td style="background:{bg};text-align:center;font-weight:bold">{avg_p:.0f}%</td></tr>')
    h.append('</table></div>')

    h.append('</div></body></html>')

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(h))
    print(f"[OK] HTML preview: {OUT}")


if __name__ == "__main__":
    main()
