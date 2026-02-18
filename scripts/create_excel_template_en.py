"""Generate KPI Action Plan Excel template (English version) for team leaders."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Styles ──
HEADER_FILL = PatternFill(start_color="1A1D25", end_color="1A1D25", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
GUIDE_FONT = Font(name="Calibri", size=9, color="6B7280")
DATA_FONT = Font(name="Calibri", size=10)
MILESTONE_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════
# Sheet 1: Action Plans
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Action Plans"

COLUMNS = [
    ("A", "Parent KPI\n(Team KPI Title)", 25),
    ("B", "Action Title", 30),
    ("C", "Icon", 6),
    ("D", "Description", 35),
    ("E", "Owner", 12),
    ("F", "Target", 8),
    ("G", "Unit", 8),
    ("H", "Priority", 10),
    ("I", "Start Date\n(YYYY-MM-DD)", 14),
    ("J", "Due Date\n(YYYY-MM-DD)", 14),
    ("K", "Output Type", 18),
    ("L", "Weight\n(0~1)", 8),
    ("M", "Milestone 1", 22),
    ("N", "Milestone 2", 22),
    ("O", "Milestone 3", 22),
    ("P", "Milestone 4", 22),
    ("Q", "Milestone 5", 22),
]

# Title row
ws.merge_cells("A1:Q1")
title_cell = ws["A1"]
title_cell.value = "KPI Cascade — Action Plan Template"
title_cell.font = Font(name="Calibri", bold=True, size=14, color="8B5CF6")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Info row
ws.merge_cells("A2:Q2")
info_cell = ws["A2"]
info_cell.value = "Team:                    Author:                    Date:          "
info_cell.font = Font(name="Calibri", size=10, color="6B7280")
info_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 24

# Header row (row 3)
for col_letter, label, width in COLUMNS:
    cell = ws[f"{col_letter}3"]
    cell.value = label
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 40

# Subheader (row 4) — field descriptions
DESCRIPTIONS = [
    "Your team's\nassigned KPI",
    "Specific action\n(use verb form)",
    "e.g.\n📞🔍📝",
    "Brief description\n(1-2 sentences)",
    "Responsible\nperson",
    "Numeric\ntarget",
    "items, %,\npts, etc.",
    "low/med\nhigh/crit",
    "Start\ndate",
    "Target\ncompletion",
    "Report/System\n/Training/etc.",
    "Weight\nwithin KPI",
    "Step 1\ncheckpoint",
    "Step 2\ncheckpoint",
    "Step 3\ncheckpoint",
    "Step 4\n(optional)",
    "Step 5\n(optional)",
]

for i, (col_letter, _, _) in enumerate(COLUMNS):
    cell = ws[f"{col_letter}4"]
    cell.value = DESCRIPTIONS[i]
    cell.font = GUIDE_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws.row_dimensions[4].height = 36

# Example rows (5-6)
EXAMPLE_1 = [
    "Sales Revenue Growth", "Q2 Key Account Meetings (5 clients)", "📞",
    "Direct meetings with strategic clients to expand sales pipeline",
    "Y.S. Kim", 5, "meetings", "high",
    "2026-02-01", "2026-06-30", "Meeting completion report",
    0.40,
    "Finalize target client list", "Complete initial outreach",
    "Confirm meeting schedule", "Send proposals",
    "Complete meetings & report results",
]

EXAMPLE_2 = [
    "Marketing Lead Gen", "Competitor Benchmarking Report", "🔍",
    "Analyze strengths/weaknesses vs. 3 key competitors and derive strategy",
    "S.J. Lee", 1, "report", "medium",
    "2026-03-01", "2026-04-15", "Analysis report (PPT)",
    0.30,
    "Define scope & target competitors", "Complete primary data collection",
    "Build analysis framework & draft", "Internal review & feedback",
    "Finalize report & distribute",
]

for row_idx, example in [(5, EXAMPLE_1), (6, EXAMPLE_2)]:
    for col_idx, value in enumerate(example):
        col_letter = COLUMNS[col_idx][0]
        cell = ws[f"{col_letter}{row_idx}"]
        cell.value = value
        cell.font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")
        cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
        cell.border = THIN_BORDER
        if col_idx >= 12:
            cell.fill = MILESTONE_FILL
    ws.row_dimensions[row_idx].height = 32

# Empty data rows (7-26)
for row_idx in range(7, 27):
    for col_idx, (col_letter, _, _) in enumerate(COLUMNS):
        cell = ws[f"{col_letter}{row_idx}"]
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
        if col_idx >= 12:
            cell.fill = MILESTONE_FILL
    ws.row_dimensions[row_idx].height = 28

# Data validations
priority_dv = DataValidation(type="list", formula1='"low,medium,high,critical"')
priority_dv.error = "Select one of: low, medium, high, critical"
priority_dv.errorTitle = "Invalid Priority"
ws.add_data_validation(priority_dv)
priority_dv.add("H5:H26")

unit_dv = DataValidation(type="list", formula1='"items,%,pts,USD,KRW,people,sessions,reports"')
ws.add_data_validation(unit_dv)
unit_dv.add("G5:G26")

ws.freeze_panes = "A5"

# ═══════════════════════════════════════════════
# Sheet 2: Guide
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Guide")

guide_lines = [
    ("KPI Cascade — Action Plan Writing Guide", True, 14),
    ("", False, 10),
    ("1. Structure", True, 12),
    ("   Strategic Goal (depth-0) > Team KPI (depth-1) > Action Plan (depth-2)", False, 10),
    ("   Team leaders fill in depth-2 items in the 'Action Plans' sheet.", False, 10),
    ("", False, 10),
    ("2. Required Elements (every action must include all of these)", True, 12),
    ("   a) Owner — One responsible person per action", False, 10),
    ("   b) Due Date — In YYYY-MM-DD format", False, 10),
    ("   c) Output Type — Concrete deliverable (report, system, training, etc.)", False, 10),
    ("   d) Target Value — Quantitative goal with unit", False, 10),
    ("   e) 3-5 Milestones — Checkpoints for interim review & feedback", False, 10),
    ("", False, 10),
    ("3. Milestone Writing Principles", True, 12),
    ("   - Each step must be verifiable as done/not done", False, 10),
    ("   - Even research tasks should be broken into 3-5 steps", False, 10),
    ("   - Use completion verbs: 'Complete...', 'Finalize...', 'Submit...'", False, 10),
    ("", False, 10),
    ("4. Weight Guide", True, 12),
    ("   - Weights of actions under the same Team KPI must sum to 1.0", False, 10),
    ("   - Example: Action A (0.4) + Action B (0.35) + Action C (0.25) = 1.0", False, 10),
    ("   - Assign higher weights to more impactful actions", False, 10),
    ("", False, 10),
    ("5. Milestone Breakdown Examples", True, 12),
    ("", False, 10),
    ("   [Research Type] 'Market Research Report'", False, 10),
    ("     M1: Define research scope & methodology", False, 10),
    ("     M2: Complete primary data collection", False, 10),
    ("     M3: Build analysis & write first draft", False, 10),
    ("     M4: Incorporate internal review feedback", False, 10),
    ("     M5: Finalize & distribute report", False, 10),
    ("", False, 10),
    ("   [Build Type] 'Process Automation System'", False, 10),
    ("     M1: Analyze current process & identify gaps", False, 10),
    ("     M2: Select tools & design architecture", False, 10),
    ("     M3: Develop prototype & internal testing", False, 10),
    ("     M4: Write manual & conduct team training", False, 10),
    ("     M5: Deploy to production", False, 10),
    ("", False, 10),
    ("   [Engagement Type] 'Key Account Meetings'", False, 10),
    ("     M1: Finalize target client list", False, 10),
    ("     M2: Complete initial outreach (email/call)", False, 10),
    ("     M3: Confirm meeting schedule", False, 10),
    ("     M4: Send proposals", False, 10),
    ("     M5: Complete meetings & submit results report", False, 10),
    ("", False, 10),
    ("6. Submission", True, 12),
    ("   - Fill in the 'Action Plans' sheet and reply with this file", False, 10),
    ("   - Example rows (gray italic) are for reference — feel free to delete them", False, 10),
    ("   - Milestones 4 & 5 are optional (minimum 3 required)", False, 10),
    ("", False, 10),
    ("7. Philosophy", True, 12),
    ("   \"Break every task into achievable milestones so that every team member", False, 10),
    ("    experiences the success of accomplishing their own goals.", False, 10),
    ("    That is the power of organizational culture.\"", False, 10),
]

for i, (text, bold, size) in enumerate(guide_lines, 1):
    cell = ws2[f"A{i}"]
    cell.value = text
    cell.font = Font(name="Calibri", bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True)

ws2.column_dimensions["A"].width = 80

# Save
output_path = r"C:\obsidian\0_Inbox\KPI_Action_Plan_Template.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
