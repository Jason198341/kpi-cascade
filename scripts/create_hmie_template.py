"""Generate HMIE Vehicle Development KPI Action Plan Excel templates (KO + EN)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Shared Styles ──
HEADER_FILL = PatternFill(start_color="1A1D25", end_color="1A1D25", fill_type="solid")
DEPTH0_FILL = PatternFill(start_color="2D1B69", end_color="2D1B69", fill_type="solid")  # purple
DEPTH1_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # blue
SUBHEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
MILESTONE_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
BIZ_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")  # light indigo
PPL_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")  # light orange
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def make_workbook(lang="ko"):
    """Create a workbook for the given language."""
    wb = openpyxl.Workbook()
    is_ko = lang == "ko"
    FONT_NAME = "맑은 고딕" if is_ko else "Calibri"

    # ═══════════════════════════════════════
    # Sheet 1: KPI 전략 구조 / Strategy Map
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "전략구조" if is_ko else "Strategy Map"
    ws1.sheet_properties.tabColor = "8B5CF6"

    # Title
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value = "2026 차량개발실 KPI 캐스케이드 전략 구조" if is_ko else "2026 Vehicle Development Division — KPI Cascade Strategy"
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="8B5CF6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 44

    # Strategy overview headers
    headers_map = [
        ("A", "구분" if is_ko else "Category", 12),
        ("B", "분야" if is_ko else "Area", 30),
        ("C", "비중" if is_ko else "Weight", 8),
        ("D", "핵심 내용" if is_ko else "Key Description", 60),
        ("E", "이모지" if is_ko else "Icon", 6),
        ("F", "관련 팀 KPI 수" if is_ko else "# Team KPIs", 12),
    ]
    for col, label, w in headers_map:
        cell = ws1[f"{col}3"]
        cell.value = label
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws1.column_dimensions[col].width = w
    ws1.row_dimensions[3].height = 32

    # ── depth-0 rows ──
    d0_rows = [
        {
            "cat": "Business\n(80%)",
            "areas": [
                ("현지 특화 사양 경쟁력 강화" if is_ko else "Local Spec Competitiveness", 0.25,
                 ("26~27년 양산 차종 현지 특화 사양 리스트업, 중장기 적용 로드맵, 권역 우선순위 정리" if is_ko
                  else "List local specs for 26-27 MY vehicles, mid/long-term roadmap, prioritize with regional teams"),
                 "🎯", 3),
                ("경제형 신기술 개발" if is_ko else "Economic New Tech Development", 0.20,
                 ("현지 주도 신기술, 남양 콜라보 항목, 모터 현지화 등 부품 현지화" if is_ko
                  else "Locally-led new tech, Namyang collab items, motor localization & parts localization"),
                 "🔬", 3),
                ("개발 강건화" if is_ko else "Development Robustness", 0.20,
                 ("충돌 교류회, DPA 체크리스트, EO 사례 예방 설계, ESIR 스케줄, 도면 완성도" if is_ko
                  else "Crash review, DPA checklist, EO-based preventive design, ESIR schedule, drawing completeness"),
                 "🛡️", 5),
                ("부품 공용화 및 Pre SR 프로세스화" if is_ko else "Parts Commonization & Pre SR Process", 0.20,
                 ("1레벨 마스터리스트 → 구매단가 정리 → 신규 vs C/O 비교 → 설계 역제안 프로세스 전차종 구축" if is_ko
                  else "L1 master list → purchase price → new vs C/O comparison → design counter-proposal process for all models"),
                 "🔄", 4),
                ("CCW 활동 (원가 절감)" if is_ko else "CCW Activities (Cost Reduction)", 0.15,
                 ("양산차 원가 절감 활동 지속 추진" if is_ko
                  else "Continuous cost reduction activities for production vehicles"),
                 "💰", 2),
            ],
        },
        {
            "cat": "People\n(20%)",
            "areas": [
                ("인재육성" if is_ko else "Talent Development", 1.0,
                 ("SW 인재육성, PCL, E2E, 매뉴얼, 컬처서베이, IIT 세미나, 테크쇼, 교육, 상용센터 지원" if is_ko
                  else "SW talent, PCL, E2E, manuals, culture survey, IIT seminar, tech show, training, commercial center support"),
                 "🎓", 10),
            ],
        },
    ]

    row = 4
    for d0 in d0_rows:
        start_row = row
        for i, (area, weight, desc, emoji, kpi_count) in enumerate(d0["areas"]):
            ws1[f"B{row}"].value = area
            ws1[f"B{row}"].font = Font(name=FONT_NAME, bold=True, size=10)
            ws1[f"C{row}"].value = f"{weight:.0%}"
            ws1[f"C{row}"].font = Font(name=FONT_NAME, size=10)
            ws1[f"D{row}"].value = desc
            ws1[f"D{row}"].font = Font(name=FONT_NAME, size=9, color="4B5563")
            ws1[f"E{row}"].value = emoji
            ws1[f"F{row}"].value = kpi_count
            ws1[f"F{row}"].font = Font(name=FONT_NAME, size=10)
            for col in "BCDEF":
                ws1[f"{col}{row}"].alignment = CENTER if col in "CEF" else LEFT_WRAP
                ws1[f"{col}{row}"].border = THIN_BORDER
                if d0["cat"].startswith("Business"):
                    ws1[f"{col}{row}"].fill = BIZ_FILL
                else:
                    ws1[f"{col}{row}"].fill = PPL_FILL
            ws1.row_dimensions[row].height = 36
            row += 1
        end_row = row - 1
        ws1.merge_cells(f"A{start_row}:A{end_row}")
        ws1[f"A{start_row}"].value = d0["cat"]
        ws1[f"A{start_row}"].font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
        ws1[f"A{start_row}"].fill = DEPTH0_FILL if d0["cat"].startswith("Business") else PatternFill(start_color="7C2D12", end_color="7C2D12", fill_type="solid")
        ws1[f"A{start_row}"].alignment = CENTER
        ws1[f"A{start_row}"].border = THIN_BORDER

    # ═══════════════════════════════════════
    # Sheet 2: 액션 플랜 / Action Plans
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("액션플랜" if is_ko else "Action Plans")
    ws2.sheet_properties.tabColor = "10B981"

    COLUMNS = [
        ("A", ("상위 KPI\n(분야명)" if is_ko else "Parent KPI\n(Area)"), 28),
        ("B", ("액션 제목" if is_ko else "Action Title"), 32),
        ("C", ("이모지" if is_ko else "Icon"), 6),
        ("D", ("설명 / 아웃풋" if is_ko else "Description / Output"), 38),
        ("E", ("담당자" if is_ko else "Owner"), 12),
        ("F", ("목표값" if is_ko else "Target"), 8),
        ("G", ("단위" if is_ko else "Unit"), 8),
        ("H", ("우선순위" if is_ko else "Priority"), 10),
        ("I", ("시작일" if is_ko else "Start Date"), 13),
        ("J", ("마감일" if is_ko else "Due Date"), 13),
        ("K", ("가중치\n(0~1)" if is_ko else "Weight\n(0~1)"), 8),
        ("L", ("마일스톤 1" if is_ko else "Milestone 1"), 24),
        ("M", ("마일스톤 2" if is_ko else "Milestone 2"), 24),
        ("N", ("마일스톤 3" if is_ko else "Milestone 3"), 24),
        ("O", ("마일스톤 4" if is_ko else "Milestone 4"), 24),
        ("P", ("마일스톤 5" if is_ko else "Milestone 5"), 24),
    ]

    # Title
    ws2.merge_cells("A1:P1")
    t = ws2["A1"]
    t.value = "2026 차량개발실 — 팀별 액션 플랜" if is_ko else "2026 Vehicle Dev. — Team Action Plans"
    t.font = Font(name=FONT_NAME, bold=True, size=14, color="10B981")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    # Info
    ws2.merge_cells("A2:P2")
    ws2["A2"].value = ("팀명:                    작성자:                    작성일:          " if is_ko
                       else "Team:                    Author:                    Date:          ")
    ws2["A2"].font = Font(name=FONT_NAME, size=10, color="6B7280")
    ws2["A2"].alignment = LEFT_WRAP
    ws2.row_dimensions[2].height = 24

    # Headers
    for col, label, w in COLUMNS:
        cell = ws2[f"{col}3"]
        cell.value = label
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[3].height = 38

    # ── Pre-filled example actions ──
    if is_ko:
        ACTIONS = [
            # ── 분야 1: 현지 특화 사양 ──
            ["현지 특화 사양 경쟁력 강화", "26~27MY 현지 특화 사양 리스트업", "📋",
             "[아웃풋: 사양 리스트 Excel] 양산 차종별 현지 특화 사양 현황 정리 및 Gap 분석",
             "", 1, "건", "high", "2026-01-06", "2026-03-31", 0.35,
             "차종별 현행 사양 수집", "경쟁사 대비 Gap 분석", "권역 VoC 반영 우선순위 정리",
             "최종 리스트 확정 및 공유", ""],
            ["현지 특화 사양 경쟁력 강화", "중장기 사양 적용 로드맵 수립", "🗺️",
             "[아웃풋: 로드맵 PPT] 28~30년 차종별 현지 특화 사양 적용 마스터플랜",
             "", 1, "건", "high", "2026-02-01", "2026-05-31", 0.35,
             "벤치마킹 데이터 수집 (인도/동남아)", "사양별 원가·효과 분석", "차종별 적용 우선순위 매트릭스",
             "로드맵 초안 작성 및 내부 리뷰", "최종 로드맵 확정"],
            ["현지 특화 사양 경쟁력 강화", "권역 소통 및 우선순위 정리", "🤝",
             "[아웃풋: 합의 의사록] 권역 담당자와 사양 우선순위 협의 및 확정",
             "", 4, "회", "medium", "2026-01-15", "2026-06-30", 0.30,
             "권역 담당자 리스트 확정", "1차 협의 (현황 공유)", "2차 협의 (우선순위 논의)",
             "최종 합의 및 의사록 배포", ""],

            # ── 분야 2: 경제형 신기술 ──
            ["경제형 신기술 개발", "인도 현지 주도 신기술 개발 항목 추진", "🔬",
             "[아웃풋: 기술 검증 보고서] 현지 주도 신기술 아이템 발굴 및 PoC 완료",
             "", 3, "건", "high", "2026-01-06", "2026-06-30", 0.40,
             "신기술 아이템 후보 리스트업", "타당성 분석 및 우선순위 선정", "PoC 계획 수립 및 착수",
             "PoC 결과 보고서 작성", "양산 적용 판단 및 후속 계획"],
            ["경제형 신기술 개발", "남양 콜라보 기술 개발 항목 진행", "🔗",
             "[아웃풋: 공동개발 결과보고] 남양 R&D와 협업 기술 항목 개발 진행",
             "", 2, "건", "high", "2026-02-01", "2026-06-30", 0.35,
             "남양 협업 항목 확정 및 R&R 정의", "1차 중간 점검 (기술 리뷰)", "시제품/시뮬레이션 검증",
             "최종 결과 보고 및 양산 이관 판단", ""],
            ["경제형 신기술 개발", "모터 및 부품 현지화 추진", "🏭",
             "[아웃풋: 현지화 진행 현황표] 모터 현지화 및 핵심 부품 현지 소싱 추진",
             "", 2, "건", "medium", "2026-03-01", "2026-06-30", 0.25,
             "현지화 대상 부품 선정", "현지 업체 소싱 및 기술 평가", "시제품 제작 및 품질 검증",
             "원가 비교 분석 보고", ""],

            # ── 분야 3: 개발 강건화 ──
            ["개발 강건화", "충돌 교류회 운영", "💥",
             "[아웃풋: 교류회 결과보고] 정기 충돌 기술 교류회 개최 및 노하우 공유",
             "", 4, "회", "high", "2026-01-15", "2026-06-30", 0.20,
             "교류회 연간 일정 수립", "1차 교류회 개최 및 결과 정리", "개선 사항 도출 및 적용 계획",
             "2차 교류회 개최", "연간 성과 정리 보고"],
            ["개발 강건화", "DPA 체크리스트 강화", "✅",
             "[아웃풋: 강화된 DPA 체크리스트] 기존 DPA 체크리스트 보완 및 차종 적용",
             "", 1, "건", "high", "2026-02-01", "2026-04-30", 0.20,
             "현행 DPA 항목 문제점 분석", "선진사 사례 벤치마킹", "개선 항목 도출 및 체크리스트 개정",
             "파일럿 차종 적용 및 검증", "전차종 배포"],
            ["개발 강건화", "EO 사례 기반 출도 전 예방 설계 강화", "🔍",
             "[아웃풋: 예방설계 가이드] EO 사례 DB화 및 출도 전 예방 설계 프로세스 구축",
             "", 1, "건", "high", "2026-01-15", "2026-05-31", 0.20,
             "최근 2년 EO 사례 수집 및 분류", "재발 방지 설계 가이드 초안 작성", "부문별 리뷰 및 피드백 반영",
             "출도 전 체크 프로세스 확정", "가이드 배포 및 교육"],
            ["개발 강건화", "ESIR 스케줄 관리 프로세스 강화", "📅",
             "[아웃풋: 프로세스 개선안] ESIR 일정 준수율 향상을 위한 관리 체계 개선",
             "", 95, "%", "medium", "2026-02-01", "2026-06-30", 0.20,
             "현행 ESIR 일정 준수율 분석", "지연 원인 파악 및 개선 방안 수립", "관리 도구/대시보드 구축",
             "월간 모니터링 체계 운영", ""],
            ["개발 강건화", "도면 완성도 향상", "📐",
             "[아웃풋: 도면 품질 리포트] 출도 도면 완성도 지표 향상 활동",
             "", 90, "%", "high", "2026-01-06", "2026-06-30", 0.20,
             "도면 완성도 기준 지표 정의", "현행 수준 측정 및 Gap 분석", "부문별 개선 활동 추진",
             "월간 도면 완성도 모니터링", "목표 달성 검증 및 보고"],

            # ── 분야 4: 부품 공용화 ──
            ["부품 공용화 및 Pre SR 프로세스화", "1레벨 마스터리스트 작성", "📊",
             "[아웃풋: 마스터리스트 Excel] 베이스 차종 기준 1레벨 부품 마스터리스트",
             "", 1, "건", "critical", "2026-01-15", "2026-03-31", 0.30,
             "베이스 차종 선정 및 BOM 확보", "1레벨 부품 리스트 정리", "부품별 사양/도면번호 매핑",
             "내부 검증 및 확정", ""],
            ["부품 공용화 및 Pre SR 프로세스화", "구매 단가 정리 및 신규 vs C/O 비교", "💹",
             "[아웃풋: 비교 분석표] 부품별 신규 개발 단가 vs C/O 단가 비교 및 절감 효과 산출",
             "", 1, "건", "high", "2026-03-01", "2026-05-15", 0.30,
             "마스터리스트 부품 구매단가 수집", "신규 개발 시 예상 단가 산출", "C/O 적용 시 비용 산출",
             "절감 효과 비교표 작성", "경영층 보고"],
            ["부품 공용화 및 Pre SR 프로세스화", "설계단 역제안 프로세스 구축", "🔄",
             "[아웃풋: 프로세스 가이드] 디자인 선정모델 前 설계단 역제안 프로세스 전차종 적용",
             "", 1, "건", "high", "2026-02-01", "2026-05-31", 0.25,
             "KS2 역제안 사례 분석 및 정리", "표준 프로세스 초안 설계", "파일럿 차종 적용 및 피드백",
             "프로세스 확정 및 가이드 배포", "전차종 적용 킥오프"],
            ["부품 공용화 및 Pre SR 프로세스화", "공용화 성과 모니터링 체계 운영", "📈",
             "[아웃풋: 월간 대시보드] 공용화 진행률 및 절감 효과 월간 트래킹",
             "", 6, "회", "medium", "2026-03-01", "2026-06-30", 0.15,
             "모니터링 지표 및 대시보드 설계", "월간 데이터 수집 체계 구축", "1차 리포트 발행",
             "3개월 누적 성과 분석", ""],

            # ── 분야 5: CCW ──
            ["CCW 활동 (원가 절감)", "양산차 CCW 절감 목표 달성", "💰",
             "[아웃풋: CCW 실적 리포트] 양산 차종별 원가 절감 활동 추진 및 실적 관리",
             "", 100, "%", "high", "2026-01-06", "2026-06-30", 0.60,
             "차종별 CCW 절감 목표 확정", "1Q 절감 활동 추진 및 실적 집계", "중간 점검 및 보완 계획 수립",
             "2Q 절감 활동 추진", "상반기 종합 실적 보고"],
            ["CCW 활동 (원가 절감)", "원가 절감 아이디어 발굴 및 공유", "💡",
             "[아웃풋: 아이디어 DB] 부문별 원가 절감 아이디어 수집 및 베스트 프랙티스 공유",
             "", 20, "건", "medium", "2026-02-01", "2026-06-30", 0.40,
             "부문별 아이디어 수집 (월 1회)", "아이디어 DB 구축", "베스트 프랙티스 선정 및 공유회 개최",
             "적용 가능 항목 실행 계획 수립", ""],

            # ── People: 인재육성 ──
            ["인재육성", "SW 인재육성 프로그램 운영", "💻",
             "[아웃풋: 교육 이수 현황표] SW 역량 강화 교육 과정 운영 및 이수율 관리",
             "", 90, "%", "high", "2026-01-15", "2026-06-30", 0.12,
             "교육 대상자 및 과정 확정", "교육 일정 수립 및 공지", "교육 진행 (분기별)",
             "이수율 모니터링 및 미이수자 관리", ""],
            ["인재육성", "PCL 시스템 운영", "🔧",
             "[아웃풋: PCL 운영 보고서] PCL 시스템 안정적 운영 및 활용도 향상",
             "", 1, "건", "medium", "2026-01-06", "2026-06-30", 0.10,
             "현행 PCL 운영 현황 점검", "개선 필요 사항 도출", "시스템 업데이트 및 교육",
             "활용도 모니터링", ""],
            ["인재육성", "E2E (Error to Excellence) 활동", "🎯",
             "[아웃풋: E2E 사례집] 오류 사례 → 우수 사례 전환 활동 및 공유",
             "", 10, "건", "medium", "2026-02-01", "2026-06-30", 0.10,
             "오류 사례 수집 (부문별)", "원인 분석 및 개선 방안 도출", "개선 적용 및 효과 검증",
             "우수 사례 정리 및 사례집 발행", ""],
            ["인재육성", "업무 매뉴얼 작성", "📖",
             "[아웃풋: 업무 매뉴얼] 주요 업무 프로세스 표준 매뉴얼 작성",
             "", 5, "건", "medium", "2026-02-01", "2026-06-30", 0.08,
             "매뉴얼 작성 대상 업무 선정", "담당자별 초안 작성", "내부 리뷰 및 수정",
             "최종 매뉴얼 발행 및 배포", ""],
            ["인재육성", "사내문화 향상 (컬처 서베이)", "🏢",
             "[아웃풋: 서베이 결과 및 개선안] 컬처 서베이 실시 및 개선 활동 추진",
             "", 1, "건", "medium", "2026-03-01", "2026-06-30", 0.10,
             "서베이 문항 설계", "서베이 실시 및 데이터 수집", "결과 분석 및 개선 포인트 도출",
             "개선 활동 실행", "후속 서베이 계획"],
            ["인재육성", "IIT 교수 초청 세미나", "🎓",
             "[아웃풋: 세미나 결과보고] IIT 교수 초청 기술 세미나 개최",
             "", 2, "회", "medium", "2026-02-15", "2026-06-30", 0.10,
             "세미나 주제 및 초청 교수 선정", "일정 확정 및 참가자 모집", "세미나 개최",
             "결과 정리 및 후속 협업 논의", ""],
            ["인재육성", "협력사 및 시스템별 테크쇼 개최", "🏗️",
             "[아웃풋: 테크쇼 결과보고] 협력사·시스템별 기술 전시/공유회 개최",
             "", 2, "회", "medium", "2026-03-01", "2026-06-30", 0.10,
             "테크쇼 주제 및 참여 업체 선정", "일정 확정 및 부스 구성", "테크쇼 개최",
             "우수 기술 선정 및 적용 검토", ""],
            ["인재육성", "공통·특화 교육 운영", "📚",
             "[아웃풋: 교육 실적표] 공통 역량 + 부문별 특화 교육 과정 운영",
             "", 10, "회", "medium", "2026-01-15", "2026-06-30", 0.10,
             "교육 과정 및 일정 수립", "1Q 교육 실시", "중간 이수율 점검",
             "2Q 교육 실시", "상반기 교육 성과 보고"],
            ["인재육성", "설계 강건화 교육", "🛡️",
             "[아웃풋: 교육 이수 현황] 설계 품질 강건화 전문 교육 운영",
             "", 4, "회", "high", "2026-02-01", "2026-06-30", 0.10,
             "교육 커리큘럼 설계", "강사 섭외 및 일정 확정", "교육 실시 (분기 1회)",
             "교육 효과 평가", ""],
            ["인재육성", "상용 센터 (B04 차종) 지원", "🚛",
             "[아웃풋: 지원 실적 보고서] B04 차종 개발 기술 지원 활동",
             "", 1, "건", "medium", "2026-01-06", "2026-06-30", 0.10,
             "지원 범위 및 R&R 확정", "기술 지원 활동 수행", "중간 점검 및 이슈 공유",
             "지원 완료 및 성과 정리", ""],
        ]
    else:
        # English version
        ACTIONS = [
            # ── Area 1: Local Spec ──
            ["Local Spec Competitiveness", "List local specs for 26-27MY vehicles", "📋",
             "[Output: Spec List Excel] Compile local spec status per vehicle model & gap analysis",
             "", 1, "report", "high", "2026-01-06", "2026-03-31", 0.35,
             "Collect current specs by model", "Gap analysis vs competitors", "Prioritize with regional VoC",
             "Finalize & distribute list", ""],
            ["Local Spec Competitiveness", "Mid/long-term spec application roadmap", "🗺️",
             "[Output: Roadmap PPT] Master plan for local spec application across 28-30MY models",
             "", 1, "report", "high", "2026-02-01", "2026-05-31", 0.35,
             "Collect benchmarking data (India/SEA)", "Cost-benefit analysis per spec", "Build priority matrix by model",
             "Draft roadmap & internal review", "Finalize roadmap"],
            ["Local Spec Competitiveness", "Regional coordination & priority alignment", "🤝",
             "[Output: Agreed minutes] Coordinate with regional teams to align spec priorities",
             "", 4, "sessions", "medium", "2026-01-15", "2026-06-30", 0.30,
             "Finalize regional contact list", "1st meeting (status sharing)", "2nd meeting (priority discussion)",
             "Final agreement & distribute minutes", ""],

            # ── Area 2: New Tech ──
            ["Economic New Tech Development", "Locally-led new technology items", "🔬",
             "[Output: Tech validation report] Discover & PoC locally-led new tech items",
             "", 3, "items", "high", "2026-01-06", "2026-06-30", 0.40,
             "List new tech candidates", "Feasibility analysis & prioritization", "PoC plan & execution",
             "PoC result report", "Mass production applicability assessment"],
            ["Economic New Tech Development", "Namyang collaboration tech items", "🔗",
             "[Output: Joint dev report] Progress on joint tech development with Namyang R&D",
             "", 2, "items", "high", "2026-02-01", "2026-06-30", 0.35,
             "Confirm collab items & define R&R", "1st interim review (tech review)", "Prototype/simulation validation",
             "Final report & mass production transfer decision", ""],
            ["Economic New Tech Development", "Motor & parts localization", "🏭",
             "[Output: Localization status table] Motor localization & key parts local sourcing",
             "", 2, "items", "medium", "2026-03-01", "2026-06-30", 0.25,
             "Select parts for localization", "Local supplier sourcing & tech evaluation", "Prototype manufacturing & QC",
             "Cost comparison analysis report", ""],

            # ── Area 3: Dev Robustness ──
            ["Development Robustness", "Crash technology exchange meetings", "💥",
             "[Output: Meeting report] Regular crash tech exchange sessions & know-how sharing",
             "", 4, "sessions", "high", "2026-01-15", "2026-06-30", 0.20,
             "Establish annual schedule", "1st exchange meeting & summary", "Derive improvements & action plan",
             "2nd exchange meeting", "Annual performance summary"],
            ["Development Robustness", "Strengthen DPA checklist", "✅",
             "[Output: Enhanced DPA checklist] Improve existing DPA checklist & apply to models",
             "", 1, "report", "high", "2026-02-01", "2026-04-30", 0.20,
             "Analyze current DPA issues", "Benchmark best practices", "Derive improvements & revise checklist",
             "Pilot model application & validation", "Full model rollout"],
            ["Development Robustness", "EO-based preventive design before drawing release", "🔍",
             "[Output: Preventive design guide] Build EO case DB & pre-release prevention process",
             "", 1, "report", "high", "2026-01-15", "2026-05-31", 0.20,
             "Collect & classify EO cases (2 years)", "Draft prevention design guide", "Cross-team review & feedback",
             "Finalize pre-release check process", "Distribute guide & training"],
            ["Development Robustness", "Strengthen ESIR schedule management", "📅",
             "[Output: Process improvement plan] Improve ESIR schedule compliance rate",
             "", 95, "%", "medium", "2026-02-01", "2026-06-30", 0.20,
             "Analyze current ESIR compliance rate", "Identify delay causes & improvement plan", "Build management dashboard",
             "Monthly monitoring operation", ""],
            ["Development Robustness", "Improve drawing completeness", "📐",
             "[Output: Drawing quality report] Improve drawing completeness metrics",
             "", 90, "%", "high", "2026-01-06", "2026-06-30", 0.20,
             "Define completeness KPI metrics", "Measure current level & gap analysis", "Drive team-level improvements",
             "Monthly completeness monitoring", "Verify target achievement & report"],

            # ── Area 4: Parts Commonization ──
            ["Parts Commonization & Pre SR Process", "Create Level-1 master parts list", "📊",
             "[Output: Master list Excel] Level-1 parts master list based on base vehicle",
             "", 1, "report", "critical", "2026-01-15", "2026-03-31", 0.30,
             "Select base vehicle & obtain BOM", "Organize L1 parts list", "Map specs/drawing numbers per part",
             "Internal validation & finalization", ""],
            ["Parts Commonization & Pre SR Process", "Purchase price & new vs C/O comparison", "💹",
             "[Output: Comparison table] Compare new development cost vs C/O cost & calculate savings",
             "", 1, "report", "high", "2026-03-01", "2026-05-15", 0.30,
             "Collect purchase prices from master list", "Estimate new development cost", "Calculate C/O application cost",
             "Create savings comparison table", "Executive report"],
            ["Parts Commonization & Pre SR Process", "Build design counter-proposal process", "🔄",
             "[Output: Process guide] Establish pre-design-selection counter-proposal process for all models",
             "", 1, "report", "high", "2026-02-01", "2026-05-31", 0.25,
             "Analyze KS2 counter-proposal cases", "Draft standard process", "Pilot model application & feedback",
             "Finalize process & distribute guide", "All-model kickoff"],
            ["Parts Commonization & Pre SR Process", "Commonization performance monitoring", "📈",
             "[Output: Monthly dashboard] Track commonization progress & savings monthly",
             "", 6, "reports", "medium", "2026-03-01", "2026-06-30", 0.15,
             "Design monitoring KPIs & dashboard", "Build monthly data collection system", "Publish 1st report",
             "3-month cumulative performance analysis", ""],

            # ── Area 5: CCW ──
            ["CCW Activities (Cost Reduction)", "Achieve production vehicle CCW targets", "💰",
             "[Output: CCW performance report] Drive cost reduction activities & track performance",
             "", 100, "%", "high", "2026-01-06", "2026-06-30", 0.60,
             "Confirm CCW targets per model", "Q1 cost reduction activities & tally", "Midpoint review & supplemental plan",
             "Q2 cost reduction activities", "H1 comprehensive performance report"],
            ["CCW Activities (Cost Reduction)", "Cost reduction idea generation & sharing", "💡",
             "[Output: Idea DB] Collect cost reduction ideas & share best practices",
             "", 20, "ideas", "medium", "2026-02-01", "2026-06-30", 0.40,
             "Monthly idea collection by team", "Build idea database", "Select & share best practices",
             "Create execution plan for applicable items", ""],

            # ── People: Talent Development ──
            ["Talent Development", "SW talent development program", "💻",
             "[Output: Training status table] Operate SW competency training & manage completion rate",
             "", 90, "%", "high", "2026-01-15", "2026-06-30", 0.12,
             "Confirm trainees & courses", "Establish training schedule", "Conduct training (quarterly)",
             "Monitor completion rate", ""],
            ["Talent Development", "PCL system operation", "🔧",
             "[Output: PCL operation report] Stable PCL system operation & usage improvement",
             "", 1, "report", "medium", "2026-01-06", "2026-06-30", 0.10,
             "Review current PCL status", "Identify improvements needed", "System update & training",
             "Usage monitoring", ""],
            ["Talent Development", "E2E (Error to Excellence) activities", "🎯",
             "[Output: E2E case book] Convert error cases to excellence cases & share",
             "", 10, "cases", "medium", "2026-02-01", "2026-06-30", 0.10,
             "Collect error cases by team", "Root cause analysis & improvement plan", "Apply improvements & verify",
             "Compile excellence case book", ""],
            ["Talent Development", "Work manual creation", "📖",
             "[Output: Work manuals] Create standard process manuals for key tasks",
             "", 5, "manuals", "medium", "2026-02-01", "2026-06-30", 0.08,
             "Select target processes", "Assign owners & draft creation", "Internal review & revision",
             "Publish & distribute final manuals", ""],
            ["Talent Development", "Culture survey & improvement", "🏢",
             "[Output: Survey results & action plan] Conduct culture survey & drive improvements",
             "", 1, "report", "medium", "2026-03-01", "2026-06-30", 0.10,
             "Design survey questions", "Conduct survey & collect data", "Analyze results & identify improvements",
             "Execute improvement activities", "Plan follow-up survey"],
            ["Talent Development", "IIT professor invited seminar", "🎓",
             "[Output: Seminar report] Host technology seminars with IIT professors",
             "", 2, "sessions", "medium", "2026-02-15", "2026-06-30", 0.10,
             "Select topics & invite professors", "Confirm schedule & recruit participants", "Host seminar",
             "Summarize results & discuss collaboration", ""],
            ["Talent Development", "Supplier & system tech shows", "🏗️",
             "[Output: Tech show report] Host supplier/system-level tech exhibitions",
             "", 2, "events", "medium", "2026-03-01", "2026-06-30", 0.10,
             "Select topics & participating suppliers", "Confirm schedule & booth setup", "Host tech show",
             "Select outstanding tech & review application", ""],
            ["Talent Development", "General & specialized training", "📚",
             "[Output: Training record] Operate general competency + team-specific training",
             "", 10, "sessions", "medium", "2026-01-15", "2026-06-30", 0.10,
             "Establish training courses & schedule", "Q1 training delivery", "Midpoint completion check",
             "Q2 training delivery", "H1 training performance report"],
            ["Talent Development", "Design robustness training", "🛡️",
             "[Output: Training completion status] Operate design quality robustness training",
             "", 4, "sessions", "high", "2026-02-01", "2026-06-30", 0.10,
             "Design training curriculum", "Arrange instructors & confirm schedule", "Deliver training (quarterly)",
             "Evaluate training effectiveness", ""],
            ["Talent Development", "Commercial center (B04) support", "🚛",
             "[Output: Support performance report] Technical support for B04 vehicle development",
             "", 1, "report", "medium", "2026-01-06", "2026-06-30", 0.10,
             "Define support scope & R&R", "Execute technical support activities", "Midpoint review & issue sharing",
             "Complete support & summarize results", ""],
        ]

    # Write action rows
    row = 4
    current_parent = ""
    for action in ACTIONS:
        row += 1
        parent = action[0]

        for col_idx, value in enumerate(action):
            col_letter = COLUMNS[col_idx][0]
            cell = ws2[f"{col_letter}{row}"]
            cell.value = value
            cell.font = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")
            cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
            cell.border = THIN_BORDER
            if col_idx >= 11:  # milestone columns
                cell.fill = MILESTONE_FILL

        # Color-code by Business vs People
        is_people = parent in ("인재육성", "Talent Development")
        for col_letter, _, _ in COLUMNS[:11]:
            cell = ws2[f"{col_letter}{row}"]
            if is_people:
                cell.fill = PPL_FILL

        # Section separator
        if parent != current_parent:
            current_parent = parent
            ws2[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=10, color="1E3A5F")

        ws2.row_dimensions[row].height = 32

    # Empty rows for team leaders to add more
    for extra_row in range(row + 1, row + 11):
        for col_idx, (col_letter, _, _) in enumerate(COLUMNS):
            cell = ws2[f"{col_letter}{extra_row}"]
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
            if col_idx >= 11:
                cell.fill = MILESTONE_FILL
        ws2.row_dimensions[extra_row].height = 28

    # Validations
    last_data_row = row + 10
    pv = DataValidation(type="list", formula1='"low,medium,high,critical"')
    pv.error = "Select: low, medium, high, critical"
    ws2.add_data_validation(pv)
    pv.add(f"H4:H{last_data_row}")

    ws2.freeze_panes = "A4"

    # ═══════════════════════════════════════
    # Sheet 3: 작성 가이드 / Guide
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("작성가이드" if is_ko else "Guide")
    ws3.sheet_properties.tabColor = "F59E0B"

    if is_ko:
        guide = [
            ("작성 가이드", True, 14),
            ("", False, 10),
            ("1. 이 양식은 참고용 예시가 사전 입력되어 있습니다 (회색 이탤릭).", False, 10),
            ("   팀 상황에 맞게 수정/삭제/추가해 주세요.", False, 10),
            ("", False, 10),
            ("2. 필수 요소 (모든 액션에 반드시 포함)", True, 11),
            ("   a) 담당자 — 책임자 1인", False, 10),
            ("   b) 마감일 — YYYY-MM-DD", False, 10),
            ("   c) 설명/아웃풋 — [아웃풋: 유형] 설명 형태로 작성", False, 10),
            ("   d) 마일스톤 3~5개 — 중간 점검 가능한 체크포인트", False, 10),
            ("", False, 10),
            ("3. 가중치 규칙", True, 11),
            ("   같은 상위 KPI 아래 액션들의 가중치 합 = 1.0", False, 10),
            ("", False, 10),
            ("4. 핵심 메시지", True, 11),
            ("   \"어떤 업무든 과정을 쪼개면 성취가 된다.\"", False, 10),
            ("   팀원들이 자기가 세운 목표를 반드시 성취해내는 경험을 하도록 하자.", False, 10),
            ("   그것이 조직문화의 힘이다.", False, 10),
            ("", False, 10),
            ("5. 제출: 작성 후 회신", True, 11),
        ]
    else:
        guide = [
            ("Writing Guide", True, 14),
            ("", False, 10),
            ("1. This template comes pre-filled with reference examples (gray italic).", False, 10),
            ("   Modify, delete, or add actions to fit your team's situation.", False, 10),
            ("", False, 10),
            ("2. Required Elements (every action must include)", True, 11),
            ("   a) Owner — One responsible person", False, 10),
            ("   b) Due Date — YYYY-MM-DD format", False, 10),
            ("   c) Description/Output — Write as [Output: type] description", False, 10),
            ("   d) 3-5 Milestones — Verifiable interim checkpoints", False, 10),
            ("", False, 10),
            ("3. Weight Rules", True, 11),
            ("   Weights of actions under the same parent KPI must sum to 1.0", False, 10),
            ("", False, 10),
            ("4. Core Philosophy", True, 11),
            ("   \"Break every task into achievable milestones.\"", False, 10),
            ("   Ensure every team member experiences achieving their own goals.", False, 10),
            ("   That is the power of organizational culture.", False, 10),
            ("", False, 10),
            ("5. Submission: Fill in and reply with this file", True, 11),
        ]

    for i, (text, bold, size) in enumerate(guide, 1):
        cell = ws3[f"A{i}"]
        cell.value = text
        cell.font = Font(name=FONT_NAME, bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 80

    return wb


# ── Generate both versions ──
for lang in ("ko", "en"):
    wb = make_workbook(lang)
    if lang == "ko":
        path_obs = r"C:\obsidian\0_Inbox\HMIE_KPI_액션플랜_2026.xlsx"
        path_pub = r"C:\Users\USER\kpi-cascade\public\HMIE_KPI_액션플랜_2026.xlsx"
    else:
        path_obs = r"C:\obsidian\0_Inbox\HMIE_KPI_Action_Plan_2026.xlsx"
        path_pub = r"C:\Users\USER\kpi-cascade\public\HMIE_KPI_Action_Plan_2026.xlsx"
    wb.save(path_obs)
    wb.save(path_pub)
    print(f"[{lang.upper()}] {path_obs}")
    print(f"[{lang.upper()}] {path_pub}")
