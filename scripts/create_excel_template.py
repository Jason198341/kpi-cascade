"""Generate KPI Action Plan Excel template for team leaders."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Styles ──
HEADER_FILL = PatternFill(start_color="1A1D25", end_color="1A1D25", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
SUBHEADER_FONT = Font(name="맑은 고딕", bold=True, color="E2E8F0", size=10)
DATA_FONT = Font(name="맑은 고딕", size=10)
GUIDE_FONT = Font(name="맑은 고딕", size=9, color="6B7280")
MILESTONE_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════
# Sheet 1: 액션 플랜 입력
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "액션플랜"

COLUMNS = [
    ("A", "상위 KPI\n(팀 KPI 제목)", 25),
    ("B", "액션 제목", 30),
    ("C", "이모지", 6),
    ("D", "설명", 35),
    ("E", "담당자", 12),
    ("F", "목표값", 8),
    ("G", "단위", 6),
    ("H", "우선순위", 10),
    ("I", "시작일\n(YYYY-MM-DD)", 14),
    ("J", "마감일\n(YYYY-MM-DD)", 14),
    ("K", "아웃풋 형태", 18),
    ("L", "가중치\n(0~1)", 8),
    ("M", "마일스톤 1", 22),
    ("N", "마일스톤 2", 22),
    ("O", "마일스톤 3", 22),
    ("P", "마일스톤 4", 22),
    ("Q", "마일스톤 5", 22),
]

# Title row
ws.merge_cells("A1:Q1")
title_cell = ws["A1"]
title_cell.value = "KPI 캐스케이드 — 액션 플랜 양식"
title_cell.font = Font(name="맑은 고딕", bold=True, size=14, color="8B5CF6")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Info row
ws.merge_cells("A2:Q2")
info_cell = ws["A2"]
info_cell.value = "팀명:                    작성자:                    작성일:          "
info_cell.font = Font(name="맑은 고딕", size=10, color="6B7280")
info_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 24

# Header row (row 3)
for col_letter, label, width in COLUMNS:
    cell = ws[f"{col_letter}3"]
    cell.value = label
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 40

# Subheader (row 4) — field descriptions
DESCRIPTIONS = [
    "귀 팀에 할당된\n팀 KPI 제목",
    "구체적 액션명\n(동사형 권장)",
    "예:\n📞🔍📝",
    "액션 설명\n(1~2문장)",
    "책임자\n이름",
    "정량\n목표",
    "건, %,\n개 등",
    "low/med\nhigh/crit",
    "시작\n날짜",
    "완료\n목표일",
    "보고서/시스템\n/교육/분석 등",
    "팀KPI 내\n비중",
    "1단계\n체크포인트",
    "2단계\n체크포인트",
    "3단계\n체크포인트",
    "4단계\n(선택)",
    "5단계\n(선택)",
]

for i, (col_letter, _, _) in enumerate(COLUMNS):
    cell = ws[f"{col_letter}4"]
    cell.value = DESCRIPTIONS[i]
    cell.font = GUIDE_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

ws.row_dimensions[4].height = 36

# Example rows (5-6)
EXAMPLE_1 = [
    "영업팀 매출 성장", "Q2 대형 고객 5건 미팅", "📞",
    "전략 고객 대상 직접 미팅을 통한 매출 파이프라인 확대",
    "김영수", 5, "건", "high",
    "2026-02-01", "2026-06-30", "미팅 완료 보고서",
    0.40,
    "타겟 고객 리스트 확정", "1차 컨택 완료", "미팅 일정 확정",
    "제안서 발송", "미팅 완료 및 결과 보고",
]

EXAMPLE_2 = [
    "마케팅 리드 생성", "경쟁사 벤치마킹 보고서 작성", "🔍",
    "주요 경쟁사 3사 대비 당사 강·약점 분석 및 전략 도출",
    "이수진", 1, "건", "medium",
    "2026-03-01", "2026-04-15", "분석 보고서 (PPT)",
    0.30,
    "벤치마킹 대상 및 범위 확정", "1차 데이터 수집 완료",
    "분석 프레임워크 수립 및 초안", "내부 리뷰 및 피드백 반영",
    "최종 보고서 완성 및 공유",
]

for row_idx, example in [(5, EXAMPLE_1), (6, EXAMPLE_2)]:
    for col_idx, value in enumerate(example):
        col_letter = COLUMNS[col_idx][0]
        cell = ws[f"{col_letter}{row_idx}"]
        cell.value = value
        cell.font = Font(name="맑은 고딕", size=10, italic=True, color="9CA3AF")
        cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
        cell.border = THIN_BORDER
        if col_idx >= 12:  # milestone columns
            cell.fill = MILESTONE_FILL
    ws.row_dimensions[row_idx].height = 32

# Empty data rows (7-26) with borders and milestone highlighting
for row_idx in range(7, 27):
    for col_idx, (col_letter, _, _) in enumerate(COLUMNS):
        cell = ws[f"{col_letter}{row_idx}"]
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = LEFT_WRAP if col_idx in (1, 3) else CENTER
        if col_idx >= 12:
            cell.fill = MILESTONE_FILL
    ws.row_dimensions[row_idx].height = 28

# Data validations
priority_dv = DataValidation(type="list", formula1='"low,medium,high,critical"')
priority_dv.error = "low, medium, high, critical 중 선택"
priority_dv.errorTitle = "우선순위 오류"
ws.add_data_validation(priority_dv)
priority_dv.add(f"H5:H26")

unit_dv = DataValidation(type="list", formula1='"건,%,개,점,원,명,회"')
ws.add_data_validation(unit_dv)
unit_dv.add(f"G5:G26")

# Freeze panes
ws.freeze_panes = "A5"

# ═══════════════════════════════════════════════
# Sheet 2: 작성 가이드
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("작성가이드")

guide_lines = [
    ("KPI 캐스케이드 — 액션 플랜 작성 가이드", True, 14),
    ("", False, 10),
    ("1. 구조", True, 12),
    ("   전략 목표 (depth-0) → 팀 KPI (depth-1) → 액션 플랜 (depth-2)", False, 10),
    ("   팀장님은 '액션플랜' 시트에 depth-2 항목을 작성합니다.", False, 10),
    ("", False, 10),
    ("2. 필수 요소 (모든 액션에 반드시 포함)", True, 12),
    ("   a) 담당자 — 책임자 1인 명기", False, 10),
    ("   b) 마감기한 — YYYY-MM-DD 형식", False, 10),
    ("   c) 아웃풋 형태 — 결과물 유형 (보고서, 시스템, 교육 완료 등)", False, 10),
    ("   d) 목표 수치 — 정량적 목표와 단위", False, 10),
    ("   e) 마일스톤 3~5개 — 중간 점검 가능한 체크포인트", False, 10),
    ("", False, 10),
    ("3. 마일스톤 작성 원칙", True, 12),
    ("   - 각 단계는 '완료/미완료'로 판별 가능해야 합니다", False, 10),
    ("   - 리서치/분석 업무도 과정을 3~5단계로 쪼개주세요", False, 10),
    ("   - '~완료', '~확정', '~제출' 등 동사형으로 작성", False, 10),
    ("", False, 10),
    ("4. 가중치(weight) 안내", True, 12),
    ("   - 같은 팀 KPI 아래의 액션들의 가중치 합 = 1.0", False, 10),
    ("   - 예: 액션A(0.4) + 액션B(0.35) + 액션C(0.25) = 1.0", False, 10),
    ("   - 중요도가 높은 액션에 더 큰 가중치 부여", False, 10),
    ("", False, 10),
    ("5. 분해 예시", True, 12),
    ("   [리서치형] '시장조사 보고서 작성'", False, 10),
    ("     M1: 조사 범위 및 방법론 확정", False, 10),
    ("     M2: 1차 데이터 수집 완료", False, 10),
    ("     M3: 분석 및 초안 작성", False, 10),
    ("     M4: 내부 리뷰 피드백 반영", False, 10),
    ("     M5: 최종본 완성 및 공유", False, 10),
    ("", False, 10),
    ("   [구축형] '시스템 자동화 구축'", False, 10),
    ("     M1: 현행 프로세스 분석", False, 10),
    ("     M2: 도구 선정 및 설계", False, 10),
    ("     M3: 프로토타입 개발 및 테스트", False, 10),
    ("     M4: 팀 교육 및 매뉴얼 작성", False, 10),
    ("     M5: 프로덕션 적용", False, 10),
    ("", False, 10),
    ("6. 제출", True, 12),
    ("   - '액션플랜' 시트 작성 후 회신", False, 10),
    ("   - 예시 행(회색 이탤릭)은 참고용이며 삭제해도 됩니다", False, 10),
    ("   - 마일스톤 4, 5는 선택사항 (3개 이상 필수)", False, 10),
]

for i, (text, bold, size) in enumerate(guide_lines, 1):
    cell = ws2[f"A{i}"]
    cell.value = text
    cell.font = Font(name="맑은 고딕", bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True)

ws2.column_dimensions["A"].width = 80

# Save
output_path = r"C:\obsidian\0_Inbox\KPI_액션플랜_양식.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
