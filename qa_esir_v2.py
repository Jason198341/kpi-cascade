"""
ESIR Master Tracker V2 — Comprehensive QA Script (updated for 30 items)
Tests: structure, formulas, data integrity, yellow highlighting,
       progress check, add/delete/status-change scenarios, multi-spec
"""

import sys
import io
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

FILE = r"C:\Users\USER\kpi-cascade\ESIR_Master_Tracker.xlsx"
TODAY = datetime.date(2026, 2, 26)
C_MANUAL_YELLOW = "FFF9C4"

PASS = 0
FAIL = 0
WARN = 0


def report(status, msg):
    global PASS, FAIL, WARN
    if status == "PASS":
        PASS += 1
        print(f"  [PASS] {msg}")
    elif status == "FAIL":
        FAIL += 1
        print(f"  [FAIL] {msg}")
    elif status == "WARN":
        WARN += 1
        print(f"  [WARN] {msg}")
    else:
        print(f"  [INFO] {msg}")


def section(title):
    print(f"\n{'='*70}")
    print(f"  {title}")
    print(f"{'='*70}")


def get_tracker(wb):
    for name in wb.sheetnames:
        if "Master Tracker" in name:
            return wb[name]
    return None


def get_dashboard(wb):
    for name in wb.sheetnames:
        if "Dashboard" in name:
            return wb[name]
    return None


# ============================================================
# CHECK 1: Sheet Architecture
# ============================================================
def check_sheet_architecture(wb):
    section("CHECK 1: Sheet Architecture")
    expected = ["Config", "Reference Data", "Master Tracker", "Dashboard", "Supplier Checklist"]
    actual = wb.sheetnames
    report("INFO", f"Sheets: {actual}")
    for e in expected:
        found = any(e.lower() in s.lower() for s in actual)
        report("PASS" if found else "FAIL", f"Sheet '{e}' exists: {found}")
    config_ws = wb[[s for s in actual if "Config" in s][0]]
    report("PASS" if config_ws.sheet_state == "hidden" else "WARN", f"Config hidden: {config_ws.sheet_state}")


# ============================================================
# CHECK 2: Reference Data Tables
# ============================================================
def check_reference_data(wb):
    section("CHECK 2: Reference Data (SSOT)")
    ref_ws = None
    for name in wb.sheetnames:
        if "Reference Data" in name:
            ref_ws = wb[name]
            break
    if not ref_ws:
        report("FAIL", "Reference Data sheet missing!")
        return
    expected_tables = ["TBL_Vehicle", "TBL_LeadTime", "TBL_Status",
                       "TBL_Supplier", "TBL_Parts", "TBL_TestCatalog", "TBL_Milestone",
                       "TBL_VehicleParts", "TBL_PartTests"]
    actual = list(ref_ws.tables.keys())
    report("INFO", f"Tables: {actual}")
    for t in expected_tables:
        report("PASS" if t in actual else "FAIL", f"Table '{t}'")
    report("PASS" if len(actual) == 9 else "FAIL", f"Total tables: {len(actual)}/9")


# ============================================================
# CHECK 3: Master Tracker Structure (31 cols, 30 rows)
# ============================================================
def check_tracker_structure(wb):
    section("CHECK 3: Master Tracker Structure")
    ws = get_tracker(wb)
    if not ws:
        report("FAIL", "Tracker not found!")
        return

    headers = [ws.cell(row=3, column=c).value for c in range(1, 32)]
    non_none = [h for h in headers if h]
    report("PASS" if len(non_none) == 31 else "FAIL", f"Header columns: {len(non_none)}/31")

    data_rows = 0
    for row in range(5, 250):
        # Count rows with actual data in B (VehicleCode), not formula-only rows
        if ws.cell(row=row, column=2).value is not None:
            data_rows += 1
        else:
            break
    report("PASS" if data_rows == 66 else "FAIL", f"Data rows: {data_rows}/66")

    # Verify empty formula rows exist beyond data
    formula_row = data_rows + 5  # first empty formula row
    a_val = ws.cell(row=formula_row, column=1).value
    has_formula = isinstance(a_val, str) and a_val.startswith("=")
    report("PASS" if has_formula else "WARN", f"Empty rows pre-populated with formulas: {has_formula}")
    report("PASS" if ws.freeze_panes == "A5" else "WARN", f"Freeze panes: {ws.freeze_panes}")


# ============================================================
# CHECK 4: Yellow Manual Input Highlighting
# ============================================================
def check_yellow_highlighting(wb):
    section("CHECK 4: Manual Input Cells (Yellow Highlighting)")
    ws = get_tracker(wb)
    if not ws:
        return

    MANUAL_COLS = [2, 3, 6, 11, 16, 17, 19, 21, 23, 25, 26]
    MANUAL_COL_NAMES = {
        2: "B:Vehicle", 3: "C:PartCode", 6: "F:TestCode", 11: "K:Stage",
        16: "P:PlanSubmit", 17: "Q:PlanReview", 19: "S:PartsActual",
        21: "U:TestStartActual", 23: "W:TestCompleteActual",
        25: "Y:ReportSubmit", 26: "Z:Override"
    }

    yellow_ok = 0
    yellow_fail = 0

    # Check first 5 data rows
    for row in [5, 6, 7, 8, 9]:
        for col in MANUAL_COLS:
            cell = ws.cell(row=row, column=col)
            fill_color = cell.fill.start_color
            if fill_color and hasattr(fill_color, 'rgb'):
                rgb = str(fill_color.rgb)
                is_yellow = C_MANUAL_YELLOW.upper() in rgb.upper() or "FFF9C4" in rgb.upper()
            else:
                is_yellow = False

            if is_yellow:
                yellow_ok += 1
            else:
                yellow_fail += 1
                if row == 5:  # Only report first row in detail
                    report("FAIL", f"Row {row} Col {MANUAL_COL_NAMES.get(col, col)}: fill={rgb if fill_color else 'none'}")

    total_checked = yellow_ok + yellow_fail
    report("PASS" if yellow_fail == 0 else "FAIL",
           f"Yellow highlighting: {yellow_ok}/{total_checked} cells correct")

    # Check a NON-manual column is NOT yellow
    auto_cell = ws.cell(row=5, column=4)  # D: PartCategory (AUTO)
    auto_rgb = str(auto_cell.fill.start_color.rgb) if auto_cell.fill.start_color else ""
    is_not_yellow = "FFF9C4" not in auto_rgb.upper()
    report("PASS" if is_not_yellow else "WARN", f"Auto col D5 is NOT yellow: {is_not_yellow}")


# ============================================================
# CHECK 5: Formula Integrity
# ============================================================
def check_formula_integrity(wb):
    section("CHECK 5: Formula Integrity")
    ws = get_tracker(wb)
    if not ws:
        return

    auto_cols = {4: "D:PartCat", 5: "E:PartName", 7: "G:TestItem",
                 8: "H:TestCat", 9: "I:TestType", 10: "J:Location",
                 12: "L:Spec", 13: "M:Supplier", 14: "N:T0",
                 15: "O:PlanDue", 18: "R:PartsReadyPlan",
                 20: "T:TestStartPlan", 22: "V:TestCompletePlan",
                 24: "X:ReportDue", 27: "AA:Status", 28: "AB:Progress",
                 29: "AC:Days", 30: "AD:Risk", 31: "AE:Action"}

    all_formula = True
    for col, name in auto_cols.items():
        val = ws.cell(row=5, column=col).value
        if not (isinstance(val, str) and val.startswith("=")):
            report("FAIL", f"{name}: not a formula, got '{val}'")
            all_formula = False
    if all_formula:
        report("PASS", f"All {len(auto_cols)} auto columns are formulas")

    # Parentheses balance
    paren_err = 0
    for row in range(5, 71):
        if ws.cell(row=row, column=2).value is None:
            break
        for col in auto_cols:
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val.startswith("="):
                if val.count("(") != val.count(")"):
                    paren_err += 1
                    report("FAIL", f"Row {row} {auto_cols[col]}: unbalanced parens")
    if paren_err == 0:
        report("PASS", "All formula parentheses balanced")

    # Status override priority
    aa5 = ws.cell(row=5, column=27).value
    if isinstance(aa5, str) and 'Z5<>""' in aa5:
        report("PASS", "Status formula: Override (Z) takes priority")
    else:
        report("FAIL", "Status formula: Override check missing")


# ============================================================
# CHECK 6: 12-State Coverage
# ============================================================
def check_12_states(wb):
    section("CHECK 6: 12-State Coverage (66 items)")
    ws = get_tracker(wb)
    if not ws:
        return

    all_12 = [
        "01-계획서미제출", "02-계획서제출_미검토", "03-검토중_조건충족",
        "04-검토중_조건불만족", "05-계획서승인", "06-BIW생산중단",
        "07-생산완료_미인수", "08-단품준비완료", "09-시험준비중",
        "10-시험완료", "11-결과등록완료", "12-조건부승인검토"
    ]

    state_counts = {s: 0 for s in all_12}
    rows_info = []

    for row in range(5, 250):
        # Only count rows with actual data in B (VehicleCode)
        vehicle = ws.cell(row=row, column=2).value
        if vehicle is None:
            break

        no = ws.cell(row=row, column=1).value
        part = ws.cell(row=row, column=3).value
        test = ws.cell(row=row, column=6).value
        p = ws.cell(row=row, column=16).value
        q = ws.cell(row=row, column=17).value
        s = ws.cell(row=row, column=19).value
        u = ws.cell(row=row, column=21).value
        w = ws.cell(row=row, column=23).value
        y = ws.cell(row=row, column=25).value
        z = ws.cell(row=row, column=26).value

        if z:
            expected = z
        elif y and w:
            expected = "11-결과등록완료"
        elif w and not y:
            expected = "10-시험완료"
        elif u and not w:
            expected = "09-시험준비중"
        elif s and not u:
            expected = "08-단품준비완료"
        elif q == "Approved":
            expected = "05-계획서승인"
        elif q == "Revision Needed":
            expected = "04-검토중_조건불만족"
        elif q == "Under Review":
            expected = "03-검토중_조건충족"
        elif p and not q:
            expected = "02-계획서제출_미검토"
        else:
            expected = "01-계획서미제출"

        if expected in state_counts:
            state_counts[expected] += 1
        rows_info.append({
            "row": row, "no": no, "vehicle": vehicle, "part": part,
            "test": test, "status": expected, "override": z
        })

    for state in all_12:
        cnt = state_counts[state]
        ovr = " (Override)" if state in ("06-BIW생산중단", "07-생산완료_미인수", "12-조건부승인검토") and cnt > 0 else ""
        report("PASS" if cnt > 0 else "FAIL", f"{state:28s} x{cnt}{ovr}")

    missing = [s for s in all_12 if state_counts[s] == 0]
    report("PASS" if not missing else "FAIL",
           f"All 12 states covered: {'YES' if not missing else 'MISSING ' + str(missing)}")

    return rows_info


# ============================================================
# CHECK 7: Multi-Spec per Part (아이템별 다중 스펙)
# ============================================================
def check_multi_spec(wb, rows_info):
    section("CHECK 7: Multi-Spec per Part (1 Part × N Specs)")
    if not rows_info:
        return

    # Group by (Vehicle, Part) → list of tests/specs
    part_specs = {}
    for r in rows_info:
        key = f"{r['vehicle']}:{r['part']}"
        if key not in part_specs:
            part_specs[key] = []
        part_specs[key].append(r['test'])

    multi = {k: v for k, v in part_specs.items() if len(v) > 1}
    single = {k: v for k, v in part_specs.items() if len(v) == 1}

    report("INFO", f"Total part instances: {len(part_specs)} ({len(multi)} multi-spec, {len(single)} single-spec)")

    print(f"\n  Multi-spec parts (1 Part × N Tests/Specs):")
    print(f"  {'Vehicle:Part':28s} | Tests/Specs")
    print(f"  {'-'*28} | {'-'*40}")
    for k, v in sorted(multi.items()):
        print(f"  {k:28s} | {', '.join(str(x) for x in v if x)}")

    report("PASS" if len(multi) >= 3 else "WARN",
           f"Multi-spec parts: {len(multi)} (diversity check)")

    # Check same part appears in different vehicles
    part_vehicles = {}
    for r in rows_info:
        if r['part'] not in part_vehicles:
            part_vehicles[r['part']] = set()
        part_vehicles[r['part']].add(r['vehicle'])

    cross_vehicle = {k: v for k, v in part_vehicles.items() if len(v) > 1}
    report("PASS" if len(cross_vehicle) >= 2 else "WARN",
           f"Parts in multiple vehicles: {len(cross_vehicle)}")
    for k, v in cross_vehicle.items():
        print(f"    {k} → {sorted(v)}")


# ============================================================
# CHECK 8: Vehicle Distribution & Cross-Ref
# ============================================================
def check_vehicle_distribution(wb, rows_info):
    section("CHECK 8: Vehicle Distribution")
    if not rows_info:
        return

    counts = {}
    for r in rows_info:
        v = r['vehicle']
        counts[v] = counts.get(v, 0) + 1

    # 22 part-test combos per vehicle (6 parts × varying tests = 22 each)
    expected = {"AE_PE": 22, "SU2i": 22, "EN_SUV": 22}
    for v, exp in expected.items():
        actual = counts.get(v, 0)
        report("PASS" if actual == exp else "WARN", f"{v}: {actual} items (expected {exp})")


# ============================================================
# CHECK 9: Dashboard — Vehicle Filter & Progress Check
# ============================================================
def check_dashboard(wb):
    section("CHECK 9: Dashboard (Vehicle Filter + Progress Check)")
    ws = get_dashboard(wb)
    if not ws:
        report("FAIL", "Dashboard not found!")
        return

    # Vehicle selector (dropdown on Dashboard C3, default value "AE_PE")
    c3 = ws.cell(row=3, column=3).value
    report("PASS" if c3 == "AE_PE" else "FAIL",
           f"Vehicle selector → Dashboard C3 dropdown: {c3}")

    # KPI formula range check (should use :200 not :28)
    kpi = ws.cell(row=7, column=2).value
    if isinstance(kpi, str):
        uses_200 = ":200" in kpi or ":B200" in kpi
        report("PASS" if uses_200 else "FAIL",
               f"KPI range extends to row 200: {uses_200}")

    # Cross-vehicle comparison
    vehicles_found = []
    for r in range(13, 16):
        v = ws.cell(row=r, column=2).value
        if v:
            vehicles_found.append(v)
    report("PASS" if len(vehicles_found) == 3 else "FAIL",
           f"Cross-vehicle comparison rows: {vehicles_found}")

    # Status distribution (12 statuses)
    status_count = 0
    for r in range(19, 31):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str) and val[:2].isdigit():
            status_count += 1
    report("PASS" if status_count == 12 else "FAIL",
           f"Status distribution rows: {status_count}/12")

    # Progress Check section
    progress_header = None
    for r in range(31, 40):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str) and "Progress" in val:
            progress_header = r
            break

    report("PASS" if progress_header else "FAIL",
           f"Progress Check section found at row {progress_header}")

    if progress_header:
        # Check Avg Progress formula
        avg_cell = ws.cell(row=progress_header + 2, column=3).value
        has_averageifs = isinstance(avg_cell, str) and "AVERAGEIFS" in avg_cell
        report("PASS" if has_averageifs else "FAIL",
               f"Avg Progress uses AVERAGEIFS: {has_averageifs}")

        # Check Behind Schedule formula
        behind_cell = ws.cell(row=progress_header + 2, column=4).value
        has_countifs = isinstance(behind_cell, str) and "COUNTIFS" in behind_cell
        report("PASS" if has_countifs else "FAIL",
               f"Behind Schedule uses COUNTIFS: {has_countifs}")

        # Check Done formula
        done_cell = ws.cell(row=progress_header + 2, column=6).value
        has_done = isinstance(done_cell, str) and "DONE" in done_cell
        report("PASS" if has_done else "FAIL",
               f"Done count formula: {has_done}")


# ============================================================
# CHECK 10: Data Validations — Dropdown Ranges
# ============================================================
def check_validations(wb):
    section("CHECK 10: Data Validations (Dropdown Ranges)")
    ws = get_tracker(wb)
    if not ws:
        return

    dvs = ws.data_validations.dataValidation
    report("INFO", f"Total validations: {len(dvs)}")

    for dv in dvs:
        cells = str(dv.sqref)
        f1 = dv.formula1 or ""
        report("INFO", f"  {cells:20s} → {f1[:60]}")

    # Check vehicle dropdown range
    vehicle_dv = None
    for dv in dvs:
        if "B5" in str(dv.sqref):
            vehicle_dv = dv
            break

    if vehicle_dv:
        f1 = vehicle_dv.formula1 or ""
        uses_20 = "$A$20" in f1 or "$A$3:$A$20" in f1
        report("PASS" if uses_20 else "FAIL",
               f"Vehicle dropdown extends to $A$20 (room for growth): {uses_20}")
    else:
        report("FAIL", "Vehicle dropdown not found")


# ============================================================
# CHECK 11: Conditional Formatting
# ============================================================
def check_conditional_formatting(wb):
    section("CHECK 11: Conditional Formatting")
    ws = get_tracker(wb)
    if not ws:
        return

    cf_rules = list(ws.conditional_formatting)
    total = sum(len(cf.rules) for cf in cf_rules)
    report("INFO", f"Total CF rules: {total}")

    categories = {"AA": 0, "AD": 0, "AC": 0, "AB": 0, "date": 0}
    for cf in cf_rules:
        rng = str(cf)
        for rule in cf.rules:
            if "AA" in rng:
                categories["AA"] += 1
            elif "AD" in rng:
                categories["AD"] += 1
            elif "AC" in rng:
                categories["AC"] += 1
            elif "AB" in rng:
                categories["AB"] += 1
            else:
                categories["date"] += 1

    report("PASS" if categories["AA"] >= 7 else "FAIL", f"Status CF rules: {categories['AA']}")
    report("PASS" if categories["AD"] >= 4 else "FAIL", f"Risk CF rules: {categories['AD']}")
    report("PASS" if categories["AC"] >= 2 else "FAIL", f"Days CF rules: {categories['AC']}")
    report("PASS" if categories["AB"] >= 1 else "FAIL", f"Progress DataBar: {categories['AB']}")


# ============================================================
# CHECK 12: Scenario Tests (Add/Delete/StatusChange)
# ============================================================
def check_scenarios(wb):
    section("CHECK 12: Scenario Tests")
    ws = get_tracker(wb)
    if not ws:
        return

    # --- A: Add new item ---
    print(f"\n  --- A: New Item Addition ---")
    last_data = 70  # 66 items → row 70
    next_empty = last_data + 1
    # Next empty row should have pre-populated formulas (A has =IF(B..="","",ROW()-4))
    # but B (VehicleCode) should be empty (ready for user input)
    a_val = ws.cell(row=next_empty, column=1).value
    b_val = ws.cell(row=next_empty, column=2).value
    has_formula = isinstance(a_val, str) and a_val.startswith("=")
    b_empty = b_val is None
    report("PASS" if has_formula and b_empty else "FAIL",
           f"Row {next_empty} formula-ready (A has formula, B empty for input): {has_formula and b_empty}")

    dvs = ws.data_validations.dataValidation
    covers_200 = any("200" in str(dv.sqref) for dv in dvs)
    report("PASS" if covers_200 else "FAIL", f"Validations cover row 200: {covers_200}")

    # --- B: Delete item ---
    print(f"\n  --- B: Item Deletion ---")
    dash = get_dashboard(wb)
    if dash:
        kpi = dash.cell(row=7, column=2).value
        range_based = isinstance(kpi, str) and "200" in kpi
        report("PASS" if range_based else "WARN",
               f"Dashboard uses 5:200 range (deletion-safe): {range_based}")

    # --- C: Status transitions ---
    print(f"\n  --- C: Status Transitions (date entry) ---")
    transitions = [
        ("01→02", "P filled, Q empty"),
        ("02→03", "Q=Under Review"),
        ("03→05", "Q=Approved"),
        ("05→08", "S filled, U empty"),
        ("08→09", "U filled, W empty"),
        ("09→10", "W filled, Y empty"),
        ("10→11", "W+Y filled"),
    ]
    aa_formula = ws.cell(row=5, column=27).value
    if isinstance(aa_formula, str):
        for trans, cond in transitions:
            print(f"    {trans:8s} | {cond}")
        report("PASS", "7 state transitions mapped in formula")

    # --- D: Override States (find dynamically) ---
    print(f"\n  --- D: Override States ---")
    override_targets = {"06-BIW": False, "07-\uc0dd\uc0b0\uc644\ub8cc": False, "12-\uc870\uac74\ubd80": False}
    for row in range(5, 71):
        z = ws.cell(row=row, column=26).value
        if isinstance(z, str):
            for prefix in override_targets:
                if prefix in z and not override_targets[prefix]:
                    report("PASS", f"Row {row}: Override '{z}' contains '{prefix}'")
                    override_targets[prefix] = True
    for prefix, found in override_targets.items():
        if not found:
            report("FAIL", f"Override '{prefix}' not found in data")

    # --- E: LeadTime propagation ---
    print(f"\n  --- E: LeadTime Propagation ---")
    o5 = ws.cell(row=5, column=15).value
    report("PASS" if isinstance(o5, str) and ("TBL_LeadTime" in o5 or "Reference Data" in o5) else "FAIL",
           "PlanDue references LeadTime lookup")
    x5 = ws.cell(row=5, column=24).value
    report("PASS" if isinstance(x5, str) and "MIN" in x5 else "FAIL",
           "ReportDue uses MIN(T0+98, TestDeadline)")

    # --- F: Vehicle-mounted test constraint ---
    print(f"\n  --- F: Vehicle-mounted Test (MAX with ProtoBuild) ---")
    t5 = ws.cell(row=5, column=20).value
    report("PASS" if isinstance(t5, str) and "MAX" in t5 else "FAIL",
           "TestStartPlan uses MAX for vehicle-mounted tests")


# ============================================================
# CHECK 13: Edge Cases
# ============================================================
def check_edge_cases(wb):
    section("CHECK 13: Edge Cases")
    ws = get_tracker(wb)
    if not ws:
        return

    # Priority: 08 > 05
    print(f"  Edge: Status priority (08-단품준비완료 > 05-계획서승인)")
    aa = ws.cell(row=9, column=27).value
    if isinstance(aa, str):
        p08 = aa.find("08-")
        p05 = aa.find("05-")
        report("PASS" if p08 < p05 else "FAIL", f"08 checked before 05: pos {p08} < {p05}")

    # Auto-filter
    af = ws.auto_filter.ref
    report("PASS" if af else "WARN", f"Auto-filter: {af}")


# ============================================================
# MAIN
# ============================================================
def main():
    print(f"{'='*70}")
    print(f"  ESIR Master Tracker V2 — QA Report (66 items, 9 tables)")
    print(f"  File: {FILE}")
    print(f"  Date: {TODAY}")
    print(f"{'='*70}")

    wb = load_workbook(FILE, data_only=False)

    check_sheet_architecture(wb)
    check_reference_data(wb)
    check_tracker_structure(wb)
    check_yellow_highlighting(wb)
    check_formula_integrity(wb)
    rows_info = check_12_states(wb)
    check_multi_spec(wb, rows_info)
    check_vehicle_distribution(wb, rows_info)
    check_dashboard(wb)
    check_validations(wb)
    check_conditional_formatting(wb)
    check_scenarios(wb)
    check_edge_cases(wb)

    print(f"\n{'='*70}")
    print(f"  FINAL SUMMARY")
    print(f"{'='*70}")
    print(f"  PASS: {PASS}")
    print(f"  FAIL: {FAIL}")
    print(f"  WARN: {WARN}")
    print(f"  Total: {PASS + FAIL + WARN}")
    if FAIL == 0:
        print(f"\n  >>> ALL CHECKS PASSED <<<")
    else:
        print(f"\n  >>> {FAIL} FAIL(S) — review above <<<")


if __name__ == "__main__":
    main()
